#!/usr/bin/env python3
"""Verify the yearly export has day-by-day details"""
import sys
sys.path.insert(0, 'E:/claude/zeiterfassung/praxiszeit/backend')

from openpyxl import load_workbook

wb = load_workbook("test_yearly_export.xlsx")

print(f"📊 Yearly Export Verification")
print(f"{'='*60}")
print(f"\n✅ Total sheets: {len(wb.sheetnames)}")
print(f"   Sheets: {', '.join(wb.sheetnames)}\n")

# Check each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    row_count = sheet.max_row
    col_count = sheet.max_column

    print(f"📄 Sheet: {sheet_name}")
    print(f"   Rows: {row_count}, Columns: {col_count}")

    # Show first few headers if available
    if row_count > 0:
        headers = []
        for col in range(1, min(col_count + 1, 11)):
            cell_value = sheet.cell(1, col).value
            if cell_value:
                headers.append(str(cell_value))
        if headers:
            print(f"   Headers: {', '.join(headers[:5])}...")

    # For employee sheets, check if they have day-by-day data
    if sheet_name not in ['Übersicht', 'Abwesenheiten']:
        # Employee sheets should have ~365-400 rows (365 days + headers + summaries + month separators)
        if row_count >= 350:
            print(f"   ✅ Has detailed day-by-day breakdown ({row_count} rows)")

            # Sample a few rows to verify structure
            print(f"\n   Sample rows:")
            for row_num in [2, 3, 50, 100]:
                if row_num <= row_count:
                    date_val = sheet.cell(row_num, 1).value
                    weekday_val = sheet.cell(row_num, 2).value
                    start_val = sheet.cell(row_num, 3).value
                    end_val = sheet.cell(row_num, 4).value
                    print(f"   Row {row_num}: {date_val} | {weekday_val} | {start_val} | {end_val}")
        else:
            print(f"   ⚠️  Warning: Only {row_count} rows (expected ~365-400 for full year)")

    print()

wb.close()

print(f"{'='*60}")
print("✅ Verification complete!")
