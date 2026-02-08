#!/usr/bin/env python3
"""Test script to verify yearly export functionality"""
import requests
import sys

BASE_URL = "http://localhost:8000"

def test_yearly_export():
    # Login as admin
    login_response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={
            "email": "admin@example.com",
            "password": "admin123"
        }
    )

    if login_response.status_code != 200:
        print(f"❌ Login failed: {login_response.status_code}")
        print(login_response.text)
        return False

    token = login_response.json()["access_token"]
    print("✅ Login successful")

    # Test yearly export
    headers = {"Authorization": f"Bearer {token}"}
    export_response = requests.get(
        f"{BASE_URL}/api/admin/reports/export-yearly?year=2026",
        headers=headers
    )

    if export_response.status_code != 200:
        print(f"❌ Yearly export failed: {export_response.status_code}")
        print(export_response.text)
        return False

    # Check if we got Excel file
    content_type = export_response.headers.get("Content-Type", "")
    if "spreadsheet" not in content_type:
        print(f"❌ Wrong content type: {content_type}")
        return False

    # Check file size (should be substantial with 365 days per employee)
    file_size = len(export_response.content)
    print(f"✅ Yearly export successful")
    print(f"   File size: {file_size:,} bytes")

    # Save to file for manual inspection
    with open("test_yearly_report_2026.xlsx", "wb") as f:
        f.write(export_response.content)
    print(f"   Saved to: test_yearly_report_2026.xlsx")

    # Try to read with openpyxl to verify it's valid
    try:
        from openpyxl import load_workbook
        wb = load_workbook("test_yearly_report_2026.xlsx")
        print(f"   Sheet count: {len(wb.sheetnames)}")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")

        # Check first employee sheet for row count (should have ~365 data rows + headers + summaries)
        if len(wb.sheetnames) > 3:  # Overview, Absences, and employee sheets
            first_employee_sheet = wb[wb.sheetnames[3]]
            max_row = first_employee_sheet.max_row
            print(f"   First employee sheet rows: {max_row} (expecting ~365-400 with headers/summaries)")

            if max_row < 350:
                print(f"   ⚠️  Warning: Expected more rows for full year breakdown")

        wb.close()
        print("✅ Excel file is valid and readable")
        return True

    except ImportError:
        print("   ⚠️  openpyxl not available for validation")
        return True
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False

if __name__ == "__main__":
    success = test_yearly_export()
    sys.exit(0 if success else 1)
