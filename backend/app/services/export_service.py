from io import BytesIO
from datetime import date, datetime, timedelta
from calendar import monthrange
from decimal import Decimal
from typing import List
from sqlalchemy import or_
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from app.models import User, TimeEntry, Absence, PublicHoliday, AbsenceType, AbsenceReason
from app.services import calculation_service, special_days_service
from app.services.arbzg_utils import is_night_work
from app.services.date_filters import date_in_year, date_in_month, date_in_year_up_to_month
from app.config import settings


# M-SEC1: spreadsheet formula / CSV injection guard. A cell whose text starts
# with one of these characters is prefixed with an apostrophe so Excel /
# LibreOffice treat it as literal text and never evaluate it as a formula.
# Employee free-text (note, sunday_exception_reason) and user names flow into
# the §16 export which an admin then opens — without this, a note like
# ``=HYPERLINK(...)`` or a DDE payload would execute on their workstation.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _group_by_date(rows) -> dict:
    """#219: group TimeEntry/Absence rows into {date: [rows…]} (Reihenfolge erhalten).
    Pro Tag eine LISTE — ein Tag kann mehrere Einträge/Abwesenheiten tragen
    (I-1: z. B. halber Tag Urlaub + halber Tag Sonstiges)."""
    by_date: dict = {}
    for r in rows:
        by_date.setdefault(r.date, []).append(r)
    return by_date


def neutralize_spreadsheet_formula(value):
    """Return ``value`` with a leading apostrophe if it could be parsed as a
    spreadsheet formula. Non-strings and safe strings are returned unchanged."""
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def export_users(db, tenant_id, period_start: date, period_end: date) -> List[User]:
    """Mitarbeiter, die in einen §16-Beleg für ``[period_start, period_end]`` gehören.

    Release-Review 1.16.0: die Exporte filterten hart auf ``is_active == True``.
    Das Handbuch weist Admins aber ausdrücklich an, ausgeschiedene Mitarbeiter auf
    „Inaktiv" zu setzen statt zu löschen (§16 ArbZG: 2 Jahre Aufbewahrung) — genau
    diese Personen fielen danach ersatzlos aus jedem Monats- und Jahresbericht,
    auch für Zeiträume, in denen sie noch gearbeitet hatten. Bei einer Prüfung
    fehlten damit die Nachweise, ohne dass irgendwo ein Hinweis erschien.

    Deshalb: aktive Mitarbeiter wie bisher, PLUS inaktive, die im Zeitraum
    tatsächlich Daten haben (Zeiteintrag oder Abwesenheit). Ausgeblendete
    (``is_hidden``) bleiben ausgeblendet — das ist eine bewusste Sichtbarkeits-
    entscheidung des Admins, keine Ausscheidens-Markierung. Ein Zeitraum ohne
    Daten des Ausgeschiedenen erzeugt weiterhin kein leeres Blatt.
    """
    base = db.query(User).filter(User.is_hidden == False)  # noqa: E712
    if tenant_id is not None:
        base = base.filter(User.tenant_id == tenant_id)

    te = db.query(TimeEntry.user_id).filter(
        TimeEntry.date >= period_start, TimeEntry.date <= period_end
    )
    ab = db.query(Absence.user_id).filter(
        Absence.date >= period_start, Absence.date <= period_end
    )
    if tenant_id is not None:
        te = te.filter(TimeEntry.tenant_id == tenant_id)
        ab = ab.filter(Absence.tenant_id == tenant_id)

    users = base.filter(User.is_active == True).all()  # noqa: E712
    users += base.filter(
        User.is_active == False,  # noqa: E712
        or_(User.id.in_(te.scalar_subquery()), User.id.in_(ab.scalar_subquery())),
    ).all()
    return sorted(users, key=lambda u: ((u.last_name or "").lower(), (u.first_name or "").lower()))


def absence_day_target(db, user, d, day_absences, holiday_dates, special_cfg, wh_changes=None):
    """Release-Review 1.16.0: Tages-Soll an einem Tag MIT Abwesenheit.

    Alle Datei-Exporte setzten hier pauschal ``Decimal('0.00')`` — „irgendeine
    Abwesenheit ⇒ kein Soll". Das ist an drei Stellen falsch und ließ die
    Summenzeilen desselben §16-Belegs seinem eigenen „Überstunden kumuliert"
    widersprechen:

    * **Halbtag** (``half_day=True``): nur 0,5 × Tagessoll fällt weg. Ein halber
      Urlaubstag plus vier gestempelte Stunden ergab im Export Soll 0 / Ist 4 →
      +4 h Überstunden statt 0.
    * **SICK/TRAINING**: nicht soll-reduzierend — das Soll bleibt stehen (die
      Gutschrift läuft über ``credited_absences``).
    * **OVERTIME**: Soll bleibt, Ist = 0 (docs/BERECHNUNGEN.md §6). Der Export
      zeigte 0/0 und verschluckte damit den Konto-Abbau.

    Delegiert an ``calculation_service._day_soll_contribution`` — dieselbe Quelle,
    die ``get_monthly_target`` und ``get_overtime_account`` nutzen. Damit können
    Bildschirm und Datei nicht mehr auseinanderlaufen. Wochenende, Beschäftigungs-
    fenster und Stichtag bleiben Sache des Aufrufers (wie beim Helper selbst);
    Feiertage behandeln die Exporte in einem eigenen Zweig davor.
    """
    # ``_soll_reducing_absence_half_map`` filtert NICHT selbst nach Typ — im
    # calculation_service übernimmt das die Query (``type.notin_([TRAINING, SICK,
    # OVERTIME])``). Hier kommen die Absencen aus dem Export-Grouping, also muss der
    # Filter an dieser Stelle stehen: TRAINING/SICK/OVERTIME lassen das Soll stehen.
    soll_reducing = [
        a for a in day_absences
        if a.type not in (
            calculation_service.AbsenceType.TRAINING,
            calculation_service.AbsenceType.SICK,
            calculation_service.AbsenceType.OVERTIME,
        )
    ]
    half_map = calculation_service._soll_reducing_absence_half_map(soll_reducing)
    return calculation_service._day_soll_contribution(
        db, user, d,
        holiday_dates=holiday_dates,
        absence_half_map=half_map,
        wh_changes=wh_changes,
        special_cfg=special_cfg,
    )


def _de_hours(value) -> str:
    """'30,0' — deutsche Dezimaldarstellung fuer eine Stundenzahl.

    NICHT mehr fuer die #415-Historie verwendet (dort steht :func:`_de_hours_exact`),
    aber weiterhin fuer andere Stundenausgaben. Achtung beim Wiederverwenden: eine
    Nachkommastelle rundet, und Python rundet ``%.1f`` half-even, waehrend
    JavaScript ``toFixed(1)`` kaufmaennisch rundet — jeder Wert auf ``.25`` faellt
    zwischen Backend und Frontend auseinander (38,25 → „38,2" hier, „38,3" dort).
    """
    return f"{float(value):.1f}".replace(".", ",")


def _de_hours_exact(value) -> str:
    """'8,0' / '8,5' / '8,25' — Stundenwert in DE-Schreibweise, verlustfrei.

    DIE Zahlformatierung der #415/#431-Vertragshistorie, in BEIDEN Modi.

    Die Tagesstunden und die daraus gebildete Wochensumme sind ``Numeric(4,2)``
    (#431 hat den Spaltentyp genau dafuer verbreitert, und
    ``WorkingHoursChangeCreate`` begrenzt nur ``0..60`` ohne Nachkommastellen-
    Limit): 8,25 h ist ein realer Vertragswert. Mit einer Nachkommastelle
    (:func:`_de_hours`) schriebe die Datei „8,2", der Frontend-Zwilling
    ``toFixed(1)`` dagegen „8,3" — Bildschirm und Datei wuerden sich
    widersprechen. Bei zwei Nachkommastellen rundet fuer diesen Spaltentyp gar
    nichts mehr, die beiden Seiten koennen also nicht auseinanderlaufen.
    Eine belanglose zweite Null faellt weg ('8,00' → '8,0').

    Byte-Identitaet zu #415: fuer JEDEN Wert mit hoechstens einer Nachkommastelle
    — also jeden, den ``Numeric(4,1)`` vor diesem Branch ueberhaupt speichern
    konnte — liefert diese Funktion exakt dieselbe Zeichenkette wie
    :func:`_de_hours`. Der Wechsel aendert nur die Werte, die vorher falsch
    (bzw. zwischen den beiden Seiten uneinig) waren. Nachgewiesen in
    ``test_415_working_hours_history_reports.py::TestDeHoursExactIsAByteIdenticalUpgrade``.
    """
    text = f"{float(value):.2f}"
    if text.endswith("0"):
        text = text[:-1]
    return text.replace(".", ",")


def _de_hours_compact(value) -> str:
    """Wie :func:`_de_hours_exact`, ohne die belanglose Nachkommastelle:
    '8' statt '8,0' ('8,25' bleibt '8,25')."""
    text = _de_hours_exact(value)
    return text[:-2] if text.endswith(",0") else text


def format_hours_de(value) -> str:
    """Oeffentlicher Name von :func:`_de_hours_exact` — dieselbe Regel, nicht
    eine zweite.

    Gibt es, damit Flaechen AUSSERHALB der Dateiexporte (Task 15: die
    Klartext-Stunden im Aenderungsprotokoll) dieselbe verlustfreie deutsche
    Schreibweise benutzen, ohne einen privaten Namen quer durch die Module zu
    importieren oder — schlimmer — ein zweites ``f"{x:.2f}".replace(...)``
    aufzumachen, das dann irgendwann anders rundet.
    """
    return _de_hours_exact(value)


# Deutsche Klartext-Labels der eingebauten Abwesenheitstypen. EINE Definition
# fuer dieses Modul (vorher stand derselbe Dict dreimal woertlich in den
# Detail-Grids) — plus Task 15: das Aenderungsprotokoll benennt die
# nachgezogene Abwesenheit im selben Vokabular wie der §16-Export.
# ``ods_export_service.ABSENCE_LABELS`` ist die eigenstaendige, inhaltsgleiche
# Konstante des ODS-Zwillings.
ABSENCE_TYPE_LABELS_DE = {
    "vacation": "Urlaub",
    "sick": "Krank",
    "training": "Fortbildung",
    "overtime": "Überstundenausgleich",
    "other": "Sonstiges",
    "paid_leave": "Bez. Freistellung",
}


_WEEKDAY_LABELS = ("Mo", "Di", "Mi", "Do", "Fr")


def format_day_plan(day_hours, compact: bool = False) -> str:
    """#431: der Tagesplan als Klartext — ``'Mo 8,0 / Di 5,0 / Mi 4,0'``.

    Wochentage ohne Stunden stehen NICHT drin, weder ``None`` noch ``0``: ein Tag
    ohne Soll ist kein Arbeitstag (er zaehlt auch nicht in die Arbeitstage), und
    ein angehaengtes „Do 0,0 / Fr 0,0" blaeht jede Kopfzeile auf, ohne etwas zu
    sagen. Leerstring, wenn kein einziger Tag Stunden traegt (per Schema
    unmoeglich, aber Bestandszeilen sind nicht garantiert).

    ``compact`` ist die Kurzform fuer die PDF-Meta (Inline-Absatz, Schriftgroesse
    8, Querformat).
    """
    fmt = _de_hours_compact if compact else _de_hours_exact
    return " / ".join(
        f"{label} {fmt(hours)}"
        for label, hours in zip(_WEEKDAY_LABELS, day_hours or ())
        if hours is not None and float(hours) > 0
    )


def format_weekly_hours_history(segments, compact: bool = False) -> str:
    """#415: die Vertragsaenderungen eines Zeitraums als Klartext.

    ``segments`` ist die Ausgabe von
    :func:`calculation_service.weekly_hours_segments`. Das ERSTE Segment ist der
    zum Zeitraumsbeginn gueltige Zustand und steht bereits als Zahl in der
    Kopfzeile/Spalte — hier interessieren nur die Aenderungen danach.

    Leerstring, wenn sich im Zeitraum nichts geaendert hat: die Aufrufer
    schreiben die Zelle dann gar nicht erst, sodass unveraenderte Berichte
    exakt so aussehen wie vorher.

    Zwei Formulierungen, je nach Modus des Segments:

    * gleichmaessig — ``ab 15.03.2026: 30,0 Std/Woche`` (woertlich wie seit #415)
    * Tagesplan (#431) — ``ab 01.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche``

    Aendern sich im gleichmaessigen Modus die ARBEITSTAGE, werden sie zusaetzlich
    genannt (``ab 16.03.2026: 40,0 Std/Woche auf 4 Arbeitstage``) — siehe
    :func:`_work_days_suffix`.

    Der Frontend-Zwilling ``utils/formatters.ts::formatWeeklyHoursChanges`` muss
    WORTGLEICH bleiben — Bildschirm und Datei duerfen nicht verschiedene Saetze
    sagen.
    """
    if len(segments) < 2:
        return ""
    return "; ".join(
        # Der Vorgaenger ist Teil der Aussage: „auf 4 Arbeitstage" steht nur da,
        # wo sich die Arbeitstage tatsaechlich geaendert haben.
        _format_segment_change(seg, previous=segments[i], compact=compact)
        for i, seg in enumerate(segments[1:])
    )


def _attach_overflow_comment(cell, text: str) -> None:
    """Fund D (Abschluss-Review #431): ein Excel-Kommentar statt eine breitere
    Spalte oder Zeilenumbruch.

    Die #415-Aenderungszeile (``format_weekly_hours_history``) steht in einer
    Metadaten-Zelle mit einer NICHT-leeren Nachbarzelle rechts (``G1`` traegt
    „Monat:") — Excel/LibreOffice ueberlaufen langen Zellinhalt nur in LEERE
    Nachbarzellen, alles darueber hinaus wird beim Anzeigen abgeschnitten. Bei
    mehreren Aenderungen im Zeitraum (``"; "``-verkettet) oder einer
    Tagesplan-Zeile („ab 15.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche",
    53 Zeichen) ist der Satz in der zwoelf Zeichen breiten Spalte F praktisch
    unsichtbar, OHNE dass ein Abschneide-Hinweis (z. B. „…") das anzeigt — die
    Zelle wirkt vollstaendig, ist es aber nicht.

    Eine breitere Spalte F wuerde den Wert JEDER Zeile der „Netto (Std)"-Spalte
    darunter mitziehen (dieselbe Spalte traegt ab Zeile 4 die taegliche
    Netto-Stunden-Zahl) — ein 50+ Zeichen breiter Spaltenkopf fuer eine
    zweistellige Zahl. Ein Zeilenumbruch (``wrap_text``) haette dieselbe
    Nebenwirkung ueber die Zeilenhoehe von Zeile 1 (dort stehen „Mitarbeiter:"
    und der Name, die keinen Umbruch brauchen). Ein Kommentar aendert weder
    Spaltenbreite noch Zeilenhoehe noch die Position einer einzigen bestehenden
    Spalte — Kundenauswertungen, die auf festen Spalten dieser Dateien laufen,
    sind unberuehrt. Excel/LibreOffice markieren die Zelle mit einem kleinen
    Eck-Indikator; der volle Satz steht beim Hovern/Oeffnen des Kommentars,
    unabhaengig von der Spaltenbreite.
    """
    cell.comment = Comment(text, "PraxisZeit")


def _work_days_suffix(segment, previous, compact: bool = False) -> str:
    """„ auf 4 Arbeitstage" — aber NUR, wenn dieses Segment die Arbeitstage
    gegenueber ``previous`` tatsaechlich aendert (Leerstring sonst).

    Grund (Abschluss-Review #431, Fund 1): ``weekly_hours_segments`` splittet auf
    dem VOLLSTAENDIGEN Snapshot, die #415-Formulierung nannte aber nur die
    Wochenstunden. „40 h auf 5 Tage" → „40 h auf 4 Tage" erzeugte damit eine
    Aenderungszeile, die zeichengleich zur Kopfzeile war (zweimal „40,0"),
    waehrend ``get_daily_target_for_date`` (= Wochenstunden ÷ Arbeitstage) das
    Tagessoll derselben Tageszeilen von 8,00 h auf 10,00 h hob — genau der
    selbstwidersprüchliche §16-Beleg, gegen den #415 angetreten ist. Und seit
    #431 ist der Dialog der EINZIGE Weg, die Arbeitstage zu aendern.

    Die Bedingung „nur bei tatsaechlicher Aenderung" haelt den eingefrorenen
    #415-Wortlaut fuer JEDEN Fall byte-identisch, der vor diesem Branch moeglich
    war: die Arbeitstage waren nicht historisiert, alle Segmente eines
    Mitarbeitenden trugen denselben Wert — der Zusatz kann dort nie erscheinen
    (Test ``TestWorkDaysChangeIsNamed::test_unchanged_work_days_keep_the_frozen_415_wording``).

    ``compact`` (PDF-Meta, Schriftgroesse 8 im Querformat) kuerzt „Arbeitstage"
    zu „Tage" — wie die Kurzform der Stundenwerte auch.
    """
    if not calculation_service.work_days_changed(previous, segment):
        return ""
    days = segment.work_days_per_week
    if days is None:
        return ""
    days = int(days)
    if compact:
        return f" auf {days} {'Tag' if days == 1 else 'Tage'}"
    return f" auf {days} {'Arbeitstag' if days == 1 else 'Arbeitstage'}"


def _format_segment_change(segment, previous=None, compact: bool = False) -> str:
    """Ein Segment als „ab <Datum>: <Zustand>". ``previous`` ist das unmittelbar
    davor gueltige Segment (``None`` = keins) und entscheidet allein darueber, ob
    die Arbeitstage genannt werden."""
    prefix = f"ab {segment.start.strftime('%d.%m.%Y')}: "
    suffix = ""
    if segment.use_daily_schedule:
        plan = format_day_plan(segment.day_hours, compact=compact)
        if plan:
            total = _de_hours_compact if compact else _de_hours_exact
            return f"{prefix}{plan} = {total(segment.weekly_hours)} h/Woche"
        # Kein einziger Tageswert → auf die gleichmaessige Formulierung
        # zurueckfallen, statt einen leeren Satz zu schreiben. Ohne Zusatz: die
        # Arbeitstage einer (per Schema unmoeglichen) Zeile ohne einen einzigen
        # Tageswert sagen nichts, und der Satz bleibt so woertlich der bisherige.
    else:
        suffix = _work_days_suffix(segment, previous, compact=compact)
    # Auch der gleichmaessige Zweig formatiert verlustfrei (`_de_hours_exact`):
    # fuer jeden vor diesem Branch speicherbaren Wert ist das zeichengleich zu
    # #415, und nur so sagen Datei und Bildschirm bei 38,25 h dasselbe.
    # `compact` wirkt auf die STUNDENZAHL hier NICHT — die #415-Formulierung
    # bleibt in der PDF-Meta woertlich (`_de_hours_exact`, nicht
    # `_de_hours_compact`).
    return f"{prefix}{_de_hours_exact(segment.weekly_hours)} Std/Woche{suffix}"


def escape_pdf_text(value):
    """Escape user-controlled text before it is placed into a reportlab
    ``Paragraph``.

    reportlab parses an intra-paragraph XML/HTML-like markup mini-language
    (``<b>``, ``<font>``, ``<a href>``, ``<img src>``, entities). Unescaped
    free text (notes, names, billing fields) is therefore an injection sink:
    content/link spoofing into the official §16 PDF, a server-side fetch of an
    attacker-chosen ``<img src>`` URL (SSRF), and — most reliably — a single
    unbalanced tag raises during ``doc.build`` and 500s the whole report.

    This is the PDF-side counterpart to :func:`neutralize_spreadsheet_formula`
    and must be applied at every ``Paragraph`` that carries user input.
    ``&`` ``<`` ``>`` are the only characters reportlab treats as markup."""
    if value is None:
        return ""
    from xml.sax.saxutils import escape as _xml_escape
    return _xml_escape(str(value))


def generate_monthly_report(db: Session, year: int, month: int, include_health_data: bool = False, tenant_id=None) -> BytesIO:
    """
    Generate Excel report for all employees for a given month.
    Creates one sheet per employee.

    Args:
        db: Database session
        year: Year
        month: Month (1-12)
        include_health_data: If False (default), sick absences are shown as "Abwesenheit" (Art. 9 DSGVO protection)
        tenant_id: F-026 belt-and-suspenders explicit tenant filter (on top of RLS)

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Get all active, non-hidden employees (F-026: explicit tenant filter on top of RLS)
    users = export_users(db, tenant_id, date(year, month, 1),
                         date(year, month, monthrange(year, month)[1]))

    for user in users:
        _create_employee_sheet(wb, db, user, year, month, include_health_data)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _load_reason_names(db: Session, tenant_id) -> dict:
    """#312: {str(reason_id): name} for the tenant — custom-reason export labels."""
    return {
        str(r.id): r.name for r in db.query(AbsenceReason.id, AbsenceReason.name)
        .filter(AbsenceReason.tenant_id == tenant_id).all()
    }


def _absence_export_label(absence, type_map: dict, reason_names: dict, include_health_data: bool):
    """#312: (label, show_note) for an absence in §16 exports.

    SICK *and* custom-reason absences are masked (label "Abwesenheit", note
    hidden) unless health data is explicitly included — a custom reason can be
    health-sensitive (e.g. "Reha"). Otherwise a custom reason shows its own
    name; built-in types use the German type label.
    """
    if (absence.type.value == "sick" and not include_health_data) or \
       (absence.reason_id is not None and not include_health_data):
        return "Abwesenheit", False
    if absence.reason_id is not None:
        return reason_names.get(str(absence.reason_id)) or type_map.get(absence.type.value, absence.type.value), True
    return type_map.get(absence.type.value, absence.type.value), True


def _create_employee_sheet(wb: Workbook, db: Session, user: User, year: int, month: int, include_health_data: bool = False):
    """
    Create a worksheet for a single employee.

    Columns:
    - Datum
    - Wochentag
    - Von
    - Bis
    - Pause (Min)
    - Netto (Std)
    - Soll (Std)
    - Differenz
    - Abwesenheit
    - Bemerkung
    """
    sheet = wb.create_sheet(title=f"{user.last_name} {user.first_name}"[:31])  # Excel sheet name max 31 chars

    reason_names = _load_reason_names(db, user.tenant_id)  # #312

    # Row 1–2: ArbZG-relevante Mitarbeiter-Metadaten (§16 ArbZG Aufzeichnungspflicht)
    sheet.cell(row=1, column=1).value = "Mitarbeiter:"
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).value = neutralize_spreadsheet_formula(f"{user.first_name} {user.last_name}")
    # #415: der zum MONATSBEGINN gueltige Vertragswert — nicht der aktuelle.
    # Sonst widerspricht die Kopfzeile den historisch gerechneten Tageszeilen.
    _last_day = monthrange(year, month)[1]
    _wh_segments = calculation_service.weekly_hours_segments(
        db, user, date(year, month, 1), date(year, month, _last_day)
    )
    sheet.cell(row=1, column=4).value = "Wochenstunden:"
    sheet.cell(row=1, column=4).font = Font(bold=True)
    sheet.cell(row=1, column=5).value = float(_wh_segments[0].weekly_hours) if _wh_segments else float(user.weekly_hours)
    _wh_history = format_weekly_hours_history(_wh_segments)
    if _wh_history:
        sheet.cell(row=1, column=6).value = _wh_history
        # Fund D: F1 hat mit G1 ("Monat:") eine nicht-leere Nachbarzelle — der
        # Satz laeuft nicht ueber, siehe _attach_overflow_comment.
        _attach_overflow_comment(sheet.cell(row=1, column=6), _wh_history)
    sheet.cell(row=1, column=7).value = "Monat:"
    sheet.cell(row=1, column=7).font = Font(bold=True)
    sheet.cell(row=1, column=8).value = f"{month:02d}/{year}"
    sheet.cell(row=2, column=1).value = "§18 ArbZG-befreit:"
    sheet.cell(row=2, column=1).font = Font(bold=True)
    sheet.cell(row=2, column=2).value = "Ja" if user.exempt_from_arbzg else "Nein"
    sheet.cell(row=2, column=4).value = "Nachtarbeitnehmer (§6 Abs. 2 ArbZG):"
    sheet.cell(row=2, column=4).font = Font(bold=True)
    sheet.cell(row=2, column=5).value = "Ja" if user.is_night_worker else "Nein"
    # Row 3: blank separator

    # Row 4: Column headers
    headers = ["Datum", "Wochentag", "Von", "Bis", "Pause (Min)", "Netto (Std)", "Soll (Std)", "Differenz", "Abwesenheit", "Bemerkung"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Get data
    _, last_day = monthrange(year, month)

    # Get all time entries for the month (list-based: multiple entries per day)
    # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
    time_entries = db.query(TimeEntry).filter(
        TimeEntry.user_id == user.id,
        TimeEntry.tenant_id == user.tenant_id,
        date_in_month(TimeEntry.date, year, month)
    ).order_by(TimeEntry.start_time).all()
    entries_by_date = _group_by_date(time_entries)  # #219: shared

    # Get all absences for the month.
    # I-1: ein Tag kann MEHRERE Absences tragen (Unique-Constraint ist je
    # (date, type) — z. B. ein halber Tag Urlaub + ein halber Tag Sonstiges).
    # Daher pro Tag eine LISTE statt einer einzelnen Absence, sonst verschluckt
    # die Anzeige die zweite stillschweigend.
    # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
    absences = db.query(Absence).filter(
        Absence.user_id == user.id,
        Absence.tenant_id == user.tenant_id,
        date_in_month(Absence.date, year, month)
    ).all()
    absences_by_date = _group_by_date(absences)  # #219: shared

    # Get public holidays (tenant-scoped: each tenant may run a different
    # state-holiday set, so a global query would leak or miss holidays).
    holidays = db.query(PublicHoliday).filter(
        PublicHoliday.tenant_id == user.tenant_id,
        date_in_month(PublicHoliday.date, year, month),
    ).all()
    holidays_by_date = {holiday.date: holiday for holiday in holidays}

    # #146: configurable 24./31.12. handling — load once, apply per working day
    # so the exported Soll column matches get_monthly_target() (half_day → ×0.5,
    # free → ×0). N-1: previously the export showed the full daily target on a
    # configured half/free day, diverging from the calculated monthly Soll.
    special_day_config = special_days_service.get_special_day_config(db, user.tenant_id, year)

    # German weekday names
    weekday_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

    row = 5  # Data starts after 3-row header + blank
    total_net = Decimal('0.00')
    total_target = Decimal('0.00')
    night_work_count = 0

    # Iterate through all days of the month
    for day in range(1, last_day + 1):
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        weekday_name = weekday_names[weekday]
        is_sunday = weekday == 6

        # Check if it's a weekend, holiday, or absence
        is_weekend = weekday >= 5
        is_holiday = current_date in holidays_by_date
        day_absences = absences_by_date.get(current_date, [])  # I-1: alle Absences des Tages

        # Date column
        sheet.cell(row=row, column=1).value = current_date
        sheet.cell(row=row, column=1).number_format = 'DD.MM.YYYY'

        # Weekday column
        sheet.cell(row=row, column=2).value = weekday_name

        # Get time entries if exist (may be multiple per day)
        day_entries = entries_by_date.get(current_date, [])

        # Night work check (§6 / §2 Abs. 4 ArbZG)
        is_night_wrk = any(
            e.end_time is not None and is_night_work(e.start_time, e.end_time)
            for e in day_entries
        )
        if is_night_wrk:
            night_work_count += 1

        if day_entries:
            first_start = day_entries[0].start_time
            # spätestes Ende des Tages, nicht der zuletzt START-ende Eintrag — bei
            # überlappenden Mehrfach-Einträgen (A 08–17, B 12–14) war "Bis" sonst
            # 14:00 statt 17:00 (Review 2026-06-23, §16-Korrektheit).
            last_end = max((e.end_time for e in day_entries if e.end_time), default=None)
            total_break = sum(e.break_minutes or 0 for e in day_entries)
            total_day_net = sum(e.net_hours for e in day_entries)
            sheet.cell(row=row, column=3).value = first_start.strftime('%H:%M')
            sheet.cell(row=row, column=4).value = last_end.strftime('%H:%M') if last_end else 'offen'
            sheet.cell(row=row, column=5).value = total_break
            sheet.cell(row=row, column=6).value = float(total_day_net)
            sheet.cell(row=row, column=6).number_format = '0.00'
            # Bemerkung (col 10): §10-Ausnahmegrund hat Vorrang, dann entry.note
            bemerkung_parts = []
            for e in day_entries:
                if e.sunday_exception_reason and (is_sunday or is_holiday):
                    bemerkung_parts.append(f"§10-Ausnahmegrund: {e.sunday_exception_reason}")
                if e.note:
                    bemerkung_parts.append(e.note)
            if bemerkung_parts:
                sheet.cell(row=row, column=10).value = neutralize_spreadsheet_formula(" | ".join(bemerkung_parts))
            net = total_day_net
            total_net += net
        else:
            net = Decimal('0.00')
            sheet.cell(row=row, column=6).value = 0.00
            sheet.cell(row=row, column=6).number_format = '0.00'

        # Per-day target using the historical contract snapshot (#431)
        schedule = calculation_service.get_schedule_for_date(db, user, current_date)
        daily_target = calculation_service.get_daily_target_for_date(user, current_date, schedule)
        # #146: apply the special-day factor (only the working-day branch below
        # consumes daily_target; the weekend/holiday/absence branches hardcode 0).
        _sd_factor = special_days_service.special_day_target_factor(current_date, special_day_config)
        if _sd_factor is not None:
            daily_target = daily_target * _sd_factor

        # Target hours + Abwesenheit (col 9) – korrekte Labels für §9/§10/§6
        if is_weekend:
            target = Decimal('0.00')
            if is_sunday and day_entries:
                abw = "Sonntagsarbeit (§9/§10 ArbZG)"
            elif is_sunday:
                abw = "Sonntag"
            else:
                abw = "Samstag"
            if is_night_wrk:
                abw += " | Nachtarbeit (§6 ArbZG)"
            sheet.cell(row=row, column=9).value = abw
            for col in range(1, 11):
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        elif is_holiday:
            target = Decimal('0.00')
            holiday = holidays_by_date[current_date]
            if day_entries:
                abw = f"Feiertagsarbeit: {holiday.name} (§9/§10 ArbZG)"
            else:
                abw = f"Feiertag: {holiday.name}"
            if is_night_wrk:
                abw += " | Nachtarbeit (§6 ArbZG)"
            sheet.cell(row=row, column=9).value = abw
            # col 10 (Bemerkung) bereits oben gesetzt – NICHT mit holiday.name überschreiben
            for col in range(1, 11):
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif day_absences:
            # Release-Review 1.16.0: zentrale Soll-Quelle statt pauschal 0.
            target = absence_day_target(db, user, current_date, day_absences, set(holidays_by_date), special_day_config)
            absence_type_map = ABSENCE_TYPE_LABELS_DE
            # I-1: ALLE Absences des Tages anzeigen (Label in Spalte 9 verbinden,
            # Notizen in Spalte 10). DSGVO F-003: Krank ohne Health-Flag maskieren
            # (Label "Abwesenheit", Notiz unterdrückt — kann Diagnose enthalten).
            abw_parts = []
            note_parts = []
            for absence in day_absences:
                type_name, show_note = _absence_export_label(
                    absence, absence_type_map, reason_names, include_health_data)
                if show_note and absence.note:
                    note_parts.append(absence.note)
                abw_parts.append(f"{type_name} ({float(absence.hours)}h)")
            sheet.cell(row=row, column=9).value = neutralize_spreadsheet_formula(" | ".join(abw_parts))  # custom-reason label = user text → neutralisieren
            if note_parts:
                sheet.cell(row=row, column=10).value = neutralize_spreadsheet_formula(" | ".join(note_parts))
        else:
            # Regulärer Arbeitstag
            target = daily_target
            if is_night_wrk:
                sheet.cell(row=row, column=9).value = "Nachtarbeit (§6 ArbZG)"

        sheet.cell(row=row, column=7).value = float(target)
        sheet.cell(row=row, column=7).number_format = '0.00'
        total_target += target

        # Difference
        diff = net - target
        sheet.cell(row=row, column=8).value = float(diff)
        sheet.cell(row=row, column=8).number_format = '0.00'

        if diff > 0:
            sheet.cell(row=row, column=8).font = Font(color="006400")
        elif diff < 0:
            sheet.cell(row=row, column=8).font = Font(color="8B0000")

        row += 1

    # Summary section
    row += 1
    sheet.cell(row=row, column=1).value = "Zusammenfassung"
    sheet.cell(row=row, column=1).font = Font(bold=True, size=12)

    # #377 Baustein 2b (Finding 1, Whole-Branch-Review): für Fix-Modus-MA
    # (use_fixed_monthly_target) MUSS die Monats-Summary mit dem modus-
    # bewussten get_monthly_target/get_monthly_actual übereinstimmen — sonst
    # widerspricht sich das §16-Dokument selbst gegen "Überstunden kumuliert"
    # weiter unten (get_overtime_account, bereits modus-bewusst). Die Per-Tag-
    # Detailzeilen oben bleiben unverändert (informativ: geplante Anwesenheit/
    # reale Erfassung, keine Tages-Soll-Zerlegung des festen Monats-Solls).
    # Nicht-Modus-MA bleiben exakt auf der alten Per-Tag-Summe (byte-identisch).
    if getattr(user, "use_fixed_monthly_target", False) and getattr(user, "agreed_monthly_hours", None):
        summary_target = calculation_service.get_monthly_target(db, user, year, month)
        summary_actual = calculation_service.get_monthly_actual(db, user, year, month)
    else:
        summary_target = total_target
        summary_actual = total_net

    row += 1
    sheet.cell(row=row, column=1).value = "Soll-Stunden Monat:"
    sheet.cell(row=row, column=2).value = float(summary_target)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    sheet.cell(row=row, column=1).value = "Ist-Stunden Monat:"
    sheet.cell(row=row, column=2).value = float(summary_actual)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    monthly_balance = summary_actual - summary_target
    sheet.cell(row=row, column=1).value = "Saldo Monat:"
    sheet.cell(row=row, column=2).value = float(monthly_balance)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)
    if monthly_balance > 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="006400")
    elif monthly_balance < 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="8B0000")

    row += 1
    overtime_account = calculation_service.get_overtime_account(db, user, year, month)
    sheet.cell(row=row, column=1).value = "Überstunden kumuliert:"
    sheet.cell(row=row, column=2).value = float(overtime_account)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)
    if overtime_account > 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="006400")
    elif overtime_account < 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="8B0000")

    row += 1
    vacation_account = calculation_service.get_vacation_account(db, user, year)
    sheet.cell(row=row, column=1).value = "Urlaub genommen (Std):"
    sheet.cell(row=row, column=2).value = float(vacation_account['used_hours'])
    sheet.cell(row=row, column=2).number_format = '0.00'

    row += 1
    sheet.cell(row=row, column=1).value = "Urlaub Rest (Std):"
    sheet.cell(row=row, column=2).value = float(vacation_account['remaining_hours'])
    sheet.cell(row=row, column=2).number_format = '0.00'

    row += 1
    sheet.cell(row=row, column=1).value = "Nachtarbeitstage (§6 ArbZG):"
    sheet.cell(row=row, column=2).value = night_work_count
    sheet.cell(row=row, column=1).font = Font(bold=True)

    # Adjust column widths
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 28
    sheet.column_dimensions['J'].width = 35


def generate_yearly_report(db: Session, year: int, include_health_data: bool = False, tenant_id=None) -> BytesIO:
    """
    Generate Excel report for all employees for a given year.
    Creates:
    - Overview sheet with summary for all employees
    - One sheet per employee with monthly breakdown
    - Absences overview

    Args:
        db: Database session
        year: Year
        tenant_id: F-026 belt-and-suspenders explicit tenant filter (on top of RLS)

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Get all active, non-hidden employees (F-026: explicit tenant filter on top of RLS)
    users = export_users(db, tenant_id, date(year, 1, 1), date(year, 12, 31))

    # Create overview sheet
    _create_yearly_overview_sheet(wb, db, users, year, include_health_data)

    # Create absences overview sheet
    _create_absences_overview_sheet(wb, db, users, year, include_health_data)

    # Create employee detail sheets
    for user in users:
        _create_employee_yearly_sheet(wb, db, user, year, include_health_data)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _create_yearly_overview_sheet(wb: Workbook, db: Session, users: List[User], year: int, include_health_data: bool = False):
    """Create overview sheet with all employees."""
    sheet = wb.create_sheet(title="Jahresübersicht", index=0)

    # Title
    sheet.cell(row=1, column=1).value = f"Jahresübersicht {year}"
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.merge_cells('A1:K1')  # #415: inkl. der angehängten Spalte "Stundenänderungen"

    # Headers
    headers = [
        "Name", "Wochenstunden", "Soll (Jahr)", "Ist (Jahr)",
        "Saldo (Jahr)", "Überstunden kum.",
        "Urlaub Budget", "Urlaub genommen", "Urlaub Rest",
        "Krankheitstage",
        # #415: ANGEHAENGT, nicht eingeschoben — bestehende Spaltenpositionen
        # (und damit Kundenauswertungen auf dieser Datei) bleiben unveraendert.
        "Stundenänderungen",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    row = 4
    for user in users:
        # Calculate yearly totals
        yearly_target = Decimal('0')
        yearly_actual = Decimal('0')

        for month in range(1, 13):
            target = calculation_service.get_monthly_target(db, user, year, month)
            actual = calculation_service.get_monthly_actual(db, user, year, month)
            yearly_target += target
            yearly_actual += actual

        yearly_balance = yearly_actual - yearly_target
        overtime = calculation_service.get_overtime_account(db, user, year, 12)

        # Vacation account
        vacation_account = calculation_service.get_vacation_account(db, user, year)

        # Krankheitstage TAGEBASIERT (GLOSSAR-Tagesprinzip) — identisch zum
        # Abwesenheiten-Sheet; NICHT die frühere naive Σh÷Tagessoll-Methode
        # (falsch bei Tagesplan/Halbtagen/track_hours=False). F-026: tenant-Filter.
        sick_absences = db.query(Absence).filter(
            Absence.user_id == user.id,
            Absence.tenant_id == user.tenant_id,
            Absence.type == AbsenceType.SICK,
            date_in_year(Absence.date, year)
        ).all()
        sick_days = float(calculation_service.absence_days(db, user, sick_absences).quantize(Decimal('0.1')))

        # #415: Wochenstunden zum JAHRESBEGINN + die Änderungen als eigene
        # (angehängte) Spalte — die Jahres-Soll/Ist-Werte daneben rechnen
        # historisch, die Spalte muss dazu passen.
        wh_segments = calculation_service.weekly_hours_segments(
            db, user, date(year, 1, 1), date(year, 12, 31)
        )

        # Write data
        sheet.cell(row=row, column=1).value = neutralize_spreadsheet_formula(f"{user.last_name}, {user.first_name}")
        sheet.cell(row=row, column=2).value = float(wh_segments[0].weekly_hours) if wh_segments else float(user.weekly_hours)
        sheet.cell(row=row, column=11).value = format_weekly_hours_history(wh_segments) or None
        sheet.cell(row=row, column=3).value = float(yearly_target)
        sheet.cell(row=row, column=3).number_format = '0.00'
        sheet.cell(row=row, column=4).value = float(yearly_actual)
        sheet.cell(row=row, column=4).number_format = '0.00'
        sheet.cell(row=row, column=5).value = float(yearly_balance)
        sheet.cell(row=row, column=5).number_format = '0.00'

        # Color code balance
        if yearly_balance > 0:
            sheet.cell(row=row, column=5).font = Font(color="006400")
        elif yearly_balance < 0:
            sheet.cell(row=row, column=5).font = Font(color="8B0000")

        sheet.cell(row=row, column=6).value = float(overtime)
        sheet.cell(row=row, column=6).number_format = '0.00'

        if overtime > 0:
            sheet.cell(row=row, column=6).font = Font(color="006400")
        elif overtime < 0:
            sheet.cell(row=row, column=6).font = Font(color="8B0000")

        sheet.cell(row=row, column=7).value = vacation_account['budget_days']
        sheet.cell(row=row, column=8).value = float(vacation_account['used_days'])
        sheet.cell(row=row, column=8).number_format = '0.0'
        sheet.cell(row=row, column=9).value = float(vacation_account['remaining_days'])
        sheet.cell(row=row, column=9).number_format = '0.0'
        if include_health_data:
            sheet.cell(row=row, column=10).value = sick_days
            sheet.cell(row=row, column=10).number_format = '0.0'
        else:
            sheet.cell(row=row, column=10).value = "–"

        row += 1

    if not include_health_data:
        # Mark the column header to indicate data is protected
        sheet.cell(row=3, column=10).value = "Krankheitstage (geschützt)"

    # Adjust column widths (#415: Spalte 11 = Stundenänderungen, breiter Freitext)
    for col in range(1, 11):
        sheet.column_dimensions[get_column_letter(col)].width = 14
    # #431: ein Tagesplan-Satz („ab 15.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 =
    # 17,0 h/Woche") ist rund 53 Zeichen lang und passte in die auf den
    # #415-Satz zugeschnittenen 34 nicht mehr. Nur die Breite, keine neue
    # oder verschobene Spalte.
    sheet.column_dimensions[get_column_letter(11)].width = 54


def _create_absences_overview_sheet(wb: Workbook, db: Session, users: List[User], year: int, include_health_data: bool = False):
    """Create absences overview sheet."""
    sheet = wb.create_sheet(title="Abwesenheiten")

    # Title
    sheet.cell(row=1, column=1).value = f"Abwesenheiten {year}"
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.merge_cells('A1:H1')

    # Headers
    headers = ["Name", "Urlaub (Tage)", "Krank (Tage)", "Fortbildung (Tage)", "ÜStd.-Ausgleich (Tage)", "Sonstiges (Tage)", "Bez. Freistellung (Tage)", "Gesamt (Tage)"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    row = 4
    for user in users:
        # Tagesprinzip (§3 BUrlG, #156/#205): die TAGE tagebasiert zählen — exakt
        # wie die Live-Reports (reports.py) und get_vacation_account, NICHT als
        # Σ(Absence.hours) ÷ ⌀-Tagessoll. Die naive Stundenrechnung driftet bei
        # ungleichmäßigem Tagesplan / Halbtagen und liefert für track_hours=False
        # 0 Tage trotz gebuchten Urlaubs.
        def _days(atype, _user=user):
            absences = db.query(Absence).filter(
                Absence.user_id == _user.id,
                Absence.tenant_id == _user.tenant_id,  # F-026
                Absence.type == atype,
                date_in_year(Absence.date, year),
            ).all()
            return float(calculation_service.absence_days(db, _user, absences).quantize(Decimal('0.1')))

        # VACATION über get_vacation_account → schließt die #146 free+counts_as_
        # vacation-Sondertage (24./31.12.) ein, konsistent mit der Urlaubs-Rechnung.
        vacation_days = round(float(calculation_service.get_vacation_account(db, user, year)["used_days"]), 1)
        sick_days = _days(AbsenceType.SICK)
        training_days = _days(AbsenceType.TRAINING)
        overtime_comp_days = _days(AbsenceType.OVERTIME)
        other_days = _days(AbsenceType.OTHER)
        paid_leave_days = _days(AbsenceType.PAID_LEAVE)

        total_days = vacation_days + (sick_days if include_health_data else 0) + training_days + overtime_comp_days + other_days + paid_leave_days

        # Write data
        sheet.cell(row=row, column=1).value = neutralize_spreadsheet_formula(f"{user.last_name}, {user.first_name}")
        sheet.cell(row=row, column=2).value = vacation_days
        sheet.cell(row=row, column=2).number_format = '0.0'
        if include_health_data:
            sheet.cell(row=row, column=3).value = sick_days
            sheet.cell(row=row, column=3).number_format = '0.0'
        else:
            sheet.cell(row=row, column=3).value = "–"
        sheet.cell(row=row, column=4).value = training_days
        sheet.cell(row=row, column=4).number_format = '0.0'
        sheet.cell(row=row, column=5).value = overtime_comp_days
        sheet.cell(row=row, column=5).number_format = '0.0'
        sheet.cell(row=row, column=6).value = other_days
        sheet.cell(row=row, column=6).number_format = '0.0'
        sheet.cell(row=row, column=7).value = paid_leave_days
        sheet.cell(row=row, column=7).number_format = '0.0'
        sheet.cell(row=row, column=8).value = total_days
        sheet.cell(row=row, column=8).number_format = '0.0'
        sheet.cell(row=row, column=8).font = Font(bold=True)

        row += 1

    if not include_health_data:
        sheet.cell(row=3, column=3).value = "Krank (Tage) (geschützt)"

    # Adjust column widths
    for col in range(1, 9):
        sheet.column_dimensions[get_column_letter(col)].width = 16


def _create_employee_yearly_sheet(wb: Workbook, db: Session, user: User, year: int, include_health_data: bool = False):
    """
    Create detailed yearly sheet for a single employee with all days.
    Similar to monthly report but for the entire year.
    """
    sheet = wb.create_sheet(title=f"{user.last_name[:20]}")
    reason_names = _load_reason_names(db, user.tenant_id)  # #312

    # Title
    sheet.cell(row=1, column=1).value = neutralize_spreadsheet_formula(f"{user.first_name} {user.last_name} - Jahresreport {year}")
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.merge_cells('A1:J1')

    # Row 2: ArbZG-relevante Mitarbeiter-Flags (§16 ArbZG Aufzeichnungspflicht)
    sheet.cell(row=2, column=1).value = "§18 ArbZG-befreit:"
    sheet.cell(row=2, column=1).font = Font(bold=True)
    sheet.cell(row=2, column=2).value = "Ja" if user.exempt_from_arbzg else "Nein"
    # DSGVO F-006: is_night_worker is health-adjacent data – only show when include_health_data=True
    sheet.cell(row=2, column=4).value = "Nachtarbeitnehmer (§6 Abs. 2 ArbZG):"
    sheet.cell(row=2, column=4).font = Font(bold=True)
    sheet.cell(row=2, column=5).value = ("Ja" if user.is_night_worker else "Nein") if include_health_data else "–"

    # #415: Wochenstunden zum Jahresbeginn + Änderungen im Jahr
    _wh_segments = calculation_service.weekly_hours_segments(
        db, user, date(year, 1, 1), date(year, 12, 31)
    )
    sheet.cell(row=2, column=7).value = "Wochenstunden:"
    sheet.cell(row=2, column=7).font = Font(bold=True)
    sheet.cell(row=2, column=8).value = float(_wh_segments[0].weekly_hours) if _wh_segments else float(user.weekly_hours)
    _wh_history = format_weekly_hours_history(_wh_segments)
    if _wh_history:
        sheet.cell(row=2, column=9).value = _wh_history

    # Row 3: Column headers
    headers = ["Datum", "Wochentag", "Von", "Bis", "Pause (Min)", "Netto (Std)", "Soll (Std)", "Differenz", "Abwesenheit", "Bemerkung"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Get all time entries for the year (list-based: multiple entries per day)
    # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
    time_entries = db.query(TimeEntry).filter(
        TimeEntry.user_id == user.id,
        TimeEntry.tenant_id == user.tenant_id,
        date_in_year(TimeEntry.date, year)
    ).order_by(TimeEntry.start_time).all()
    entries_by_date = _group_by_date(time_entries)  # #219: shared

    # Get all absences for the year.
    # I-1: pro Tag eine LISTE (mehrere Absences je Tag möglich, s. _create_employee_sheet).
    # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
    absences = db.query(Absence).filter(
        Absence.user_id == user.id,
        Absence.tenant_id == user.tenant_id,
        date_in_year(Absence.date, year)
    ).all()
    absences_by_date = _group_by_date(absences)  # #219: shared

    # Get public holidays for the year (tenant-scoped; see generate_monthly_report).
    holidays = db.query(PublicHoliday).filter(
        PublicHoliday.tenant_id == user.tenant_id,
        PublicHoliday.year == year,
    ).all()
    holidays_by_date = {holiday.date: holiday for holiday in holidays}

    # #146: configurable 24./31.12. handling (see _create_employee_sheet). N-1.
    special_day_config = special_days_service.get_special_day_config(db, user.tenant_id, year)

    # German weekday names
    weekday_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

    row = 4
    total_net = Decimal('0.00')
    total_target = Decimal('0.00')
    night_work_count = 0
    current_month = 0

    # Iterate through all days of the year
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    current_date = start_date

    while current_date <= end_date:
        # Add month separator
        if current_date.month != current_month:
            if current_month > 0:
                row += 1  # Empty row between months

            current_month = current_date.month
            month_names = [
                "Januar", "Februar", "März", "April", "Mai", "Juni",
                "Juli", "August", "September", "Oktober", "November", "Dezember"
            ]

            sheet.cell(row=row, column=1).value = month_names[current_month - 1]
            sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
            sheet.cell(row=row, column=1).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            sheet.merge_cells(f'A{row}:J{row}')
            row += 1

        weekday = current_date.weekday()
        weekday_name = weekday_names[weekday]
        is_sunday = weekday == 6

        is_weekend = weekday >= 5
        is_holiday = current_date in holidays_by_date
        day_absences = absences_by_date.get(current_date, [])  # I-1: alle Absences des Tages

        sheet.cell(row=row, column=1).value = current_date
        sheet.cell(row=row, column=1).number_format = 'DD.MM.YYYY'
        sheet.cell(row=row, column=2).value = weekday_name

        day_entries = entries_by_date.get(current_date, [])

        # Night work check (§6 / §2 Abs. 4 ArbZG)
        is_night_wrk = any(
            e.end_time is not None and is_night_work(e.start_time, e.end_time)
            for e in day_entries
        )
        if is_night_wrk:
            night_work_count += 1

        if day_entries:
            first_start = day_entries[0].start_time
            # spätestes Ende des Tages, nicht der zuletzt START-ende Eintrag — bei
            # überlappenden Mehrfach-Einträgen (A 08–17, B 12–14) war "Bis" sonst
            # 14:00 statt 17:00 (Review 2026-06-23, §16-Korrektheit).
            last_end = max((e.end_time for e in day_entries if e.end_time), default=None)
            total_break = sum(e.break_minutes or 0 for e in day_entries)
            total_day_net = sum(e.net_hours for e in day_entries)
            sheet.cell(row=row, column=3).value = first_start.strftime('%H:%M')
            sheet.cell(row=row, column=4).value = last_end.strftime('%H:%M') if last_end else 'offen'
            sheet.cell(row=row, column=5).value = total_break
            sheet.cell(row=row, column=6).value = float(total_day_net)
            sheet.cell(row=row, column=6).number_format = '0.00'
            # Bemerkung (col 10): §10-Ausnahmegrund hat Vorrang, dann entry.note
            bemerkung_parts = []
            for e in day_entries:
                if e.sunday_exception_reason and (is_sunday or is_holiday):
                    bemerkung_parts.append(f"§10-Ausnahmegrund: {e.sunday_exception_reason}")
                if e.note:
                    bemerkung_parts.append(e.note)
            if bemerkung_parts:
                sheet.cell(row=row, column=10).value = neutralize_spreadsheet_formula(" | ".join(bemerkung_parts))
            net = total_day_net
            total_net += net
        else:
            net = Decimal('0.00')
            sheet.cell(row=row, column=6).value = 0.00
            sheet.cell(row=row, column=6).number_format = '0.00'

        schedule = calculation_service.get_schedule_for_date(db, user, current_date)
        daily_target = calculation_service.get_daily_target_for_date(user, current_date, schedule)
        _sd_factor = special_days_service.special_day_target_factor(current_date, special_day_config)
        if _sd_factor is not None:
            daily_target = daily_target * _sd_factor

        # Target hours + Abwesenheit (col 9) – korrekte Labels für §9/§10/§6
        if is_weekend:
            target = Decimal('0.00')
            if is_sunday and day_entries:
                abw = "Sonntagsarbeit (§9/§10 ArbZG)"
            elif is_sunday:
                abw = "Sonntag"
            else:
                abw = "Samstag"
            if is_night_wrk:
                abw += " | Nachtarbeit (§6 ArbZG)"
            sheet.cell(row=row, column=9).value = abw
            for col in range(1, 11):
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        elif is_holiday:
            target = Decimal('0.00')
            holiday = holidays_by_date[current_date]
            if day_entries:
                abw = f"Feiertagsarbeit: {holiday.name} (§9/§10 ArbZG)"
            else:
                abw = f"Feiertag: {holiday.name}"
            if is_night_wrk:
                abw += " | Nachtarbeit (§6 ArbZG)"
            sheet.cell(row=row, column=9).value = abw
            for col in range(1, 11):
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        elif day_absences:
            # Release-Review 1.16.0: zentrale Soll-Quelle statt pauschal 0.
            target = absence_day_target(db, user, current_date, day_absences, set(holidays_by_date), special_day_config)
            absence_type_map = ABSENCE_TYPE_LABELS_DE
            # I-1: ALLE Absences des Tages anzeigen; DSGVO F-003: Krank ohne
            # Health-Flag maskieren (Label "Abwesenheit", Notiz unterdrückt).
            abw_parts = []
            note_parts = []
            for absence in day_absences:
                type_name, show_note = _absence_export_label(
                    absence, absence_type_map, reason_names, include_health_data)
                if show_note and absence.note:
                    note_parts.append(absence.note)
                abw_parts.append(f"{type_name} ({float(absence.hours)}h)")
            sheet.cell(row=row, column=9).value = neutralize_spreadsheet_formula(" | ".join(abw_parts))  # custom-reason label = user text → neutralisieren
            if note_parts:
                sheet.cell(row=row, column=10).value = neutralize_spreadsheet_formula(" | ".join(note_parts))
        else:
            target = daily_target
            if is_night_wrk:
                sheet.cell(row=row, column=9).value = "Nachtarbeit (§6 ArbZG)"

        sheet.cell(row=row, column=7).value = float(target)
        sheet.cell(row=row, column=7).number_format = '0.00'
        total_target += target

        diff = net - target
        sheet.cell(row=row, column=8).value = float(diff)
        sheet.cell(row=row, column=8).number_format = '0.00'

        if diff > 0:
            sheet.cell(row=row, column=8).font = Font(color="006400")
        elif diff < 0:
            sheet.cell(row=row, column=8).font = Font(color="8B0000")

        row += 1
        current_date += timedelta(days=1)

    # Summary section
    row += 2
    sheet.cell(row=row, column=1).value = "Jahressumme"
    sheet.cell(row=row, column=1).font = Font(bold=True, size=12)

    # #377 Baustein 2b (Follow-up zu Finding 1, Whole-Branch-Review): für
    # Fix-Modus-MA (use_fixed_monthly_target) MUSS die Jahres-Summary mit dem
    # modus-bewussten Σ get_monthly_target/get_monthly_actual übereinstimmen —
    # exakt derselbe Fix wie im Monatsexport (2ae1c6e0), sonst widerspricht
    # sich das §16-Jahresdokument selbst gegen die bereits modus-bewusste
    # Jahresübersicht (_create_yearly_overview_sheet) und "Überstunden
    # kumuliert" darunter (get_overtime_account, bereits modus-bewusst). Die
    # Per-Tag-Detailzeilen oben bleiben unverändert. Nicht-Modus-MA bleiben
    # exakt auf der alten Per-Tag-Summe (byte-identisch).
    if getattr(user, "use_fixed_monthly_target", False) and getattr(user, "agreed_monthly_hours", None):
        summary_target = sum(
            (calculation_service.get_monthly_target(db, user, year, m) for m in range(1, 13)),
            start=Decimal('0'),
        )
        summary_actual = sum(
            (calculation_service.get_monthly_actual(db, user, year, m) for m in range(1, 13)),
            start=Decimal('0'),
        )
    else:
        summary_target = total_target
        summary_actual = total_net

    row += 1
    sheet.cell(row=row, column=1).value = "Soll-Stunden Jahr:"
    sheet.cell(row=row, column=2).value = float(summary_target)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    sheet.cell(row=row, column=1).value = "Ist-Stunden Jahr:"
    sheet.cell(row=row, column=2).value = float(summary_actual)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)

    row += 1
    yearly_balance = summary_actual - summary_target
    sheet.cell(row=row, column=1).value = "Saldo Jahr:"
    sheet.cell(row=row, column=2).value = float(yearly_balance)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)
    if yearly_balance > 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="006400")
    elif yearly_balance < 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="8B0000")

    row += 1
    overtime_account = calculation_service.get_overtime_account(db, user, year, 12)
    sheet.cell(row=row, column=1).value = "Überstunden kumuliert:"
    sheet.cell(row=row, column=2).value = float(overtime_account)
    sheet.cell(row=row, column=2).number_format = '0.00'
    sheet.cell(row=row, column=1).font = Font(bold=True)
    if overtime_account > 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="006400")
    elif overtime_account < 0:
        sheet.cell(row=row, column=2).font = Font(bold=True, color="8B0000")

    row += 1
    vacation_account = calculation_service.get_vacation_account(db, user, year)
    sheet.cell(row=row, column=1).value = "Urlaub genommen (Tage):"
    sheet.cell(row=row, column=2).value = float(vacation_account['used_days'])
    sheet.cell(row=row, column=2).number_format = '0.0'

    row += 1
    sheet.cell(row=row, column=1).value = "Urlaub Rest (Tage):"
    sheet.cell(row=row, column=2).value = float(vacation_account['remaining_days'])
    sheet.cell(row=row, column=2).number_format = '0.0'

    row += 1
    sheet.cell(row=row, column=1).value = "Nachtarbeitstage (§6 ArbZG):"
    sheet.cell(row=row, column=2).value = night_work_count
    sheet.cell(row=row, column=1).font = Font(bold=True)

    # Adjust column widths
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 28
    sheet.column_dimensions['J'].width = 35


def generate_yearly_report_classic(db: Session, year: int, include_health_data: bool = False, tenant_id=None) -> BytesIO:
    """
    Generate classic yearly report (compact format with months as columns).
    Creates one sheet per employee.

    Args:
        db: Database session
        year: Year
        tenant_id: F-026 belt-and-suspenders explicit tenant filter (on top of RLS)

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Get all active, non-hidden employees (F-026: explicit tenant filter on top of RLS)
    users = export_users(db, tenant_id, date(year, 1, 1), date(year, 12, 31))

    for user in users:
        _create_employee_classic_sheet(wb, db, user, year, include_health_data)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _create_employee_classic_sheet(wb: Workbook, db: Session, user: User, year: int, include_health_data: bool = False):
    """
    Create classic yearly overview sheet for one employee.
    Format: Months as columns, compact overview with running balances.

    #377 Baustein 2b — bekannte Grenze: dieser klassische Kompaktbericht baut auf
    dem Brutto-Soll-Modell ({Arbeit, Krank, Urlaub} pro Monat, Per-Tag-Soll) auf
    und ist für `use_fixed_monthly_target`-MA NICHT modus-korrekt (das feste
    Monats-Soll + die Feiertags-/Fehltags-Gutschrift passen nicht in dieses
    Layout). Für Minijob-Konten mit festem Monats-Soll den Standard-Export
    (`generate_yearly_report`/`_yearly_employee_sheet`) nutzen — der ist
    modus-korrekt. Eine Anpassung des Legacy-Formats ist bewusst offen.
    """
    sheet = wb.create_sheet(title=f"{user.last_name}")

    # Styles
    header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Month names
    month_names = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                   'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']

    # Row 1: Practice header (DSGVO F-016: use configurable env vars)
    sheet.cell(row=1, column=1).value = settings.PRACTICE_NAME
    sheet.cell(row=1, column=1).font = header_font
    if settings.PRACTICE_ADDRESS:
        sheet.cell(row=1, column=2).value = settings.PRACTICE_ADDRESS

    # Row 2: Employee name
    sheet.cell(row=2, column=1).value = "Mitarbeiterin"
    sheet.cell(row=2, column=1).font = bold_font
    sheet.cell(row=2, column=2).value = neutralize_spreadsheet_formula(f"{user.first_name} {user.last_name}")

    # Row 3: Year title + ArbZG-Flags
    sheet.cell(row=3, column=1).value = "Jahresarbeitszeiten"
    sheet.cell(row=3, column=1).font = bold_font
    sheet.cell(row=3, column=2).value = year
    sheet.cell(row=3, column=15).value = "§18 ArbZG-befreit:"
    sheet.cell(row=3, column=15).font = bold_font
    sheet.cell(row=3, column=16).value = "Ja" if user.exempt_from_arbzg else "Nein"
    # DSGVO F-006: protect health-adjacent is_night_worker
    sheet.cell(row=2, column=15).value = "Nachtarbeitnehmer (§6 Abs. 2):"
    sheet.cell(row=2, column=15).font = bold_font
    sheet.cell(row=2, column=16).value = ("Ja" if user.is_night_worker else "Nein") if include_health_data else "–"

    # Row 4: Month headers (columns 3-14 for Jan-Dec)
    sheet.cell(row=4, column=2).value = "Übertrag"
    sheet.cell(row=4, column=2).font = bold_font
    sheet.cell(row=4, column=2).alignment = center_align
    for i, month_name in enumerate(month_names, start=3):
        sheet.cell(row=4, column=i).value = month_name
        sheet.cell(row=4, column=i).font = bold_font
        sheet.cell(row=4, column=i).alignment = center_align

    # Row 5: Previous year carry-over
    sheet.cell(row=5, column=2).value = "Vorjahr"
    sheet.cell(row=5, column=2).font = normal_font

    # Get previous year overtime
    prev_year_overtime = calculation_service.get_overtime_account(db, user, year - 1, 12)
    sheet.cell(row=5, column=3).value = float(prev_year_overtime)
    sheet.cell(row=5, column=3).number_format = '0.0'

    # Row labels
    sheet.cell(row=6, column=1).value = "Zahl Arbeitstage im Monat"
    sheet.cell(row=7, column=1).value = "Sollstunden im Monat"
    sheet.cell(row=8, column=1).value = "minus Krank  Std."
    sheet.cell(row=9, column=1).value = "minus Urlaub  Std."
    sheet.cell(row=10, column=1).value = "aktuelle Sollstundenzahl"
    sheet.cell(row=11, column=1).value = "erbrachte Stunden"
    sheet.cell(row=12, column=1).value = "StundenSaldo"
    sheet.cell(row=13, column=1).value = "Jan.- Monatsende"
    sheet.cell(row=14, column=1).value = "Überstunden / Minusstunden"
    sheet.cell(row=15, column=1).value = "Resturlaub + Urlaub in Std."
    sheet.cell(row=16, column=1).value = "Nachtarbeitstage (§6 ArbZG)"

    for row in range(6, 17):
        sheet.cell(row=row, column=1).font = normal_font

    # #150/Fix #4: das kumulative Überstundenkonto je Monat (Row 14) EINMAL als
    # Single-Pass holen, statt get_overtime_account pro Monat zu rufen (jede
    # Einzelrufung iteriert ab Carryover-Start neu -> O(Monate²)).
    # history[(year, m)] entspricht bitgenau get_overtime_account(year, m)
    # (gepinnt: test_overtime_history_matches_account); Monate vor dem
    # History-Bereich liefert get_overtime_account 0.00 -> Default.
    overtime_history = calculation_service.get_overtime_history(db, user, year, 12)

    # Fix #7: das Urlaubskonto hängt nur an (user, year) und ist über alle 12
    # Monate identisch — EINMAL vor der Schleife berechnen statt pro Monat.
    vacation_account = calculation_service.get_vacation_account(db, user, year)

    # Calculate data for each month
    for month in range(1, 13):
        col = month + 2  # Column 3 = January, ..., Column 14 = December

        # Row 6: Working days in month
        working_days = calculation_service.get_working_days_in_month(db, year, month)
        sheet.cell(row=6, column=col).value = working_days
        sheet.cell(row=6, column=col).alignment = center_align

        # Row 7: GROSS target hours (all working days). The explicit "minus Krank
        # / minus Urlaub" rows below reduce it — traditional Stundenkonto layout.
        # NOT get_monthly_target (which already nets vacation out → using it here
        # double-subtracted vacation and produced a phantom +Saldo).
        target_hours = calculation_service.get_gross_monthly_target(db, user, year, month)
        sheet.cell(row=7, column=col).value = float(target_hours)
        sheet.cell(row=7, column=col).number_format = '0.0'
        sheet.cell(row=7, column=col).alignment = right_align

        # Row 8: Sick hours
        # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
        sick_absences = db.query(Absence).filter(
            Absence.user_id == user.id,
            Absence.tenant_id == user.tenant_id,
            Absence.type == AbsenceType.SICK,
            date_in_month(Absence.date, year, month)
        ).all()
        # #198: gefenstert wie das Ist (Row 11 = get_monthly_worked_hours, #195).
        # Sonst verschoebe eine out-of-window-SICK-Absence den adjusted_target
        # (Row 10 = target − sick − vacation), ohne die Ist-Seite zu beruehren.
        sick_hours = sum(float(a.hours) for a in sick_absences
                         if calculation_service._within_employment_window(user, a.date))
        real_sick = sick_hours  # kept for the masked-case credit on Row 11
        if include_health_data:
            sheet.cell(row=8, column=col).value = sick_hours
            sheet.cell(row=8, column=col).number_format = '0.0'
        else:
            sheet.cell(row=8, column=col).value = "–"
            sick_hours = 0  # not subtracted from soll when masked (Art. 9)
        sheet.cell(row=8, column=col).alignment = right_align

        # Row 9: Vacation hours
        # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
        vacation_absences = db.query(Absence).filter(
            Absence.user_id == user.id,
            Absence.tenant_id == user.tenant_id,
            Absence.type == AbsenceType.VACATION,
            date_in_month(Absence.date, year, month)
        ).all()
        # #198: gefenstert wie Ist + sick (Row 10/11-Konsistenz).
        vacation_hours = sum(float(a.hours) for a in vacation_absences
                             if calculation_service._within_employment_window(user, a.date))
        sheet.cell(row=9, column=col).value = vacation_hours
        sheet.cell(row=9, column=col).number_format = '0.0'
        sheet.cell(row=9, column=col).alignment = right_align

        # Row 10: Adjusted target (Target - Sick - Vacation)
        adjusted_target = float(target_hours) - sick_hours - vacation_hours
        sheet.cell(row=10, column=col).value = adjusted_target
        sheet.cell(row=10, column=col).number_format = '0.0'
        sheet.cell(row=10, column=col).alignment = right_align

        # Row 11: hours physically WORKED (no credited sick/training), paired with
        # the gross-minus-absences adjusted target so the balance equals the
        # canonical Ist − Soll for the {Arbeit, Krank, Urlaub} model this classic
        # format represents. When health data is masked, Krank is NOT subtracted
        # above; credit it to the worked side instead so the masked Saldo stays
        # correct AND Krank is never derivable from the displayed cells (Art. 9).
        # NOTE: TRAINING / OTHER / PAID_LEAVE / OVERTIME are not modelled by this
        # 2-row layout — use the standard report for richer absence mixes.
        actual_hours = float(calculation_service.get_monthly_worked_hours(db, user, year, month))
        if not include_health_data:
            actual_hours += real_sick
        sheet.cell(row=11, column=col).value = actual_hours
        sheet.cell(row=11, column=col).number_format = '0.0'
        sheet.cell(row=11, column=col).alignment = right_align

        # Row 12: Monthly balance (worked − adjusted target = canonical Ist − Soll)
        monthly_balance = float(actual_hours) - adjusted_target
        sheet.cell(row=12, column=col).value = monthly_balance
        sheet.cell(row=12, column=col).number_format = '0.0'
        sheet.cell(row=12, column=col).alignment = right_align

        # Apply color coding to balance
        if monthly_balance > 0:
            sheet.cell(row=12, column=col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif monthly_balance < 0:
            sheet.cell(row=12, column=col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Row 14: Cumulative overtime (Fix #4: aus dem Single-Pass-History)
        cumulative_overtime = overtime_history.get((year, month), Decimal('0.00'))
        sheet.cell(row=14, column=col).value = float(cumulative_overtime)
        sheet.cell(row=14, column=col).number_format = '0.0'
        sheet.cell(row=14, column=col).alignment = right_align
        sheet.cell(row=14, column=col).font = bold_font

        # Apply color coding to cumulative
        if cumulative_overtime > 0:
            sheet.cell(row=14, column=col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif cumulative_overtime < 0:
            sheet.cell(row=14, column=col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Row 15: Remaining vacation in hours (vacation_account: s. o., 1× berechnet)
        # Calculate remaining vacation up to this month.
        # F-026 (Review 2026-06-23): explizit auf den Tenant scopen.
        vacation_used_ytd = sum(
            float(a.hours) for a in db.query(Absence).filter(
                Absence.user_id == user.id,
                Absence.tenant_id == user.tenant_id,
                Absence.type == AbsenceType.VACATION,
                date_in_year_up_to_month(Absence.date, year, month),
            ).all()
        )
        # Jahresend-Spalte: den autoritativen remaining_hours-Wert aus
        # get_vacation_account nehmen — er beruecksichtigt die Sondertags-Abzuege
        # (24./31.12 als Urlaub, #188), die als Absence-Rows nicht existieren und
        # die naive budget − ytd-Rechnung sonst um einen Tagessoll ueberzeichnet.
        # Frühere Monate: Sondertage (Jahresende) sind noch nicht eingetreten.
        if month == 12:
            vacation_remaining = float(vacation_account['remaining_hours'])
        else:
            vacation_remaining = float(vacation_account['budget_hours']) - vacation_used_ytd
        sheet.cell(row=15, column=col).value = vacation_remaining
        sheet.cell(row=15, column=col).number_format = '0.0'
        sheet.cell(row=15, column=col).alignment = right_align

        # Row 16: Night work days per month (§6 ArbZG)
        # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
        month_entries = db.query(TimeEntry).filter(
            TimeEntry.user_id == user.id,
            TimeEntry.tenant_id == user.tenant_id,
            date_in_month(TimeEntry.date, year, month),
            TimeEntry.end_time.isnot(None),
        ).all()
        night_days = len({e.date for e in month_entries if is_night_work(e.start_time, e.end_time)})
        sheet.cell(row=16, column=col).value = night_days
        sheet.cell(row=16, column=col).alignment = center_align

    # Add daily hours info in corner (current value — informational)
    #
    # Fund G (Abschluss-Review #431): ``get_daily_target(user)`` war die letzte
    # verbliebene Bypass-Stelle der Export-Schicht — sie liest
    # ``user.weekly_hours``/``work_days_per_week`` direkt und kennt laut
    # eigenem Docstring weder Historie noch Tagesplan.
    #
    # Beurteilung (wie vom Fund gefordert), ob dieser klassische Jahresbericht
    # die Zahl ueberhaupt sinnvoll ausweisen KANN: für eine gleichmässige
    # Woche ja — die Division bleibt ein einzelner, in sich stimmiger
    # "aktueller" Tageswert (byte-identisch zu vorher, s. Docstring oben:
    # dieselbe bewusste Brutto-Legacy-Beschränkung wie beim #377-Fix-Modus).
    # Für einen TAGESPLAN-MA dagegen nicht: die Division (Summe ÷ Anzahl Tage)
    # wäre ein arithmetisches Mittel, das mit KEINEM einzelnen Wochentag
    # übereinstimmt (Mo 8 / Di 5 / Mi 4 → 5,67 h — kein Tag hat 5,67 h) und
    # der historisch aufgelösten Tagesplan-Zeile widerspräche, die derselbe
    # Bericht weiter oben (Zeile 7, ``get_gross_monthly_target``) für denselben
    # Mitarbeitenden zeigt. Ein einzelner Jahres-Skalar kann einen Tagesplan
    # grundsätzlich nicht abbilden — deshalb hier der Tagesplan-Text
    # (``format_day_plan``, dieselbe Klartext-Funktion wie in den #415/#431-
    # Kopfzeilen der übrigen Exporte) statt eines falschen Mittelwerts. Keine
    # Spalte verschoben — nur der Inhalt dieser einen informativen Ecke.
    sheet.cell(row=6, column=17).value = "Tagesplan:" if user.use_daily_schedule else "tägl. Std:"
    sheet.cell(row=6, column=17).font = normal_font
    if user.use_daily_schedule:
        day_hours = [user.hours_monday, user.hours_tuesday, user.hours_wednesday,
                     user.hours_thursday, user.hours_friday]
        sheet.cell(row=6, column=18).value = format_day_plan(day_hours) or "–"
    else:
        daily_hours = calculation_service.get_daily_target(user)
        sheet.cell(row=6, column=18).value = float(daily_hours)
        sheet.cell(row=6, column=18).number_format = '0.0'

    # Set column widths
    sheet.column_dimensions['A'].width = 28
    sheet.column_dimensions['B'].width = 12
    for col in range(3, 15):
        sheet.column_dimensions[get_column_letter(col)].width = 11
    sheet.column_dimensions['Q'].width = 10
    sheet.column_dimensions['R'].width = 8

    # Add borders to data area
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(4, 17):
        for col in range(2, 15):
            sheet.cell(row=row, column=col).border = thin_border


# ---------------------------------------------------------------------------
# PDF Export (reportlab)
# ---------------------------------------------------------------------------

def generate_monthly_report_pdf(db: Session, year: int, month: int, include_health_data: bool = False, tenant_id=None) -> BytesIO:
    """
    Generate PDF monthly report for all employees.
    One page per employee, landscape A4.
    Same data as Excel monthly report.
    F-026: pass tenant_id for belt-and-suspenders explicit filter.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"PraxisZeit Monatsreport {month:02d}/{year}",
    )

    # Styles
    s_normal = ParagraphStyle('n', fontName='Helvetica', fontSize=7, leading=9)
    s_bold = ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=7, leading=9)
    s_center = ParagraphStyle('c', fontName='Helvetica', fontSize=7, leading=9, alignment=TA_CENTER)
    s_title = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=10, leading=13)
    s_sum_lbl = ParagraphStyle('sl', fontName='Helvetica-Bold', fontSize=7.5, leading=10)
    s_sum_val = ParagraphStyle('sv', fontName='Helvetica', fontSize=7.5, leading=10)

    def colored(text, hex_color, bold=False):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        return Paragraph(text, ParagraphStyle('col', fontName=fn, fontSize=7, leading=9,
                                              textColor=colors.HexColor(hex_color)))

    month_names = ['Januar', 'Februar', 'Maerz', 'April', 'Mai', 'Juni',
                   'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']

    # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
    users = export_users(db, tenant_id, date(year, month, 1),
                         date(year, month, monthrange(year, month)[1]))

    # Landscape A4: 297mm − 30mm margins = 267mm usable
    col_widths = [22*mm, 10*mm, 13*mm, 13*mm, 15*mm, 16*mm, 14*mm, 16*mm, 74*mm, 74*mm]
    weekday_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    absence_type_map = ABSENCE_TYPE_LABELS_DE

    story = []

    for i, user in enumerate(users):
        if i > 0:
            story.append(PageBreak())
        reason_names = _load_reason_names(db, user.tenant_id)  # #312

        # ── Title ──
        story.append(Paragraph(
            f"PraxisZeit \u2013 Monatsreport {month_names[month - 1]} {year}",
            s_title
        ))
        story.append(Spacer(1, 2 * mm))

        # ── Employee meta ──
        arbzg_flag = " | \u00a718-befreit" if user.exempt_from_arbzg else ""
        night_flag = " | Nachtarbeitnehmer (\u00a76)" if user.is_night_worker else ""
        # #415: Wochenstunden zum Monatsbeginn + \u00c4nderungen im Monat.
        _wh_segments = calculation_service.weekly_hours_segments(
            db, user, date(year, month, 1), date(year, month, monthrange(year, month)[1])
        )
        _wh_start = _wh_segments[0].weekly_hours if _wh_segments else user.weekly_hours
        # #431: `compact` — die Meta ist ein Inline-Absatz in Schriftgroesse 8 auf
        # Querformat; ein ausgeschriebener Tagesplan („Mo 8,0 / Di 5,0 / …")
        # sprengt die Zeile. Der Schalter wirkt NUR im Tagesplan-Zweig, die
        # #415-Formulierung fuer alle uebrigen Mitarbeitenden bleibt woertlich.
        _wh_history = format_weekly_hours_history(_wh_segments, compact=True)
        _wh_flag = f" | Stunden\u00e4nderung: {_wh_history}" if _wh_history else ""
        # #431: der Startwert lief bis hierher ueber `f"{\u2026:.1f}"` \u2014 mit
        # Dezimal-PUNKT und gerundet. Bei einer Tagesplan-Summe von 17,75 stand
        # dann \u201e\u2026 17.8h/Woche | Stundenaenderung: \u2026 = 17,75 h/Woche" in EINEM
        # Satz: zwei Schreibweisen und zwei Werte fuer dieselbe Zahl. Jetzt
        # dieselbe verlustfreie deutsche Formatierung wie der Rest der Zeile.
        meta_label = f"{user.first_name} {user.last_name}  \u2013  {_de_hours_exact(_wh_start)}h/Woche{_wh_flag}{arbzg_flag}{night_flag}"
        story.append(Paragraph(escape_pdf_text(meta_label), ParagraphStyle('meta', fontName='Helvetica', fontSize=8, leading=10,
                                                           textColor=colors.HexColor('#374151'))))
        story.append(Spacer(1, 2 * mm))

        # ── Fetch data ──
        _, last_day = monthrange(year, month)

        # F-026: explicit tenant filter (belt-and-suspenders on top of RLS)
        time_entries = db.query(TimeEntry).filter(
            TimeEntry.user_id == user.id,
            TimeEntry.tenant_id == user.tenant_id,
            date_in_month(TimeEntry.date, year, month),
        ).order_by(TimeEntry.start_time).all()
        entries_by_date: dict = {}
        for te in time_entries:
            entries_by_date.setdefault(te.date, []).append(te)

        absences = db.query(Absence).filter(
            Absence.user_id == user.id,
            Absence.tenant_id == user.tenant_id,
            date_in_month(Absence.date, year, month),
        ).all()
        # I-1: pro Tag eine LISTE (mehrere Absences je Tag möglich, s. _create_employee_sheet).
        absences_by_date: dict = {}
        for a in absences:
            absences_by_date.setdefault(a.date, []).append(a)

        holidays = db.query(PublicHoliday).filter(
            PublicHoliday.tenant_id == user.tenant_id,
            date_in_month(PublicHoliday.date, year, month),
        ).all()
        holidays_by_date = {h.date: h for h in holidays}

        # #146: configurable 24./31.12. handling (see _create_employee_sheet). N-1.
        special_day_config = special_days_service.get_special_day_config(db, user.tenant_id, year)

        # ── Build table ──
        headers = ['Datum', 'WT', 'Von', 'Bis', 'Pause\n(Min)', 'Netto\n(Std)', 'Soll\n(Std)', 'Diff.', 'Abwesenheit', 'Bemerkung']
        table_data = [[Paragraph(h, ParagraphStyle('hdr', fontName='Helvetica-Bold', fontSize=7,
                                                    leading=9, alignment=TA_CENTER))
                       for h in headers]]
        row_bgs = {}  # row_index -> HexColor

        total_net = Decimal('0.00')
        total_target = Decimal('0.00')
        night_work_count = 0

        for day in range(1, last_day + 1):
            cur = date(year, month, day)
            wd = cur.weekday()
            is_weekend = wd >= 5
            is_sunday = wd == 6
            is_holiday = cur in holidays_by_date
            day_absences = absences_by_date.get(cur, [])  # I-1: alle Absences des Tages
            day_entries = entries_by_date.get(cur, [])

            is_night = any(
                e.end_time is not None and is_night_work(e.start_time, e.end_time)
                for e in day_entries
            )
            if is_night:
                night_work_count += 1

            if day_entries:
                von = day_entries[0].start_time.strftime('%H:%M')
                # spätestes Ende des Tages, nicht der zuletzt START-ende Eintrag — bei
                # überlappenden Mehrfach-Einträgen (A 08–17, B 12–14) war "Bis" sonst
                # 14:00 statt 17:00 (Review 2026-06-23, §16-Korrektheit).
                last_end = max((e.end_time for e in day_entries if e.end_time), default=None)
                bis = last_end.strftime('%H:%M') if last_end else 'offen'
                pause_str = str(sum(e.break_minutes or 0 for e in day_entries))
                total_day_net = sum(e.net_hours for e in day_entries)
                netto_val = float(total_day_net)
                net = total_day_net
                total_net += net
                bem_parts = []
                for e in day_entries:
                    if e.sunday_exception_reason and (is_sunday or is_holiday):
                        bem_parts.append(f"\u00a710: {e.sunday_exception_reason}")
                    if e.note:
                        bem_parts.append(e.note)
                bem = " | ".join(bem_parts)
            else:
                von = bis = pause_str = bem = ''
                netto_val = 0.0
                net = Decimal('0.00')

            # Per-day target using the historical contract snapshot (#431)
            schedule = calculation_service.get_schedule_for_date(db, user, cur)
            daily_target = calculation_service.get_daily_target_for_date(user, cur, schedule)
            _sd_factor = special_days_service.special_day_target_factor(cur, special_day_config)
            if _sd_factor is not None:
                daily_target = daily_target * _sd_factor

            if is_weekend:
                target = Decimal('0.00')
                if is_sunday and day_entries:
                    abw = 'Sonntagsarbeit (\u00a79/\u00a710)'
                elif is_sunday:
                    abw = 'Sonntag'
                else:
                    abw = 'Samstag'
                if is_night:
                    abw += ' | Nachtarbeit'
                bg = colors.HexColor('#E8E8E8')
            elif is_holiday:
                target = Decimal('0.00')
                hname = holidays_by_date[cur].name
                abw = f"Feiertagsarbeit: {hname}" if day_entries else f"Feiertag: {hname}"
                if is_night:
                    abw += ' | Nachtarbeit'
                bg = colors.HexColor('#FFFFCC')
            elif day_absences:
                # Release-Review 1.16.0: zentrale Soll-Quelle statt pauschal 0.
                target = absence_day_target(db, user, cur, day_absences, set(holidays_by_date), special_day_config)
                # I-1: ALLE Absences des Tages zeigen; DSGVO F-003: Krank ohne
                # Health-Flag maskieren (Label 'Abwesenheit', Notiz unterdrückt).
                abw_parts = []
                note_parts = []
                for absence in day_absences:
                    type_name, show_note = _absence_export_label(
                        absence, absence_type_map, reason_names, include_health_data)
                    if show_note and absence.note:
                        note_parts.append(absence.note)
                    abw_parts.append(f"{type_name} ({float(absence.hours):.1f}h)")
                abw = " | ".join(abw_parts)
                bem = " | ".join(note_parts)  # ersetzt evtl. Eintrags-Bemerkung wie zuvor
                bg = None
            else:
                target = daily_target
                abw = 'Nachtarbeit (\u00a76 ArbZG)' if is_night else ''
                bg = None

            total_target += target
            diff = net - target
            diff_str = f"{float(diff):+.2f}"
            if diff > 0:
                diff_cell = colored(diff_str, '#006400', bold=True)
            elif diff < 0:
                diff_cell = colored(diff_str, '#8B0000', bold=True)
            else:
                diff_cell = Paragraph(diff_str, s_center)

            row = [
                Paragraph(cur.strftime('%d.%m.%Y'), s_normal),
                Paragraph(weekday_names[wd], s_center),
                Paragraph(von, s_center),
                Paragraph(bis, s_center),
                Paragraph(pause_str, s_center),
                Paragraph(f"{netto_val:.2f}", s_center),
                Paragraph(f"{float(target):.2f}", s_center),
                diff_cell,
                Paragraph(escape_pdf_text(abw), s_normal),
                Paragraph(escape_pdf_text(bem), s_normal),
            ]
            table_data.append(row)
            if bg:
                row_bgs[len(table_data) - 1] = bg

        tbl_style = [
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCE5FF')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]
        for row_idx, bg_color in row_bgs.items():
            tbl_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))

        main_tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        main_tbl.setStyle(TableStyle(tbl_style))
        story.append(main_tbl)

        # ── Summary ──
        story.append(Spacer(1, 4 * mm))
        # Release-Review 1.16.0: #377 Baustein 2b — der Fixmodus-Branch fehlte hier
        # als einziger der drei Monats-Exportflächen (XLSX Z.439, ODS-Monat), sodass
        # dieselbe §16-Auskunft je Dateiformat unterschiedliche Soll-Zahlen trug und
        # das PDF sich gegen sein eigenes „Überstunden kumuliert" (bereits
        # modus-bewusst) stellte. Nicht-Modus-MA bleiben auf der Per-Tag-Summe.
        if getattr(user, "use_fixed_monthly_target", False) and getattr(user, "agreed_monthly_hours", None):
            summary_target = calculation_service.get_monthly_target(db, user, year, month)
            summary_actual = calculation_service.get_monthly_actual(db, user, year, month)
        else:
            summary_target = total_target
            summary_actual = total_net
        monthly_balance = summary_actual - summary_target
        overtime_account = calculation_service.get_overtime_account(db, user, year, month)
        vacation_account = calculation_service.get_vacation_account(db, user, year)

        bal_color = '#006400' if monthly_balance > 0 else ('#8B0000' if monthly_balance < 0 else '#1e293b')
        ot_color = '#006400' if overtime_account > 0 else ('#8B0000' if overtime_account < 0 else '#1e293b')

        summary_rows = [
            [Paragraph('Zusammenfassung', ParagraphStyle('st', fontName='Helvetica-Bold', fontSize=8, leading=10)), ''],
            [Paragraph('Soll-Stunden:', s_sum_lbl), Paragraph(f"{float(summary_target):.2f} h", s_sum_val)],
            [Paragraph('Ist-Stunden:', s_sum_lbl), Paragraph(f"{float(summary_actual):.2f} h", s_sum_val)],
            [Paragraph('Saldo Monat:', s_sum_lbl),
             Paragraph(f"{float(monthly_balance):+.2f} h",
                       ParagraphStyle('sb', fontName='Helvetica-Bold', fontSize=7.5,
                                      textColor=colors.HexColor(bal_color)))],
            [Paragraph('\u00dcberstunden kumuliert:', s_sum_lbl),
             Paragraph(f"{float(overtime_account):+.2f} h",
                       ParagraphStyle('so', fontName='Helvetica-Bold', fontSize=7.5,
                                      textColor=colors.HexColor(ot_color)))],
            [Paragraph('Urlaub genommen:', s_sum_lbl),
             Paragraph(f"{float(vacation_account['used_hours']):.2f} h", s_sum_val)],
            [Paragraph('Urlaub Rest:', s_sum_lbl),
             Paragraph(f"{float(vacation_account['remaining_hours']):.2f} h", s_sum_val)],
            [Paragraph('Nachtarbeitstage (\u00a76 ArbZG):', s_sum_lbl),
             Paragraph(str(night_work_count), s_sum_val)],
        ]
        sum_tbl = Table(summary_rows, colWidths=[55 * mm, 35 * mm])
        sum_tbl.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CCE5FF')),
            ('SPAN', (0, 0), (1, 0)),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ]))
        story.append(sum_tbl)

    doc.build(story)
    output.seek(0)
    return output
