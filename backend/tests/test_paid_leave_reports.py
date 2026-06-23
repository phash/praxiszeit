"""Tests für PAID_LEAVE (bezahlte Freistellung) in Reports/Exports (#145 REQ-4).

AbsenceType.PAID_LEAVE darf NICHT stillschweigend in den Auswertungen
verschwinden, sondern muss als eigene Kategorie ("Bezahlte Freistellung")
neben OTHER/"Sonstiges" auftauchen — im JSON-Report sowie im XLSX- und
ODS-Jahresexport.
"""
import zipfile
from datetime import date, time
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import Base, get_db
from app.middleware.auth import get_current_user, require_admin
from app.models import Absence, AbsenceType
from app.services.export_service import generate_yearly_report
from app.services.ods_export_service import (
    generate_yearly_report as generate_yearly_ods,
    generate_yearly_report_classic as generate_yearly_ods_classic,
)
from tests.conftest import DEFAULT_TENANT_ID, engine, TestingSessionLocal


YEAR = 2026


def _mk_absence(db, user, d, typ, hours=8.0):
    a = Absence(
        user_id=user.id,
        tenant_id=DEFAULT_TENANT_ID,
        date=d,
        type=typ,
        hours=hours,
    )
    db.add(a)
    db.commit()
    return a


# ---------------------------------------------------------------------------
# JSON report endpoint  (GET /api/admin/reports/yearly-absences)
# ---------------------------------------------------------------------------

def _reports_app() -> FastAPI:
    from app.routers import reports as reports_router

    app = FastAPI(title="PraxisZeit PaidLeave Report Test")

    from app.core.limiter import limiter
    from slowapi import _rate_limit_exceeded_handler
    from slowapi.errors import RateLimitExceeded

    limiter.enabled = False
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    app.include_router(reports_router.router)
    return app


_app = _reports_app()


def test_yearly_absences_endpoint_includes_paid_leave(db, test_user, test_admin):
    """Der JSON-Report muss paid_leave_days als eigene Kategorie liefern und in
    total_days einrechnen — nicht stillschweigend verwerfen."""
    # 2 Tage Urlaub + 3 Tage bezahlte Freistellung (PAID_LEAVE)
    _mk_absence(db, test_user, date(YEAR, 3, 2), AbsenceType.VACATION, 8.0)
    _mk_absence(db, test_user, date(YEAR, 3, 3), AbsenceType.VACATION, 8.0)
    _mk_absence(db, test_user, date(YEAR, 3, 4), AbsenceType.PAID_LEAVE, 8.0)
    _mk_absence(db, test_user, date(YEAR, 3, 5), AbsenceType.PAID_LEAVE, 8.0)
    _mk_absence(db, test_user, date(YEAR, 3, 6), AbsenceType.PAID_LEAVE, 8.0)

    def override_db():
        yield db

    _app.dependency_overrides[get_db] = override_db
    _app.dependency_overrides[get_current_user] = lambda: test_admin
    _app.dependency_overrides[require_admin] = lambda: test_admin
    try:
        client = TestClient(_app)
        resp = client.get(f"/api/admin/reports/yearly-absences?year={YEAR}")
        assert resp.status_code == 200, resp.text
        rows = resp.json()
    finally:
        _app.dependency_overrides.clear()

    row = next(r for r in rows if r["user_id"] == str(test_user.id))
    # daily_target = 40/5 = 8h → 3 PAID_LEAVE-Tage à 8h = 3.0 Tage
    assert row["paid_leave_days"] == pytest.approx(3.0)
    assert row["vacation_days"] == pytest.approx(2.0)
    # total_days enthält Urlaub + bezahlte Freistellung
    assert row["total_days"] == pytest.approx(5.0)


def test_monthly_report_reports_vacation_and_sick_in_days(db, test_user, test_admin):
    """Bug-Fix: die Monatsübersicht muss Urlaub/Krank auch in TAGEN liefern (nicht
    nur in Stunden), damit die Anzeige Tage zeigen kann. Krank ist ohne
    include_health_data 0 (DSGVO Art. 9), mit Flag in Tagen sichtbar."""
    _mk_absence(db, test_user, date(YEAR, 6, 1), AbsenceType.VACATION, 8.0)  # 1 Tag Urlaub
    _mk_absence(db, test_user, date(YEAR, 6, 2), AbsenceType.SICK, 8.0)      # 1 Tag krank

    def override_db():
        yield db

    _app.dependency_overrides[get_db] = override_db
    _app.dependency_overrides[get_current_user] = lambda: test_admin
    _app.dependency_overrides[require_admin] = lambda: test_admin
    try:
        client = TestClient(_app)
        r1 = client.get(f"/api/admin/reports/monthly?month={YEAR}-06")
        assert r1.status_code == 200, r1.text
        row1 = next(r for r in r1.json() if r["user_id"] == str(test_user.id))
        assert row1["vacation_used_days"] == pytest.approx(1.0)   # 8h ÷ 8h Tagessoll
        assert row1["sick_days"] == 0.0                            # DSGVO-maskiert ohne Flag
        r2 = client.get(f"/api/admin/reports/monthly?month={YEAR}-06&include_health_data=true")
        row2 = next(r for r in r2.json() if r["user_id"] == str(test_user.id))
        assert row2["sick_days"] == pytest.approx(1.0)             # 8h ÷ 8h, jetzt sichtbar
    finally:
        _app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# XLSX yearly export — "Bez. Freistellung" column on the absences overview
# ---------------------------------------------------------------------------

def test_xlsx_yearly_export_has_paid_leave_column_and_value(db, test_user):
    _mk_absence(db, test_user, date(YEAR, 4, 1), AbsenceType.PAID_LEAVE, 8.0)
    _mk_absence(db, test_user, date(YEAR, 4, 2), AbsenceType.PAID_LEAVE, 8.0)

    out = generate_yearly_report(db, YEAR)
    out.seek(0)
    data = out.read()
    assert zipfile.is_zipfile(BytesIO(data)), "output is not a valid xlsx zip"
    out.seek(0)
    wb = load_workbook(out, read_only=True)
    assert "Abwesenheiten" in wb.sheetnames
    ws = wb["Abwesenheiten"]

    cells = [
        str(c.value) if c.value is not None else ""
        for row in ws.iter_rows(values_only=False)
        for c in row
    ]
    # Header column for the new own category must be present
    assert any("Bez. Freistellung" in c for c in cells), (
        f"PAID_LEAVE column header missing: {cells}"
    )
    # The 2.0-days value (2 * 8h / 8h daily target) must be rendered
    assert any(c in ("2.0", "2") for c in cells), (
        f"expected paid-leave days value 2.0 in sheet, got: {cells}"
    )


# ---------------------------------------------------------------------------
# ODS yearly export — "Bez. Freistellung" present in both overview + monthly
# ---------------------------------------------------------------------------

def test_ods_yearly_export_has_paid_leave_overview_header(db, test_user):
    _mk_absence(db, test_user, date(YEAR, 5, 4), AbsenceType.PAID_LEAVE, 8.0)

    out = generate_yearly_ods(db, YEAR)
    out.seek(0)
    data = out.read()
    with zipfile.ZipFile(BytesIO(data)) as zf:
        content = zf.read("content.xml").decode("utf-8", errors="replace")

    # Yearly absences-overview sheet header
    assert "Bez. Freistellung (Tage)" in content, "ODS overview missing paid-leave column"


def test_ods_classic_yearly_export_has_paid_leave_monthly_header(db, test_user):
    """Die klassische Jahresübersicht (12 Monatszeilen pro MA) muss die
    Bez.-Freistellung-Spalte in Stunden führen."""
    _mk_absence(db, test_user, date(YEAR, 5, 4), AbsenceType.PAID_LEAVE, 8.0)

    out = generate_yearly_ods_classic(db, YEAR)
    out.seek(0)
    data = out.read()
    with zipfile.ZipFile(BytesIO(data)) as zf:
        content = zf.read("content.xml").decode("utf-8", errors="replace")

    # Per-employee monthly-rows sheet header
    assert "Bez. Freistellung (Std)" in content, "ODS monthly sheet missing paid-leave column"


# ---------------------------------------------------------------------------
# §3 24-week-average (M-ARB2, Review 2026-05-29): scheduled_work_days must be
# computed per actual working day (honouring holidays + absences + contract
# history), not a flat 24 × work_days_per_week.
# ---------------------------------------------------------------------------

def test_24_week_average_excludes_holidays_and_absences(db, test_user, test_admin):
    from app.models.public_holiday import PublicHoliday

    end = date(2026, 3, 31)

    def override_db():
        yield db

    _app.dependency_overrides[get_db] = override_db
    _app.dependency_overrides[get_current_user] = lambda: test_admin
    _app.dependency_overrides[require_admin] = lambda: test_admin
    try:
        client = TestClient(_app)
        base = client.get(f"/api/admin/reports/24-week-average?end_date={end.isoformat()}")
        assert base.status_code == 200, base.text
        base_row = next(r for r in base.json()["employees"] if r["user_id"] == str(test_user.id))
        base_days = base_row["scheduled_work_days"]
        assert base_days > 0

        # One public holiday + one full-day VACATION, both on distinct weekdays
        # inside the window (2026-02-03 Tue, 2026-02-04 Wed).
        db.add(PublicHoliday(date=date(2026, 2, 3), name="Testtag", year=2026, tenant_id=DEFAULT_TENANT_ID))
        _mk_absence(db, test_user, date(2026, 2, 4), AbsenceType.VACATION, 8.0)
        db.commit()

        after = client.get(f"/api/admin/reports/24-week-average?end_date={end.isoformat()}")
        after_row = next(r for r in after.json()["employees"] if r["user_id"] == str(test_user.id))
        # Holiday + vacation must each remove one scheduled working day.
        assert after_row["scheduled_work_days"] == base_days - 2
    finally:
        _app.dependency_overrides.clear()


def test_monthly_report_includes_exempt_flag(db, test_user, test_admin):
    """#159: jede Report-Zeile trägt exempt_from_arbzg, damit das Dashboard
    leitende MA aus dem Ø-Saldo filtern kann."""
    import uuid as _uuid
    from app.models import User, UserRole
    from app.services import auth_service as _auth

    boss = User(
        id=_uuid.uuid4(), username="boss", email="boss@t.local",
        password_hash=_auth.hash_password("Test2025!Password"),
        first_name="Boss", last_name="Chef", role=UserRole.EMPLOYEE,
        weekly_hours=40.0, vacation_days=30, work_days_per_week=5,
        is_active=True, tenant_id=DEFAULT_TENANT_ID, exempt_from_arbzg=True,
    )
    db.add(boss)
    db.commit()

    def override_db():
        yield db

    _app.dependency_overrides[get_db] = override_db
    _app.dependency_overrides[get_current_user] = lambda: test_admin
    _app.dependency_overrides[require_admin] = lambda: test_admin
    try:
        client = TestClient(_app)
        resp = client.get(f"/api/admin/reports/monthly?month={YEAR}-03")
        assert resp.status_code == 200, resp.text
        rows = resp.json()
    finally:
        _app.dependency_overrides.clear()

    boss_row = next(r for r in rows if r["user_id"] == str(boss.id))
    assert boss_row["exempt_from_arbzg"] is True
    non_exempt = [r for r in rows if r["user_id"] != str(boss.id)]
    assert non_exempt and all(r["exempt_from_arbzg"] is False for r in non_exempt)
