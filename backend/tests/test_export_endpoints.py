"""Die sieben Datei-Ausleitungen unter ``/api/admin/reports/export*``.

Warum diese Datei existiert
===========================
Diese sieben Endpunkte liefern die **§16-Belege** aus — die Dateien, die bei
einer Pruefung durch das Gewerbeaufsichtsamt oder den Zoll vorgelegt werden —
und drei von ihnen (``/export-yearly``, ``/export-yearly-ods``,
``/export-yearly-classic-ods``) hatten bis hierher auf **keiner** Ebene einen
Test. Der einzige scheinbare Treffer der Suite
(``test_license_readonly_middleware.py``, ``/api/reports/export``) laeuft wegen
eines falschen Praefixes in einen 404, den seine Zusicherung erfuellt — ein
toter Pfad. Die E2E-Suite laedt vier der Dateien herunter, prueft aber nur den
vorgeschlagenen Dateinamen und setzt den Gesundheitsdaten-Haken nie.

Damit lag die gesamte **Nachweiskette der Datei-Herausgabe** blank:

* Jeder der sieben Endpunkte schreibt bei ``include_health_data=true`` eine
  ``TimeEntryAuditLog``-Zeile mit ``action="health_export"`` — der Art.-5(2)-
  Nachweis darueber, wer wann Gesundheitsdaten (Art. 9 DSGVO) ausgeleitet hat.
  Die sieben Bloecke sind Kopien voneinander; genau diese Bauform hat dieses
  Projekt wiederholt an einer Stelle gepflegt und an der Nachbarstelle
  verloren. ``health_export`` kommt repoweit nur in diesen sieben Bloecken vor.
* Die Rollenpruefung (``require_admin``) und die Mandantentrennung (F-026)
  entscheiden, wer fremde Arbeitszeitnachweise als Datei mitnehmen kann.
* Die Maskierung nach Art. 9 (Krankheits-Freitext, Krank-Spalte) ist der
  Unterschied zwischen einem zulaessigen und einem unzulaessigen Beleg.

Der Decimal-Fall
================
``TestDecimalAusDerDatenbank`` ist der wichtigste Test der Datei. Zweimal
(#383 ``vacation_days``, #408 ``weekly_hours``) wurde eine Integer-Spalte auf
``Numeric`` umgestellt; SQLAlchemy liefert fuer ``Numeric`` beim Lesen
``Decimal``, und in jedem Pfad, der roh serialisiert oder rechnet statt ueber
FastAPIs ``jsonable_encoder`` zu gehen, riss das einen **HTTP 500 fuer jeden
Nutzer** auf. Die Datei-Exporte sind genau solche Pfade: sie rechnen die
Summenzeilen selbst und mischen dabei ``float`` und ``Decimal``.

Der Test haelt deshalb zuerst fest, dass die Vorrichtung ueberhaupt ``Decimal``
liefert (SQLite tut das fuer ``Numeric`` genau wie PostgreSQL, sobald die
Objekte frisch aus der Datenbank kommen) — sonst verkaeme er still zu einem
weiteren Statuscode-Test — und faehrt danach alle sieben Ausleitungen an.

Wo diese Tests laufen
=====================
SQLite, also im normalen ``pytest tests/``-Lauf (Schritt 1 von
``scripts/local-ci.sh``) und im SQLite-Schritt der CI.
"""
from __future__ import annotations

import base64
import io
import re
import uuid
import zipfile
import zlib
from datetime import date, time, timedelta
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text

from app.database import Base, get_db
from app.middleware.auth import get_current_user
from app.models import (
    Absence,
    AbsenceType,
    TimeEntry,
    TimeEntryAuditLog,
    User,
    UserRole,
)
from app.models.tenant import Tenant
from app.services import auth_service
from tests.conftest import TestingSessionLocal, engine
from tests.test_endpoints import test_app

TENANT_A_ID = uuid.UUID("11111111-0000-4000-8000-00000000000a")
TENANT_B_ID = uuid.UUID("22222222-0000-4000-8000-00000000000b")

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ODS_MIME = "application/vnd.oasis.opendocument.spreadsheet"
PDF_MIME = "application/pdf"

# Der Berichtsmonat. Bewusst fest verdrahtet (nicht "heute"), damit die
# Zusicherungen ueber Soll/Ist/Maskierung nicht vom Laufdatum abhaengen.
YEAR = 2026
MONTH = 3

# Freitext-Marker. Sie stehen in ``Absence.note`` bzw. im Namen der Person und
# sind so gewaehlt, dass sie in KEINER Beschriftung des Exports vorkommen
# koennen — die Maskierungs-Zusicherungen haengen daran.
SICK_NOTE_A = "MARKERKRANKHEITAAA"
EMPLOYEE_A_LAST = "Ausleitungstest"
EMPLOYEE_B_LAST = "Fremdmandant"
SICK_NOTE_B = "MARKERKRANKHEITBBB"


# ---------------------------------------------------------------------------
# Format-Leser: aus jeder ausgelieferten Datei den enthaltenen Text gewinnen
# ---------------------------------------------------------------------------

def _xlsx_text(blob: bytes) -> str:
    """Alle Zellwerte + Blattnamen einer XLSX-Datei als Text.

    Gleichzeitig der Beweis, dass die Datei ueberhaupt eine gueltige
    Arbeitsmappe ist — ``load_workbook`` wirft sonst.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    parts: list[str] = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def _ods_text(blob: bytes) -> str:
    """Der ``content.xml``-Rumpf einer ODS-Datei.

    ODS ist ein ZIP; ist die Datei kaputt, wirft ``ZipFile`` bzw. das fehlende
    Mitglied ``content.xml``.
    """
    with zipfile.ZipFile(io.BytesIO(blob)) as archive:
        return archive.read("content.xml").decode("utf-8", "replace")


def _pdf_text(blob: bytes) -> str:
    """Der Text aus den Inhaltsstroemen einer PDF-Datei.

    ReportLab schreibt die Seiten als ``ASCII85Decode``+``FlateDecode``. Im
    Container ist keine PDF-Bibliothek installiert (und fuer eine Zusicherung
    ueber vorhandenen/fehlenden Text braucht es auch keine): die Stroeme werden
    direkt dekodiert. Faellt beides aus, bleibt der Rohbytes-Strom uebrig —
    dann greift die Zusicherung immer noch, nur unschaerfer.
    """
    parts: list[str] = []
    for match in re.finditer(rb"stream\r?\n(.*?)endstream", blob, re.S):
        chunk = match.group(1).strip()
        decoded = chunk
        for decoder in (
            lambda x: zlib.decompress(base64.a85decode(x, adobe=True)),
            zlib.decompress,
        ):
            try:
                decoded = decoder(chunk)
                break
            except Exception:  # noqa: BLE001 — naechster Dekodierversuch
                continue
        parts.append(decoded.decode("latin-1", "replace"))
    return "\n".join(parts)


_READERS = {"xlsx": _xlsx_text, "ods": _ods_text, "pdf": _pdf_text}


def _enthaelt(content: str, needle: str) -> bool:
    """``needle in content`` — als BOOL, bevor daraus eine Zusicherung wird.

    Kein Schoenheitsfehler: schriebe man ``assert needle in content`` direkt,
    baut pytest bei einem Fehlschlag seine Erklaerung ueber BEIDE Operanden —
    und der Textrumpf eines ODS-Jahresberichts hat ~360 KB. Die Gegenprobe zu
    diesen Zusicherungen brauchte damit **35 Minuten**, um rot zu werden,
    statt Sekunden. Ueber den Bool sind die Operanden ``True``/``False``, die
    Meldung traegt den Grund.
    """
    return needle in content


# ---------------------------------------------------------------------------
# Die sieben Ausleitungen
# ---------------------------------------------------------------------------
# (Kennung, Pfad, Parameter, MIME-Typ, Dateiname, Format, zeigt-Freitext)
#
# ``zeigt_freitext`` unterscheidet die beiden klassischen Jahresberichte: sie
# stellen Monate als Spalten dar und geben Abwesenheits-Notizen ueberhaupt
# nicht wieder. Ihre Art.-9-Maskierung sitzt statt dessen in der Zeile
# "minus Krank Std." und wird von ``TestKlassischerJahresbericht`` geprueft.
EXPORT_ROUTES = [
    ("xlsx-monat", "/api/admin/reports/export", {"month": f"{YEAR}-{MONTH:02d}"},
     XLSX_MIME, f"PraxisZeit_Monatsreport_{YEAR}_{MONTH:02d}.xlsx", "xlsx", True),
    ("xlsx-jahr", "/api/admin/reports/export-yearly", {"year": YEAR},
     XLSX_MIME, f"PraxisZeit_Jahresreport_{YEAR}.xlsx", "xlsx", True),
    ("xlsx-jahr-klassisch", "/api/admin/reports/export-yearly-classic", {"year": YEAR},
     XLSX_MIME, f"PraxisZeit_Jahresreport_Classic_{YEAR}.xlsx", "xlsx", False),
    ("ods-monat", "/api/admin/reports/export-ods", {"month": f"{YEAR}-{MONTH:02d}"},
     ODS_MIME, f"PraxisZeit_Monatsreport_{YEAR}_{MONTH:02d}.ods", "ods", True),
    ("pdf-monat", "/api/admin/reports/export-pdf", {"month": f"{YEAR}-{MONTH:02d}"},
     PDF_MIME, f"PraxisZeit_Monatsreport_{YEAR}_{MONTH:02d}.pdf", "pdf", True),
    ("ods-jahr", "/api/admin/reports/export-yearly-ods", {"year": YEAR},
     ODS_MIME, f"PraxisZeit_Jahresreport_{YEAR}.ods", "ods", True),
    ("ods-jahr-klassisch", "/api/admin/reports/export-yearly-classic-ods", {"year": YEAR},
     ODS_MIME, f"PraxisZeit_Jahresreport_Classic_{YEAR}.ods", "ods", False),
]

MONTH_ROUTES = [r for r in EXPORT_ROUTES if "month" in r[2]]

_IDS = [r[0] for r in EXPORT_ROUTES]


def _params(route, **extra):
    merged = dict(route[2])
    merged.update(extra)
    return merged


# ---------------------------------------------------------------------------
# Vorrichtung
# ---------------------------------------------------------------------------

@pytest.fixture(scope="function")
def _db_session():
    with engine.connect() as conn:
        conn.execute(text("PRAGMA foreign_keys = OFF"))
        conn.commit()
    Base.metadata.drop_all(bind=engine, checkfirst=True)
    Base.metadata.create_all(bind=engine)
    session = TestingSessionLocal()
    try:
        yield session
    finally:
        session.close()
        with engine.connect() as conn:
            conn.execute(text("PRAGMA foreign_keys = OFF"))
            conn.commit()
        Base.metadata.drop_all(bind=engine, checkfirst=True)


def _make_user(db, tenant_id, *, role, username, last_name):
    user = User(
        username=username,
        email=f"{username}@export.test",
        password_hash=auth_service.hash_password("Export2026!Test"),
        first_name="Vorname",
        last_name=last_name,
        role=role,
        weekly_hours=40.0,
        vacation_days=30,
        work_days_per_week=5,
        is_active=True,
        tenant_id=tenant_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def _seed_month(db, user, tenant_id, sick_note):
    """Ein Arbeitstag, ein Krankentag, ein Urlaubstag im Berichtsmonat.

    Der Krankentag traegt den Freitext-Marker — er ist der Traeger jeder
    Art.-9-Zusicherung dieser Datei.
    """
    first = date(YEAR, MONTH, 1)
    monday = first + timedelta(days=(7 - first.weekday()) % 7)  # erster Montag
    db.add(TimeEntry(
        user_id=user.id, date=monday, start_time=time(8, 0), end_time=time(16, 30),
        break_minutes=30, tenant_id=tenant_id,
    ))
    db.add(Absence(
        user_id=user.id, date=monday + timedelta(days=1), type=AbsenceType.SICK,
        hours=8.0, note=sick_note, tenant_id=tenant_id,
    ))
    db.add(Absence(
        user_id=user.id, date=monday + timedelta(days=2), type=AbsenceType.VACATION,
        hours=8.0, tenant_id=tenant_id,
    ))
    db.commit()


@pytest.fixture(scope="function")
def welt(_db_session):
    """Zwei Mandanten mit je einem Admin und einem Mitarbeiter samt Daten.

    Rueckgabe sind einfache Werte (IDs, Namen), KEINE ORM-Objekte: die
    Endpunkte rufen ``db.close()`` vor dem Ausliefern des Datenstroms (F-053),
    was alle Objekte der geteilten Sitzung abhaengt.
    """
    for tid, name in ((TENANT_A_ID, "Mandant A"), (TENANT_B_ID, "Mandant B")):
        _db_session.add(Tenant(
            id=tid, name=name, slug=f"t-{tid.hex[:8]}", is_active=True, mode="multi",
        ))
    _db_session.commit()

    admin_a = _make_user(_db_session, TENANT_A_ID, role=UserRole.ADMIN,
                         username="export_admin_a", last_name="AdminA")
    employee_a = _make_user(_db_session, TENANT_A_ID, role=UserRole.EMPLOYEE,
                            username="export_employee_a", last_name=EMPLOYEE_A_LAST)
    admin_b = _make_user(_db_session, TENANT_B_ID, role=UserRole.ADMIN,
                         username="export_admin_b", last_name="AdminB")
    employee_b = _make_user(_db_session, TENANT_B_ID, role=UserRole.EMPLOYEE,
                            username="export_employee_b", last_name=EMPLOYEE_B_LAST)

    _seed_month(_db_session, employee_a, TENANT_A_ID, SICK_NOTE_A)
    _seed_month(_db_session, employee_b, TENANT_B_ID, SICK_NOTE_B)

    return {
        "admin_a_id": admin_a.id,
        "admin_a_username": admin_a.username,
        "employee_a_id": employee_a.id,
        "admin_b_id": admin_b.id,
    }


def _client_as(db_session, user_id):
    """Client, der als ``user_id`` angemeldet ist.

    Ueberschrieben wird ausschliesslich ``get_current_user`` — ``require_admin``
    laeuft ECHT. Sonst wuerde die Rollenpruefung der sieben Endpunkte
    wegmodelliert statt geprueft.
    """
    from app.core.limiter import limiter

    user = db_session.query(User).filter(User.id == user_id).first()

    def _override_db():
        yield db_session

    def _override_user():
        return user

    test_app.dependency_overrides[get_db] = _override_db
    test_app.dependency_overrides[get_current_user] = _override_user
    # Die Endpunkte tragen ``@limiter.limit("20/minute")``. Der Limiter ist ein
    # Modul-Singleton; eine andere Datei der Suite kann ihn eingeschaltet
    # zuruecklassen (test_real_app_middleware stellt ihn bewusst scharf). Diese
    # Datei feuert mehr als 20 Anfragen — also hier explizit aus und danach
    # exakt auf den vorgefundenen Wert zurueck.
    saved = limiter.enabled
    limiter.enabled = False
    client = TestClient(test_app)
    client.__enter__()
    return client, saved


@pytest.fixture(scope="function")
def admin_client(_db_session, welt):
    client, saved = _client_as(_db_session, welt["admin_a_id"])
    try:
        yield client
    finally:
        from app.core.limiter import limiter
        limiter.enabled = saved
        client.__exit__(None, None, None)
        test_app.dependency_overrides.clear()


@pytest.fixture(scope="function")
def admin_b_client(_db_session, welt):
    client, saved = _client_as(_db_session, welt["admin_b_id"])
    try:
        yield client
    finally:
        from app.core.limiter import limiter
        limiter.enabled = saved
        client.__exit__(None, None, None)
        test_app.dependency_overrides.clear()


@pytest.fixture(scope="function")
def employee_client(_db_session, welt):
    client, saved = _client_as(_db_session, welt["employee_a_id"])
    try:
        yield client
    finally:
        from app.core.limiter import limiter
        limiter.enabled = saved
        client.__exit__(None, None, None)
        test_app.dependency_overrides.clear()


@pytest.fixture(scope="function")
def anonymous_client(_db_session, welt):
    from app.core.limiter import limiter

    def _override_db():
        yield _db_session

    test_app.dependency_overrides[get_db] = _override_db
    saved = limiter.enabled
    limiter.enabled = False
    with TestClient(test_app) as client:
        yield client
    limiter.enabled = saved
    test_app.dependency_overrides.clear()


def _health_export_rows(db):
    return db.query(TimeEntryAuditLog).filter(
        TimeEntryAuditLog.action == "health_export"
    ).all()


# ---------------------------------------------------------------------------
# 1. Die Datei kommt an — Statuscode, Inhaltstyp, Dateiname, Lesbarkeit
# ---------------------------------------------------------------------------

class TestAusleitungLiefertDatei:

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_liefert_lesbare_datei(self, admin_client, route):
        """200, korrekter Inhaltstyp, korrekter Dateiname, parsbare Datei."""
        _, path, params, mime, filename, fmt, _ = route
        resp = admin_client.get(path, params=params)

        assert resp.status_code == 200, resp.text[:300]
        assert resp.headers["content-type"].split(";")[0] == mime
        disposition = resp.headers["content-disposition"]
        assert "attachment" in disposition
        assert f'filename="{filename}"' in disposition
        # RFC 5987: der Dateiname wird ZUSAETZLICH prozentkodiert mitgegeben —
        # ohne ihn zeigen Browser bei Umlauten/Sonderzeichen Muell an.
        assert "filename*=UTF-8''" in disposition

        blob = resp.content
        assert len(blob) > 1000, "verdaechtig kleine Datei"
        if fmt == "pdf":
            assert blob.startswith(b"%PDF-")
        else:
            assert blob[:2] == b"PK", "kein ZIP-Container (xlsx/ods)"
        # Wirft, wenn die Datei nicht lesbar ist.
        assert bool(_READERS[fmt](blob)), "Datei enthaelt keinen lesbaren Text"

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_datei_enthaelt_den_mitarbeiter(self, admin_client, route):
        """Der Beleg ist nicht leer: der Mitarbeiter des Mandanten steht drin.

        Ohne diese Zusicherung koennte jeder folgende Negativtest ("Fremder
        nicht enthalten", "Freitext maskiert") schon durch eine LEERE Datei
        erfuellt werden.
        """
        _, path, params, _, _, fmt, _ = route
        resp = admin_client.get(path, params=params)
        assert resp.status_code == 200
        content = _READERS[fmt](resp.content)
        assert _enthaelt(content, EMPLOYEE_A_LAST), \
            f"{EMPLOYEE_A_LAST} fehlt in der Datei"

    @pytest.mark.parametrize("route", MONTH_ROUTES,
                             ids=[r[0] for r in MONTH_ROUTES])
    def test_ungueltiges_monatsformat_400(self, admin_client, route):
        _, path, params, _, _, _, _ = route
        resp = admin_client.get(path, params=_params(route, month="2026/03"))
        assert resp.status_code == 400
        assert "Monatsformat" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# 2. Wer darf die Datei ueberhaupt holen
# ---------------------------------------------------------------------------

class TestZugriffsschutz:

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_mitarbeiter_bekommt_403(self, employee_client, route):
        """``require_admin``: ein Mitarbeiter darf die Arbeitszeitnachweise der
        gesamten Belegschaft nicht als Datei mitnehmen."""
        _, path, params, _, _, _, _ = route
        resp = employee_client.get(path, params=params)
        assert resp.status_code == 403
        assert "Admin" in resp.json()["detail"]

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_ohne_anmeldung_kein_zugriff(self, anonymous_client, route):
        _, path, params, _, _, _, _ = route
        resp = anonymous_client.get(path, params=params)
        assert resp.status_code in (401, 403)

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_fremder_mandant_ist_nicht_enthalten(self, admin_client, route):
        """F-026: der Admin von Mandant A darf in seiner Datei niemanden aus
        Mandant B finden — weder Namen noch Freitext."""
        _, path, params, _, _, fmt, _ = route
        resp = admin_client.get(path, params=_params(route, include_health_data="true"))
        assert resp.status_code == 200
        content = _READERS[fmt](resp.content)
        assert _enthaelt(content, EMPLOYEE_A_LAST), "eigener Mitarbeiter fehlt"
        assert not _enthaelt(content, EMPLOYEE_B_LAST), \
            "FREMDER Mitarbeiter in der Datei"
        assert not _enthaelt(content, SICK_NOTE_B), \
            "FREMDE Krankheitsnotiz in der Datei"

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_umgekehrte_richtung_ebenfalls_dicht(self, admin_b_client, route):
        """Die Gegenrichtung — sonst bliebe ein Filter unentdeckt, der
        versehentlich auf einen festen Mandanten verdrahtet ist."""
        _, path, params, _, _, fmt, _ = route
        resp = admin_b_client.get(path, params=_params(route, include_health_data="true"))
        assert resp.status_code == 200
        content = _READERS[fmt](resp.content)
        assert _enthaelt(content, EMPLOYEE_B_LAST), "eigener Mitarbeiter fehlt"
        assert not _enthaelt(content, EMPLOYEE_A_LAST), \
            "FREMDER Mitarbeiter in der Datei"
        assert not _enthaelt(content, SICK_NOTE_A), \
            "FREMDE Krankheitsnotiz in der Datei"


# ---------------------------------------------------------------------------
# 3. Art. 9 DSGVO — Maskierung der Gesundheitsdaten
# ---------------------------------------------------------------------------

class TestGesundheitsdatenMaskierung:

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_ohne_flag_kein_krankheits_freitext(self, admin_client, route):
        """Standardfall (kein Parameter): der Freitext zur Krankmeldung darf in
        KEINEM der sieben Formate erscheinen."""
        _, path, params, _, _, fmt, _ = route
        resp = admin_client.get(path, params=params)
        assert resp.status_code == 200
        assert not _enthaelt(_READERS[fmt](resp.content), SICK_NOTE_A), \
            "Krankheits-Freitext trotz fehlendem Art.-9-Haken in der Datei"

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_flag_false_wie_kein_flag(self, admin_client, route):
        _, path, params, _, _, fmt, _ = route
        resp = admin_client.get(path, params=_params(route, include_health_data="false"))
        assert resp.status_code == 200
        assert not _enthaelt(_READERS[fmt](resp.content), SICK_NOTE_A), \
            "Krankheits-Freitext trotz include_health_data=false in der Datei"

    @pytest.mark.parametrize(
        "route", [r for r in EXPORT_ROUTES if r[6]],
        ids=[r[0] for r in EXPORT_ROUTES if r[6]],
    )
    def test_mit_flag_erscheint_der_freitext(self, admin_client, route):
        """Gegenprobe zur Maskierung: mit ausdruecklichem Haken IST der Freitext
        drin. Ohne diesen Test koennte die Maskierung auch dadurch "gruen" sein,
        dass der Export den Freitext generell nie ausgibt."""
        _, path, params, _, _, fmt, _ = route
        resp = admin_client.get(path, params=_params(route, include_health_data="true"))
        assert resp.status_code == 200
        assert _enthaelt(_READERS[fmt](resp.content), SICK_NOTE_A), \
            "Freitext fehlt trotz gesetztem Art.-9-Haken"


class TestKlassischerJahresbericht:
    """Die beiden klassischen Jahresberichte zeigen keine Notizen — ihre
    Art.-9-Maskierung sitzt in der Zeile "minus Krank Std.": ohne Haken steht
    dort ein Gedankenstrich und die Stunden werden der Arbeitsseite
    zugeschlagen, mit Haken stehen die Krankstunden als Zahl."""

    ROW_LABEL = "minus Krank"
    MASK = "–"  # Gedankenstrich

    def _xlsx_krank_row(self, blob: bytes) -> list:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(blob))
        # Der klassische Bericht legt je Mitarbeiter ein Blatt an, benannt nach
        # dem Nachnamen — und der Admin des Mandanten bekommt ebenfalls eines.
        # Blatt 0 waere seines (alphabetisch vor dem Mitarbeiter) und stuende
        # ohne jede Abwesenheit da: die Zusicherung waere still wertlos.
        titles = [t for t in wb.sheetnames if EMPLOYEE_A_LAST in t]
        assert titles, wb.sheetnames
        ws = wb[titles[0]]
        for row in ws.iter_rows():
            if row[0].value and self.ROW_LABEL in str(row[0].value):
                return [c.value for c in row[2:14]]
        raise AssertionError(f"Zeile '{self.ROW_LABEL}' nicht gefunden")

    def test_xlsx_ohne_flag_maskiert_die_krankspalte(self, admin_client):
        resp = admin_client.get("/api/admin/reports/export-yearly-classic",
                                params={"year": YEAR})
        assert resp.status_code == 200
        values = self._xlsx_krank_row(resp.content)
        assert values, "Krank-Zeile ist leer"
        assert all(v in (None, self.MASK) for v in values), values

    def test_xlsx_mit_flag_zeigt_die_krankstunden(self, admin_client):
        resp = admin_client.get("/api/admin/reports/export-yearly-classic",
                                params={"year": YEAR, "include_health_data": "true"})
        assert resp.status_code == 200
        values = self._xlsx_krank_row(resp.content)
        numeric = [v for v in values if isinstance(v, (int, float))]
        assert numeric, values
        assert max(numeric) > 0, "Krankstunden fehlen trotz gesetztem Haken"

    def test_ods_maskierung_unterscheidet_sich_vom_flag_lauf(self, admin_client):
        """ODS-Zwilling. Der Inhalt MUSS sich zwischen maskiert und unmaskiert
        unterscheiden — waeren beide Laeufe gleich, ginge der Haken ins Leere."""
        masked = admin_client.get("/api/admin/reports/export-yearly-classic-ods",
                                  params={"year": YEAR})
        unmasked = admin_client.get("/api/admin/reports/export-yearly-classic-ods",
                                    params={"year": YEAR, "include_health_data": "true"})
        assert masked.status_code == unmasked.status_code == 200
        masked_text = _ods_text(masked.content)
        unmasked_text = _ods_text(unmasked.content)
        assert bool(masked_text != unmasked_text), \
            "maskierter und unmaskierter Lauf liefern dieselbe Datei"
        assert _enthaelt(masked_text, self.MASK), "Maskierungszeichen fehlt"


# ---------------------------------------------------------------------------
# 4. Art. 5(2) DSGVO — die Nachweiszeile je Ausleitung
# ---------------------------------------------------------------------------

class TestNachweiszeile:
    """``health_export`` existiert repoweit nur als sieben Kopien desselben
    Blocks in ``reports.py``. Jede einzelne wird hier festgenagelt."""

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_mit_flag_genau_eine_nachweiszeile(self, admin_client, _db_session, welt, route):
        _, path, params, _, _, _, _ = route
        assert _health_export_rows(_db_session) == []

        resp = admin_client.get(path, params=_params(route, include_health_data="true"))
        assert resp.status_code == 200

        rows = _health_export_rows(_db_session)
        assert len(rows) == 1, f"{len(rows)} Nachweiszeilen statt genau einer"
        row = rows[0]
        assert row.source == "dsgvo"
        assert row.user_id == welt["admin_a_id"]
        assert row.changed_by == welt["admin_a_id"]
        assert row.tenant_id == TENANT_A_ID
        assert row.time_entry_id is None
        # Der Nachweis muss sagen WER ausgeleitet hat — sonst belegt er nichts.
        assert welt["admin_a_username"] in (row.new_note or "")
        assert "Art. 9" in (row.new_note or "")

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_ohne_flag_keine_nachweiszeile(self, admin_client, _db_session, route):
        """Kein Gesundheitsdaten-Export ⇒ kein Eintrag. Ein Protokoll, das bei
        JEDEM Abruf schreibt, verwaessert den Nachweis."""
        _, path, params, _, _, _, _ = route
        resp = admin_client.get(path, params=params)
        assert resp.status_code == 200
        assert _health_export_rows(_db_session) == []

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    def test_403_schreibt_keine_nachweiszeile(self, employee_client, _db_session, route):
        """Ein abgewiesener Zugriff darf das Protokoll nicht fuellen."""
        _, path, params, _, _, _, _ = route
        resp = employee_client.get(path, params=_params(route, include_health_data="true"))
        assert resp.status_code == 403
        assert _health_export_rows(_db_session) == []

    def test_jeder_abruf_schreibt_eine_eigene_zeile(self, admin_client, _db_session):
        """Zwei Ausleitungen ⇒ zwei Nachweise (kein Dedup, kein Upsert)."""
        for _ in range(2):
            resp = admin_client.get("/api/admin/reports/export",
                                    params={"month": f"{YEAR}-{MONTH:02d}",
                                            "include_health_data": "true"})
            assert resp.status_code == 200
        assert len(_health_export_rows(_db_session)) == 2

    def test_nachweiszeile_traegt_den_zeitraum(self, admin_client, _db_session):
        """Ohne Zeitraum im Text laesst sich nicht sagen, WELCHE Daten den
        Betrieb verlassen haben."""
        admin_client.get("/api/admin/reports/export-yearly",
                         params={"year": YEAR, "include_health_data": "true"})
        rows = _health_export_rows(_db_session)
        assert len(rows) == 1
        assert str(YEAR) in rows[0].new_note


# ---------------------------------------------------------------------------
# 5. Der Decimal-Fall (#383 / #408)
# ---------------------------------------------------------------------------

class TestDecimalAusDerDatenbank:

    def test_vorrichtung_liefert_wirklich_decimal(self, _db_session, welt):
        """Selbstkontrolle. Faellt sie, ist der Test darunter wertlos: er wuerde
        dann nur noch Statuscodes gegen float-Werte pruefen und genau die
        Fehlerklasse verfehlen, die er abdecken soll.

        ``expire_all`` erzwingt das Nachladen aus der Datenbank — genau das tut
        auch der Endpunkt, der mit einer frischen Sitzung arbeitet.
        """
        _db_session.expire_all()
        user = _db_session.query(User).filter(
            User.id == welt["employee_a_id"]).first()
        assert isinstance(user.weekly_hours, Decimal)    # #431 Numeric(4,2)
        assert isinstance(user.vacation_days, Decimal)   # #408 Numeric(4,1)

        absence = _db_session.query(Absence).filter(
            Absence.user_id == user.id).first()
        assert isinstance(absence.hours, Decimal)        # Numeric(4,2)

        entry = _db_session.query(TimeEntry).filter(
            TimeEntry.user_id == user.id).first()
        assert isinstance(entry.net_hours, Decimal)

    @pytest.mark.parametrize("route", EXPORT_ROUTES, ids=_IDS)
    @pytest.mark.parametrize("health", [False, True], ids=["ohne-art9", "mit-art9"])
    def test_ausleitung_ueberlebt_decimal_werte(self, admin_client, _db_session,
                                                route, health):
        """Der eigentliche Regressionsschutz gegen #383/#408.

        Die Datei-Exporte rechnen ihre Summenzeilen selbst und mischen dabei
        ``float`` und ``Decimal``. Faellt irgendwo ein ``float(...)``-Cast weg,
        wirft die Ausleitung ``TypeError`` bzw. ``decimal.InvalidOperation`` —
        HTTP 500 fuer JEDEN Nutzer, nicht nur fuer einen Sonderfall. Genau so
        sind #383 und #408 in Produktion aufgeschlagen.
        """
        _db_session.expire_all()
        _, path, params, _, _, fmt, _ = route
        resp = admin_client.get(
            path, params=_params(route, include_health_data=str(health).lower()))
        assert resp.status_code == 200, resp.text[:400]
        assert bool(_READERS[fmt](resp.content)), "Datei nicht lesbar"

    def test_krumme_dezimalwerte_ueberleben_alle_formate(self, admin_client, _db_session,
                                                         welt):
        """#408 im Original: ``vacation_days`` 16,8 statt 30 — ein Wert, der
        ohne Nachkommastelle gar nicht darstellbar ist. Dazu krumme
        Wochenstunden (17,75) und ein halber Abwesenheitstag (3,75 h).
        """
        user = _db_session.query(User).filter(
            User.id == welt["employee_a_id"]).first()
        user.vacation_days = Decimal("16.8")
        user.weekly_hours = Decimal("17.75")
        absence = _db_session.query(Absence).filter(
            Absence.user_id == user.id, Absence.type == AbsenceType.VACATION).first()
        absence.hours = Decimal("3.75")
        absence.half_day = True
        _db_session.commit()
        _db_session.expire_all()

        for kennung, path, params, _, _, fmt, _ in EXPORT_ROUTES:
            _db_session.expire_all()
            resp = admin_client.get(
                path, params={**params, "include_health_data": "true"})
            assert resp.status_code == 200, f"{kennung}: {resp.text[:300]}"
            assert bool(_READERS[fmt](resp.content)), \
                f"{kennung}: Datei nicht lesbar"
