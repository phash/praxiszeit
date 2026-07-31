"""F2 (Release-Gate 1.18.0): §16-Datei-Exporte respektieren das Beschäftigungsfenster.

Fünf Per-Tag-Schleifen (XLSX-Monatsblatt, XLSX-Jahres-Mitarbeiterblatt, PDF,
ODS-Monat, ODS-Jahres-Mitarbeiterblatt) schrieben Soll und Ist ohne
``_within_employment_window``-Prüfung. ``get_daily_target_for_date`` kennt kein
Fenster und liefert vor ``first_work_day`` das volle Tagessoll — im SELBEN Blatt
rechnen „Überstunden kumuliert" (``get_overtime_account``) und im selben Workbook
die Jahresübersicht (Σ ``get_monthly_target``) dagegen fenster-korrekt.

Beispiel: Eintritt 16.03.2026, 40 h/5 Tage ⇒ Datei „Soll-Stunden Monat" 176,00 h
statt 96,00 h, „Saldo Monat" 80 h zu negativ. Die Ist-Seite ist spiegelbildlich
ungefiltert (ein Zeiteintrag VOR dem Eintritt — Import/Rehire/Datumskorrektur —
zählte mit, obwohl ``get_monthly_actual`` ihn seit #195 ausschließt).
"""
import base64
import re
import zlib
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import TimeEntry, User, UserRole
from app.services import calculation_service, export_service, ods_export_service
from tests.conftest import DEFAULT_TENANT_ID

FIRST_WORK_DAY = date(2026, 3, 16)      # Montag
BEFORE = date(2026, 3, 10)              # Dienstag, VOR dem Eintritt
AFTER = date(2026, 3, 17)               # Dienstag, NACH dem Eintritt


@pytest.fixture
def late_starter(db, default_tenant):
    from app.services import auth_service
    u = User(
        username="latestart", email="late@example.com",
        password_hash=auth_service.hash_password("test123"),
        first_name="Spät", last_name="Starter",
        role=UserRole.EMPLOYEE, weekly_hours=40.0, vacation_days=30,
        work_days_per_week=5, is_active=True, tenant_id=DEFAULT_TENANT_ID,
        first_work_day=FIRST_WORK_DAY,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    for d in (BEFORE, AFTER):
        db.add(TimeEntry(user_id=u.id, tenant_id=DEFAULT_TENANT_ID, date=d,
                         start_time=time(8, 0), end_time=time(16, 0), break_minutes=0))
    db.commit()
    return u


def _sheet(bio: BytesIO):
    bio.seek(0)
    wb = load_workbook(bio)
    return wb[wb.sheetnames[0]]


def _summary(sheet, label):
    """Wert der Zusammenfassungszeile mit ``label`` in Spalte A."""
    for row in sheet.iter_rows(min_col=1, max_col=2):
        if row[0].value == label:
            return row[1].value
    raise AssertionError(f"Zusammenfassungszeile {label!r} nicht gefunden")


def _xlsx_day_row(sheet, d: date):
    """Detailzeile eines Datums: (Netto Spalte 6, Soll Spalte 7)."""
    for row in sheet.iter_rows(min_col=1, max_col=7):
        cell = row[0].value
        if isinstance(cell, datetime):
            cell = cell.date()
        if cell == d:
            return row[5].value, row[6].value
    raise AssertionError(f"Tageszeile {d} nicht gefunden")


def _ods_rows(bio: BytesIO):
    """ODS → Liste von Zeilen als Liste von Zell-Strings (erstes Blatt)."""
    from odf.opendocument import load as odf_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    bio.seek(0)
    doc = odf_load(bio)
    table = doc.spreadsheet.getElementsByType(Table)[0]
    rows = []
    for tr in table.getElementsByType(TableRow):
        cells = []
        for tc in tr.getElementsByType(TableCell):
            val = tc.getAttribute("value")
            if val is None:
                ps = tc.getElementsByType(P)
                val = "".join(str(p) for p in ps)
            cells.append(str(val))
        rows.append(cells)
    return rows


def _pdf_text(bio: BytesIO) -> str:
    """Textinhalt eines reportlab-PDF (FlateDecode-Streams entpacken)."""
    bio.seek(0)
    raw = bio.read()
    out = []
    for m in re.finditer(rb"stream\r?\n(.*?)endstream", raw, re.S):
        chunk = m.group(1)
        try:  # reportlab: ASCII85 ueber FlateDecode
            chunk = base64.a85decode(chunk.strip(), adobe=True)
        except Exception:
            pass
        try:
            chunk = zlib.decompress(chunk)
        except zlib.error:
            pass
        out.append(chunk.decode("latin-1", errors="ignore"))
    return "\n".join(out)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def test_xlsx_monthly_summary_matches_calculation(db, late_starter):
    """„Soll-Stunden Monat" muss get_monthly_target entsprechen (96, nicht 176)."""
    expected = calculation_service.get_monthly_target(db, late_starter, 2026, 3)
    assert expected == Decimal('96.00'), expected  # Kontrolle des Szenarios

    sheet = _sheet(export_service.generate_monthly_report(db, 2026, 3))
    assert _summary(sheet, "Soll-Stunden Monat:") == float(expected)


def test_xlsx_monthly_ist_excludes_pre_employment_entry(db, late_starter):
    """Die Ist-Seite ist spiegelbildlich gefenstert wie get_monthly_actual."""
    expected = calculation_service.get_monthly_actual(db, late_starter, 2026, 3)
    assert expected == Decimal('8.00'), expected

    sheet = _sheet(export_service.generate_monthly_report(db, 2026, 3))
    assert _summary(sheet, "Ist-Stunden Monat:") == float(expected)


def test_xlsx_monthly_day_rows_are_windowed(db, late_starter):
    """Detailzeile vor dem Eintritt: Soll 0 UND Ist 0 — sonst widerspricht die
    Summenzeile den Zeilen darüber."""
    sheet = _sheet(export_service.generate_monthly_report(db, 2026, 3))
    net_before, target_before = _xlsx_day_row(sheet, BEFORE)
    assert target_before == 0.00, target_before
    assert net_before == 0.00, net_before
    # Kontrolle: der Tag NACH dem Eintritt bleibt unverändert
    net_after, target_after = _xlsx_day_row(sheet, AFTER)
    assert target_after == 8.00, target_after
    assert net_after == 8.00, net_after


def test_xlsx_yearly_employee_sheet_matches_calculation(db, late_starter):
    """Jahres-Mitarbeiterblatt: Soll-Jahr = Σ get_monthly_target."""
    expected = sum((calculation_service.get_monthly_target(db, late_starter, 2026, m)
                    for m in range(1, 13)), start=Decimal('0'))
    bio = export_service.generate_yearly_report(db, 2026)
    bio.seek(0)
    wb = load_workbook(bio)
    sheet = wb[[n for n in wb.sheetnames if "Starter" in n][0]]
    assert _summary(sheet, "Soll-Stunden Jahr:") == float(expected)


def test_pdf_monthly_summary_matches_calculation(db, late_starter):
    """PDF-Zusammenfassung trägt dieselbe Soll-Zahl wie die Berechnung."""
    expected = calculation_service.get_monthly_target(db, late_starter, 2026, 3)
    text = _pdf_text(export_service.generate_monthly_report_pdf(db, 2026, 3))
    assert "Soll-Stunden:" in text, "PDF-Text nicht lesbar — Testhelfer prüfen"
    assert f"{float(expected):.2f} h" in text, text[-2000:]
    assert "176.00 h" not in text


# ---------------------------------------------------------------------------
# ODS
# ---------------------------------------------------------------------------

def test_ods_monthly_summary_matches_calculation(db, late_starter):
    expected = calculation_service.get_monthly_target(db, late_starter, 2026, 3)
    rows = _ods_rows(ods_export_service.generate_monthly_report(db, 2026, 3))
    soll = [r for r in rows if r and r[0] == "Soll-Stunden Monat:"]
    assert soll, "Summenzeile nicht gefunden"
    assert float(soll[0][1]) == float(expected)
    ist = [r for r in rows if r and r[0] == "Ist-Stunden Monat:"]
    assert float(ist[0][1]) == float(
        calculation_service.get_monthly_actual(db, late_starter, 2026, 3))


def test_ods_yearly_employee_day_rows_are_windowed(db, late_starter):
    """ODS-Jahres-Mitarbeiterblatt hat keine Summenzeile — die Tageszeilen selbst
    müssen vor dem Eintritt Soll 0 / Ist 0 tragen."""
    from odf.opendocument import load as odf_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    bio = ods_export_service.generate_yearly_report(db, 2026)
    bio.seek(0)
    doc = odf_load(bio)
    table = [t for t in doc.spreadsheet.getElementsByType(Table)
             if "Starter" in str(t.getAttribute("name"))][0]
    for tr in table.getElementsByType(TableRow):
        cells = []
        for tc in tr.getElementsByType(TableCell):
            val = tc.getAttribute("value")
            if val is None:
                val = "".join(str(p) for p in tc.getElementsByType(P))
            cells.append(str(val))
        if cells and cells[0] == BEFORE.strftime('%d.%m.%Y'):
            assert float(cells[5]) == 0.0, cells   # Netto
            assert float(cells[6]) == 0.0, cells   # Soll
            return
    raise AssertionError("Tageszeile vor dem Eintritt nicht gefunden")


# ---------------------------------------------------------------------------
# Kontrolltests: ohne Eintritts-/Austrittsdatum ändert sich nichts
# ---------------------------------------------------------------------------

def test_control_user_without_window_unchanged(db, test_user):
    """MA ohne first_work_day/last_work_day: volle 176 h Soll wie bisher."""
    db.add(TimeEntry(user_id=test_user.id, tenant_id=DEFAULT_TENANT_ID, date=BEFORE,
                     start_time=time(8, 0), end_time=time(16, 0), break_minutes=0))
    db.commit()
    sheet = _sheet(export_service.generate_monthly_report(db, 2026, 3))
    assert _summary(sheet, "Soll-Stunden Monat:") == 176.00
    assert _summary(sheet, "Ist-Stunden Monat:") == 8.00
    net_before, target_before = _xlsx_day_row(sheet, BEFORE)
    assert target_before == 8.00
    assert net_before == 8.00
