"""Tests for configurable special days 24./31.12. (#146).

Verifies the per-practice handling of 24.12. and 31.12. as
working_day | half_day | free, and — when free — the Urlaub vs. bezahlte
Freistellung distinction (counts_as_vacation):

- working_day (default): the monthly target is unchanged (backward compatible).
- half_day: that day's target = 0.5 × the per-day target.
- free: that day's target = 0 (like a public holiday).
- free + counts_as_vacation=False (paid leave): target 0, NO vacation deduction.
- free + counts_as_vacation=True (vacation): target 0 AND one vacation day used.

December 2025: 24.12. (Wed) and 31.12. (Wed) are both ordinary weekdays with
no seeded public holiday, so the special-day rules are the only thing that
changes the target on those days.
"""

from decimal import Decimal
from datetime import date

import pytest

from app.models import User, UserRole, Absence, AbsenceType, PublicHoliday
from app.models.system_setting import SystemSetting
from app.services import calculation_service, special_days_service
from tests.conftest import DEFAULT_TENANT_ID


# 24./31.12.2025 are both Wednesdays.
DEC24 = date(2025, 12, 24)
DEC31 = date(2025, 12, 31)


def _set_setting(db, key: str, value: str, tenant_id=DEFAULT_TENANT_ID):
    db.add(SystemSetting(key=key, value=value, description=key, tenant_id=tenant_id))
    db.commit()


def _set_special_day(db, prefix: str, mode: str, counts_as_vacation: bool = True):
    _set_setting(db, f"{prefix}_mode", mode)
    _set_setting(db, f"{prefix}_counts_as_vacation", "true" if counts_as_vacation else "false")


# ---------------------------------------------------------------------------
# Defaults / backward compatibility
# ---------------------------------------------------------------------------

def test_default_is_working_day_no_change(db, test_user):
    """No settings -> 24./31.12. behave like normal working days (REQ-5)."""
    target = calculation_service.get_monthly_target(db, test_user, 2025, 12)

    # Full month target with no special handling: every Mon-Fri counts 8h.
    # 24. and 31.12. are both Wednesdays and must be included at full target.
    daily = calculation_service.get_daily_target(test_user)
    assert daily == Decimal("8.00")

    # Recompute the expected full-month target by hand (no holidays/absences).
    expected = Decimal("0")
    for day in range(1, 32):
        d = date(2025, 12, day)
        if d.weekday() < 5:
            expected += daily
    assert target == expected
    # Sanity: both special days are part of that full target.
    assert DEC24.weekday() < 5 and DEC31.weekday() < 5


def test_config_resolves_defaults(db, test_user):
    """get_special_day_config falls back to working_day with no rows."""
    cfg = special_days_service.get_special_day_config(db, DEFAULT_TENANT_ID, 2025)
    assert cfg[DEC24]["mode"] == "working_day"
    assert cfg[DEC31]["mode"] == "working_day"
    # working_day imposes no factor.
    assert special_days_service.special_day_target_factor(DEC24, cfg) is None
    assert special_days_service.special_day_target_factor(DEC31, cfg) is None


# ---------------------------------------------------------------------------
# half_day
# ---------------------------------------------------------------------------

def test_half_day_halves_target_for_that_day(db, test_user):
    """24.12. as half_day reduces the monthly target by half a daily target."""
    target_before = calculation_service.get_monthly_target(db, test_user, 2025, 12)

    _set_special_day(db, "special_day_dec24", "half_day")

    target_after = calculation_service.get_monthly_target(db, test_user, 2025, 12)
    daily = calculation_service.get_daily_target(test_user)  # 8.00

    # The day drops from full (8h) to half (4h) -> month target drops by 4h.
    assert target_before - target_after == (daily * Decimal("0.5"))


def test_half_day_factor(db, test_user):
    _set_special_day(db, "special_day_dec31", "half_day")
    cfg = special_days_service.get_special_day_config(db, DEFAULT_TENANT_ID, 2025)
    assert special_days_service.special_day_target_factor(DEC31, cfg) == Decimal("0.5")


def test_both_days_half(db, test_user):
    """Both 24. and 31.12. as half_day each remove half a daily target."""
    target_before = calculation_service.get_monthly_target(db, test_user, 2025, 12)

    _set_special_day(db, "special_day_dec24", "half_day")
    _set_special_day(db, "special_day_dec31", "half_day")

    target_after = calculation_service.get_monthly_target(db, test_user, 2025, 12)
    daily = calculation_service.get_daily_target(test_user)
    assert target_before - target_after == daily  # 0.5 + 0.5 = 1 full day


# ---------------------------------------------------------------------------
# free
# ---------------------------------------------------------------------------

def test_free_zeroes_target_for_that_day(db, test_user):
    """24.12. as free reduces the monthly target by one full daily target."""
    target_before = calculation_service.get_monthly_target(db, test_user, 2025, 12)

    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=False)

    target_after = calculation_service.get_monthly_target(db, test_user, 2025, 12)
    daily = calculation_service.get_daily_target(test_user)
    assert target_before - target_after == daily


def test_free_factor(db, test_user):
    _set_special_day(db, "special_day_dec24", "free")
    cfg = special_days_service.get_special_day_config(db, DEFAULT_TENANT_ID, 2025)
    assert special_days_service.special_day_target_factor(DEC24, cfg) == Decimal("0")


def test_free_on_holiday_not_double_handled(db, test_user):
    """A 'free' special day that already is a public holiday is not counted twice."""
    # Seed 24.12.2025 as a public holiday.
    db.add(PublicHoliday(date=DEC24, name="Heiligabend", year=2025, tenant_id=DEFAULT_TENANT_ID))
    db.commit()

    target_with_holiday_only = calculation_service.get_monthly_target(db, test_user, 2025, 12)

    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=False)
    target_with_free = calculation_service.get_monthly_target(db, test_user, 2025, 12)

    # The holiday already reduced the day to 0; marking it free changes nothing.
    assert target_with_free == target_with_holiday_only


# ---------------------------------------------------------------------------
# free + counts_as_vacation: vacation budget effect
# ---------------------------------------------------------------------------

def test_free_paid_leave_does_not_touch_vacation_budget(db, test_user):
    """free + counts_as_vacation=False (paid leave): vacation budget unchanged."""
    acct_before = calculation_service.get_vacation_account(db, test_user, 2025)

    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=False)
    _set_special_day(db, "special_day_dec31", "free", counts_as_vacation=False)

    acct_after = calculation_service.get_vacation_account(db, test_user, 2025)

    assert acct_after["used_hours"] == acct_before["used_hours"] == Decimal("0.0")
    assert acct_after["used_days"] == acct_before["used_days"] == Decimal("0.0")
    assert acct_after["remaining_days"] == acct_before["remaining_days"]


def test_free_vacation_consumes_one_day(db, test_user):
    """free + counts_as_vacation=True deducts one vacation day per such date."""
    daily = calculation_service.get_daily_target(test_user)  # 8.00

    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=True)

    acct = calculation_service.get_vacation_account(db, test_user, 2025)
    # Exactly one vacation day (one daily target worth of hours) is consumed.
    assert acct["used_hours"] == Decimal(str(daily.quantize(Decimal("0.01"))))
    assert acct["used_days"] == Decimal("1.0")
    assert acct["remaining_days"] == Decimal("29.0")  # 30 - 1


def test_both_free_vacation_consume_two_days(db, test_user):
    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=True)
    _set_special_day(db, "special_day_dec31", "free", counts_as_vacation=True)

    acct = calculation_service.get_vacation_account(db, test_user, 2025)
    assert acct["used_days"] == Decimal("2.0")
    assert acct["remaining_days"] == Decimal("28.0")


def test_free_vacation_plus_real_vacation_no_double_count(db, test_user):
    """If the employee already has a real VACATION absence on 24.12., the
    special-day rule must not deduct a second time for the same date."""
    daily = calculation_service.get_daily_target(test_user)
    db.add(Absence(
        user_id=test_user.id,
        tenant_id=DEFAULT_TENANT_ID,
        date=DEC24,
        type=AbsenceType.VACATION,
        hours=daily,
    ))
    db.commit()

    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=True)

    acct = calculation_service.get_vacation_account(db, test_user, 2025)
    # Only the single real VACATION absence counts -> exactly one day.
    assert acct["used_days"] == Decimal("1.0")


def test_free_vacation_on_holiday_no_deduction(db, test_user):
    """A free+vacation day that is already a public holiday costs no vacation."""
    db.add(PublicHoliday(date=DEC24, name="Heiligabend", year=2025, tenant_id=DEFAULT_TENANT_ID))
    db.commit()

    _set_special_day(db, "special_day_dec24", "free", counts_as_vacation=True)

    acct = calculation_service.get_vacation_account(db, test_user, 2025)
    assert acct["used_days"] == Decimal("0.0")


# ---------------------------------------------------------------------------
# Settings validation surface
# ---------------------------------------------------------------------------

def test_get_special_day_settings_shape(db, test_user):
    _set_special_day(db, "special_day_dec24", "half_day")
    settings = special_days_service.get_special_day_settings(db, DEFAULT_TENANT_ID)
    assert settings["special_day_dec24_mode"] == "half_day"
    assert settings["special_day_dec31_mode"] == "working_day"
    assert settings["special_day_dec24_counts_as_vacation"] is True


def test_vacation_deduction_dates_excludes_weekend_and_holiday(db, test_user):
    """Only weekday, non-holiday free+vacation dates are deduction dates."""
    # 2022: 24.12. is a Saturday, 31.12. is a Saturday -> both weekend.
    _set_setting(db, "special_day_dec24_mode", "free")
    _set_setting(db, "special_day_dec24_counts_as_vacation", "true")
    _set_setting(db, "special_day_dec31_mode", "free")
    _set_setting(db, "special_day_dec31_counts_as_vacation", "true")

    dates_2022 = special_days_service.vacation_deduction_dates_for_year(
        db, DEFAULT_TENANT_ID, 2022, holiday_dates=set()
    )
    assert dates_2022 == set()  # both fall on a weekend

    dates_2025 = special_days_service.vacation_deduction_dates_for_year(
        db, DEFAULT_TENANT_ID, 2025, holiday_dates=set()
    )
    assert dates_2025 == {DEC24, DEC31}  # both Wednesdays


# ---------------------------------------------------------------------------
# N-1: the configured factor must reach the §16 export Soll column (xlsx/ods),
# not just get_monthly_target — otherwise the export documentation diverges
# from the calculated monthly target on a configured 24./31.12.
# ---------------------------------------------------------------------------

def _xlsx_soll_for_december_day(wb, sheet_name: str, day: int):
    """Return the 'Soll (Std)' cell (column 7) for the given Dec-2025 day."""
    ws = wb[sheet_name[:31]]
    for row in ws.iter_rows(min_row=5):
        c0 = row[0].value
        if hasattr(c0, "month") and c0.year == 2025 and c0.month == 12 and c0.day == day:
            return row[6].value
    return None


def test_monthly_xlsx_export_soll_reflects_half_day(db, test_user):
    from openpyxl import load_workbook
    from app.services.export_service import generate_monthly_report

    _set_special_day(db, "special_day_dec24", "half_day")
    wb = load_workbook(generate_monthly_report(db, 2025, 12))
    sheet = f"{test_user.last_name} {test_user.first_name}"

    # 24.12. (half_day, Wed) → 4.0; 17.12. (ordinary Wed) → 8.0 (sanity).
    assert _xlsx_soll_for_december_day(wb, sheet, 24) == 4.0
    assert _xlsx_soll_for_december_day(wb, sheet, 17) == 8.0


def test_monthly_xlsx_export_soll_reflects_free(db, test_user):
    from openpyxl import load_workbook
    from app.services.export_service import generate_monthly_report

    _set_special_day(db, "special_day_dec31", "free", counts_as_vacation=False)
    wb = load_workbook(generate_monthly_report(db, 2025, 12))
    sheet = f"{test_user.last_name} {test_user.first_name}"

    # 31.12. (free, Wed) → 0.0; 30.12. (ordinary Tue) → 8.0 (sanity).
    assert _xlsx_soll_for_december_day(wb, sheet, 31) == 0.0
    assert _xlsx_soll_for_december_day(wb, sheet, 30) == 8.0


# ---------------------------------------------------------------------------
# Admin router: GET /settings/special-days + PUT validation
# ---------------------------------------------------------------------------

def _settings_client(db_session, current_user):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
    from app.database import get_db
    from app.middleware.auth import require_admin
    from app.routers import admin_settings

    app = FastAPI()
    app.include_router(admin_settings.router)

    def override_db():
        yield db_session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[require_admin] = lambda: current_user
    return TestClient(app)


def test_router_get_special_days_defaults(db, test_admin):
    client = _settings_client(db, test_admin)
    res = client.get("/api/admin/settings/special-days")
    assert res.status_code == 200
    body = res.json()
    assert body["special_day_dec24_mode"] == "working_day"
    assert body["special_day_dec31_mode"] == "working_day"
    assert body["special_day_dec24_counts_as_vacation"] is True
    assert body["special_day_dec31_counts_as_vacation"] is True


def test_router_put_mode_and_readback(db, test_admin):
    client = _settings_client(db, test_admin)

    res = client.put("/api/admin/settings/special_day_dec24_mode", json={"value": "half_day"})
    assert res.status_code == 200
    res = client.put(
        "/api/admin/settings/special_day_dec24_counts_as_vacation", json={"value": "false"}
    )
    assert res.status_code == 200

    body = client.get("/api/admin/settings/special-days").json()
    assert body["special_day_dec24_mode"] == "half_day"
    assert body["special_day_dec24_counts_as_vacation"] is False


def test_router_put_invalid_mode_rejected(db, test_admin):
    client = _settings_client(db, test_admin)
    res = client.put("/api/admin/settings/special_day_dec24_mode", json={"value": "bogus"})
    assert res.status_code == 400


def test_router_put_invalid_bool_rejected(db, test_admin):
    client = _settings_client(db, test_admin)
    res = client.put(
        "/api/admin/settings/special_day_dec31_counts_as_vacation", json={"value": "maybe"}
    )
    assert res.status_code == 400
