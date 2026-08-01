"""Audit 2026-07-31 — Nachzug N2: die Datei-Exporte rechneten die Ist-Gutschrift nicht.

TRAINING/SICK reduzieren das Soll NICHT — stattdessen bekommt die Ist-Seite die
Stunden der Abwesenheit gutgeschrieben (§ 3 EntgFG, ``credited_absences``). Die
Anwendung tut das ueber ``get_monthly_actual``. Die Summenzeilen der
Datei-Exporte dagegen summierten ``total_net`` — ausschliesslich erfasste
Zeiteintraege.

Ein Monat mit einem Kranktag stand damit in der Datei als::

    Soll-Stunden Monat:      176,00
    Ist-Stunden Monat:       160,00      <- die 16 h Gutschrift fehlen
    Saldo Monat:             -16,00
    Ueberstunden kumuliert:    0,00      <- get_overtime_account, korrekt

Das Dokument widersprach sich also in zwei benachbarten Zeilen selbst — und es
ist ein Nachweis nach § 16 Abs. 2 ArbZG. Der Bildschirm zeigte 0.

**Warum es erst jetzt sichtbar wurde:** vor dem 1.16.0-Fix der SOLL-Seite setzte
der Export an JEDEM Tag mit Abwesenheit das Soll pauschal auf 0. An einem
Kranktag stand damit 0 gegen 0 — der Fehler hob sich zufaellig auf. Seit das
Soll korrekt stehen bleibt (``absence_day_target`` delegiert an
``_day_soll_contribution``), wird die fehlende Gutschrift als Minus sichtbar.

**Betroffen waren vier Flaechen** (die Jahresuebersicht und der ODS-Klassiker
delegierten laengst, das ODS-Jahres-Mitarbeiterblatt hat gar keine Summenzeile):
XLSX-Monatsblatt, XLSX-Jahres-Mitarbeiterblatt, PDF-Monat, ODS-Monatsblatt.
Besonders deutlich im XLSX-Jahresbericht: die Jahresuebersicht und das
Mitarbeiterblatt im SELBEN Workbook nannten unterschiedliche Ist-Zahlen.

**Nicht angeglichen — der klassische Jahresbericht** (``generate_yearly_report_classic``).
Er ist ein bewusst anderes, aelteres Brutto-Modell {Arbeit, Krank, Urlaub}: dort
wird Krank vom SOLL abgezogen (Zeile „minus Krank Std.") statt dem Ist
gutgeschrieben, und Zeile 11 zeigt ausdruecklich die physisch GEARBEITETEN
Stunden (``get_monthly_worked_hours``). In sich stimmig, nur nach einem anderen
Modell — ein Angleichen wuerde ihn zerstoeren. Der Kontrolltest unten pinnt das.
"""
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import Absence, AbsenceType, TimeEntry
from app.services import calculation_service as cs
from tests.conftest import DEFAULT_TENANT_ID


# Januar 2026: 05.01. = Montag. 06.01. (Di) wird zum Kranktag, 13.01. (Di) zum
# Fortbildungstag. Beide sind regulaere Werktage — kein Feiertag, kein Sondertag,
# also greift credit_day_weight mit 1,0 und die Gutschrift ist voll.
SICK_DAY = date(2026, 1, 6)
TRAINING_DAY = date(2026, 1, 13)
WORK_DAY = date(2026, 1, 5)


def _entry(db, user, d, start_h=8, end_h=17, break_min=60):
    e = TimeEntry(
        user_id=user.id, tenant_id=DEFAULT_TENANT_ID, date=d,
        start_time=time(start_h, 0), end_time=time(end_h, 0), break_minutes=break_min,
    )
    db.add(e)
    db.commit()
    return e


def _absence(db, user, d, typ, hours=8.0):
    a = Absence(
        user_id=user.id, tenant_id=DEFAULT_TENANT_ID, date=d, type=typ, hours=hours,
    )
    db.add(a)
    db.commit()
    return a


@pytest.fixture
def credited_month(db, test_user):
    """Ein gearbeiteter Tag + ein Kranktag + ein Fortbildungstag im Januar 2026.

    Erwartung an ALLEN Flaechen: Ist = 8 (Stempelzeit) + 8 (Krank) + 8
    (Fortbildung) = 24 h — genau das, was ``get_monthly_actual`` liefert.
    """
    _entry(db, test_user, WORK_DAY)
    _absence(db, test_user, SICK_DAY, AbsenceType.SICK)
    _absence(db, test_user, TRAINING_DAY, AbsenceType.TRAINING)
    return test_user


def _expect_month(db, user):
    return (cs.get_monthly_target(db, user, 2026, 1),
            cs.get_monthly_actual(db, user, 2026, 1))


def _xlsx_summary(bio: BytesIO, user, labels):
    bio.seek(0)
    wb = load_workbook(bio, read_only=True)
    ws = next(
        wb[n] for n in wb.sheetnames
        if user.last_name in n and n not in ("Jahresübersicht", "Abwesenheiten")
    )
    return {
        row[0].value: row[1].value
        for row in ws.iter_rows(values_only=False)
        if row[0].value in labels
    }


# --------------------------------------------------------------------------
# Vorbedingung: die Gutschrift ist ueberhaupt im Spiel
# --------------------------------------------------------------------------

def test_fixture_actually_produces_a_credit(db, credited_month):
    """Ohne diese Vorbedingung waeren alle Tests unten vakuum-gruen.

    Die bestehenden Export-Tests gingen genau daran vorbei: sie pinnen
    ``export Ist == get_monthly_actual`` seit 1.16.0, aber ihre Fixtures
    enthalten nur einen blanken Zeiteintrag und keine gutgeschriebene
    Abwesenheit — die Gleichung hielt dort auch ohne die Gutschrift.
    """
    target, actual = _expect_month(db, credited_month)
    worked = cs.get_monthly_worked_hours(db, credited_month, 2026, 1)
    assert worked == 8, "genau ein gestempelter 8-h-Tag"
    assert actual == 24, "8 h gestempelt + 8 h Krank + 8 h Fortbildung"
    assert actual - worked == 16, "16 h Gutschrift muessen den Unterschied machen"
    assert target > 0


# --------------------------------------------------------------------------
# N2 — die vier betroffenen Summenzeilen
# --------------------------------------------------------------------------

def test_xlsx_monthly_summary_includes_credit(db, credited_month):
    """XLSX-Monatsblatt: Ist/Saldo folgen der Anwendung."""
    from app.services.export_service import generate_monthly_report

    target, actual = _expect_month(db, credited_month)
    rows = _xlsx_summary(
        generate_monthly_report(db, 2026, 1, tenant_id=DEFAULT_TENANT_ID),
        credited_month,
        ("Soll-Stunden Monat:", "Ist-Stunden Monat:", "Saldo Monat:",
         "Überstunden kumuliert:"),
    )
    assert rows["Ist-Stunden Monat:"] == pytest.approx(float(actual))
    assert rows["Soll-Stunden Monat:"] == pytest.approx(float(target))
    assert rows["Saldo Monat:"] == pytest.approx(float(actual - target))
    # Der Kern des Findings: das Dokument darf sich nicht selbst widersprechen.
    assert rows["Saldo Monat:"] == pytest.approx(rows["Überstunden kumuliert:"])


def test_xlsx_yearly_employee_summary_includes_credit(db, credited_month):
    """XLSX-Jahres-Mitarbeiterblatt — und Gleichstand mit der Jahresuebersicht
    im SELBEN Workbook, die schon immer delegierte."""
    from app.services.export_service import generate_yearly_report

    result = generate_yearly_report(db, 2026, tenant_id=DEFAULT_TENANT_ID)
    rows = _xlsx_summary(
        result, credited_month,
        ("Soll-Stunden Jahr:", "Ist-Stunden Jahr:", "Saldo Jahr:"),
    )
    expected_actual = sum(
        (cs.get_monthly_actual(db, credited_month, 2026, m) for m in range(1, 13)),
        start=Decimal("0"),
    )
    assert rows["Ist-Stunden Jahr:"] == pytest.approx(float(expected_actual))

    result.seek(0)
    wb = load_workbook(result, read_only=True)
    overview = wb["Jahresübersicht"]
    ov = next(
        r for r in overview.iter_rows(min_row=4, values_only=False)
        if r[0].value == f"{credited_month.last_name}, {credited_month.first_name}"
    )
    assert ov[3].value == pytest.approx(rows["Ist-Stunden Jahr:"]), (
        "Jahresuebersicht und Mitarbeiterblatt desselben Workbooks muessen "
        "dieselbe Ist-Zahl nennen"
    )
    assert ov[4].value == pytest.approx(rows["Saldo Jahr:"])


def test_ods_monthly_summary_includes_credit(db, credited_month):
    """ODS-Monatsblatt — dieselbe Auskunft muss in jedem Dateiformat dieselbe
    Zahl tragen."""
    from app.services.ods_export_service import _doc_with_styles, _monthly_sheet
    from tests.test_ods_export_service import _summary_value

    target, actual = _expect_month(db, credited_month)
    doc, bold, normal = _doc_with_styles()
    _monthly_sheet(doc, db, credited_month, 2026, 1, bold, normal)
    sheet = f"{credited_month.last_name} {credited_month.first_name}"[:31]

    assert _summary_value(doc, sheet, "Ist-Stunden Monat:") == pytest.approx(float(actual))
    assert _summary_value(doc, sheet, "Soll-Stunden Monat:") == pytest.approx(float(target))
    assert _summary_value(doc, sheet, "Saldo Monat:") == pytest.approx(float(actual - target))
    assert _summary_value(doc, sheet, "Saldo Monat:") == pytest.approx(
        _summary_value(doc, sheet, "Überstunden kumuliert:"))


def test_pdf_monthly_summary_includes_credit(db, credited_month):
    """PDF-Monat. Der Text im PDF ist nicht ohne Weiteres auslesbar, deshalb
    wird die Summary-Quelle direkt beobachtet: ``get_monthly_actual`` MUSS beim
    Bauen des PDFs fuer diesen MA aufgerufen worden sein (vorher lief der
    Nicht-Modus-Zweig komplett an ihr vorbei), und das PDF muss valide bleiben.
    """
    from app.services import export_service

    seen = []
    orig = export_service.calculation_service.get_monthly_actual

    def _spy(db_, user_, year_, month_, **kw):
        if user_.id == credited_month.id and (year_, month_) == (2026, 1):
            seen.append(kw)
        return orig(db_, user_, year_, month_, **kw)

    export_service.calculation_service.get_monthly_actual = _spy
    try:
        result = export_service.generate_monthly_report_pdf(
            db, 2026, 1, tenant_id=DEFAULT_TENANT_ID)
    finally:
        export_service.calculation_service.get_monthly_actual = orig

    assert seen, "die PDF-Summary muss ueber get_monthly_actual gehen"
    assert all(kw.get("up_to_date") is None for kw in seen), (
        "§16-Dateien rechnen den VOLLEN Monat, nicht bis zum Saldo-Stichtag"
    )
    result.seek(0)
    assert result.read(4) == b"%PDF", "PDF muss valide bleiben"


# --------------------------------------------------------------------------
# Kontrollen: was NICHT angefasst werden durfte
# --------------------------------------------------------------------------

def test_classic_yearly_report_keeps_gross_model(db, credited_month):
    """Der klassische Jahresbericht bleibt beim Brutto-Modell.

    Zeile „Gearbeitete Std." zeigt weiterhin die physisch gestempelten 8 h und
    NICHT die 24 h aus ``get_monthly_actual`` — dort wird Krank ueber die Zeile
    „minus Krank Std." vom Soll abgezogen statt dem Ist gutgeschrieben. Ein
    Angleichen wuerde beides zaehlen.
    """
    from app.services.export_service import generate_yearly_report_classic

    result = generate_yearly_report_classic(db, 2026, tenant_id=DEFAULT_TENANT_ID)
    result.seek(0)
    wb = load_workbook(result, read_only=True)
    ws = next(wb[n] for n in wb.sheetnames if credited_month.last_name in n)
    labels = {}
    for row in ws.iter_rows(values_only=False):
        if row[0].value and isinstance(row[0].value, str):
            labels[row[0].value.strip()] = row
    worked_label = next((k for k in labels if "erbrachte Stunden" in k), None)
    assert worked_label, f"Zeile mit erbrachten Stunden nicht gefunden: {list(labels)}"
    # Spalte 1 = Beschriftung, Spalte 2 = "Übertrag Vorjahr", Spalte 3 = Januar.
    jan = labels[worked_label][2].value
    # 8 h gestempelt + 8 h Krank. Das Krank kommt hier NICHT aus der
    # Ist-Gutschrift, sondern aus der Art.-9-Maskierung: ohne
    # ``include_health_data`` wird Krank nicht vom Soll abgezogen (Zeile „minus
    # Krank Std." zeigt „–") und stattdessen der Arbeitsseite gutgeschrieben,
    # damit der Saldo stimmt und Krank aus keiner sichtbaren Zelle ableitbar
    # ist. Die 8 h FORTBILDUNG fehlen bewusst — das 2-Zeilen-Layout modelliert
    # nur {Arbeit, Krank, Urlaub}.
    assert jan == pytest.approx(16.0), (
        "der Klassiker bleibt beim Brutto-Modell (8 h gearbeitet + 8 h "
        f"maskiertes Krank), nicht get_monthly_actual (24 h) — erhalten: {jan}"
    )
    _, actual = _expect_month(db, credited_month)
    assert jan != pytest.approx(float(actual)), (
        "der Klassiker darf NICHT auf get_monthly_actual angeglichen werden"
    )


def test_daily_rows_still_show_stamped_time(db, credited_month):
    """Bewusste Grenze: die PER-TAG-Zeilen bleiben Stempelzeit.

    Spalte „Netto (Std)" steht zwischen „Pause (Min)" und „Soll (Std)" und ist
    der § 16-Nachweis der tatsaechlichen Anwesenheit — an einem Kranktag 0.
    Woher die Gutschrift kommt, sagt die Spalte „Abwesenheit" („Krank (8.0h)").
    Verbindlich ist die Summenzeile. Wer das aendern will, aendert die Bedeutung
    einer § 16-Spalte — deshalb hier festgehalten statt stillschweigend.
    """
    from app.services.export_service import generate_monthly_report

    result = generate_monthly_report(db, 2026, 1, tenant_id=DEFAULT_TENANT_ID)
    result.seek(0)
    wb = load_workbook(result, read_only=True)
    ws = next(
        wb[n] for n in wb.sheetnames
        if credited_month.last_name in n and n not in ("Jahresübersicht", "Abwesenheiten")
    )
    def _as_date(v):
        # openpyxl liefert Datumszellen als ``datetime`` zurueck.
        return v.date() if isinstance(v, datetime) else v

    sick_row = next(
        r for r in ws.iter_rows(values_only=False)
        if _as_date(r[0].value) == SICK_DAY
    )
    assert sick_row[5].value == pytest.approx(0.0), "Netto (Std) = Stempelzeit"
    assert sick_row[6].value == pytest.approx(8.0), "Soll bleibt stehen (§3 EntgFG)"
    assert "8.0h" in (sick_row[8].value or ""), "Spalte Abwesenheit weist die Stunden aus"


def test_user_without_credit_unchanged(db, test_user):
    """Byte-Identitaets-Kontrolle: ohne gutgeschriebene Abwesenheit aendert sich
    nichts — die Summenzeile bleibt exakt die alte Per-Tag-Summe."""
    from app.services.export_service import generate_monthly_report

    _entry(db, test_user, WORK_DAY)
    target, actual = _expect_month(db, test_user)
    rows = _xlsx_summary(
        generate_monthly_report(db, 2026, 1, tenant_id=DEFAULT_TENANT_ID),
        test_user,
        ("Soll-Stunden Monat:", "Ist-Stunden Monat:", "Saldo Monat:"),
    )
    assert rows["Ist-Stunden Monat:"] == pytest.approx(8.0)
    assert rows["Ist-Stunden Monat:"] == pytest.approx(float(actual))
    assert rows["Soll-Stunden Monat:"] == pytest.approx(float(target))
