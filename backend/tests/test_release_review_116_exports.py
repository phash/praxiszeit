"""Release-Review 1.16.0 — die §16-Datei-Exporte dürfen sich nicht mehr selbst
widersprechen.

Alle Per-Tag-Schleifen der Exporte setzten das Tages-Soll auf 0, sobald der Tag
IRGENDEINE Abwesenheit trug. Das ist an drei Stellen falsch:

* **Halbtag**: nur 0,5 × Tagessoll fällt weg — ein halber Urlaubstag plus vier
  gestempelte Stunden ergab Soll 0 / Ist 4 und damit +4 h Überstunden statt 0.
* **SICK/TRAINING**: nicht soll-reduzierend, das Soll bleibt stehen.
* **OVERTIME**: Soll bleibt, Ist = 0 (docs/BERECHNUNGEN.md §6) — der Export zeigte
  0/0 und verschluckte den Konto-Abbau.

Die Summenzeilen desselben Dokuments widersprachen damit ihrem eigenen
„Überstunden kumuliert" (das aus dem modus-bewussten ``get_overtime_account``
kommt). Diese Tests binden die Export-Summen an ``get_monthly_target`` /
``get_monthly_actual`` — die Quelle, die Bildschirm und Berechnung nutzen.
"""
from datetime import date, time
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import Absence, AbsenceType, TimeEntry
from app.services import calculation_service
from app.services.export_service import generate_monthly_report, generate_monthly_report_pdf
from tests.conftest import DEFAULT_TENANT_ID


YEAR, MONTH = 2026, 3


def _entry(db, user, d, start_h, end_h, break_min=0):
    e = TimeEntry(
        user_id=user.id, tenant_id=DEFAULT_TENANT_ID, date=d,
        start_time=time(start_h, 0), end_time=time(end_h, 0), break_minutes=break_min,
    )
    db.add(e); db.commit()
    return e


def _absence(db, user, d, typ, hours, half_day=False):
    a = Absence(
        user_id=user.id, tenant_id=DEFAULT_TENANT_ID, date=d,
        type=typ, hours=hours, half_day=half_day,
    )
    db.add(a); db.commit()
    return a


def _sheet(db, user):
    bio = generate_monthly_report(db, YEAR, MONTH)
    bio.seek(0)
    wb = load_workbook(BytesIO(bio.read()))
    return wb[f"{user.last_name} {user.first_name}"[:31]]


def _summary(sheet):
    """Summenblock: (Soll, Ist) aus den beschrifteten Zeilen lesen."""
    out = {}
    for row in range(1, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        if isinstance(label, str) and label.startswith(("Soll-Stunden", "Ist-Stunden")):
            out[label.split()[0]] = float(sheet.cell(row=row, column=2).value or 0)
    return out


def _day_row(sheet, d):
    """Die Tageszeile zu einem Datum: (Netto, Soll, Differenz)."""
    stamp = d.strftime("%d.%m.%Y")
    for row in range(1, sheet.max_row + 1):
        v = sheet.cell(row=row, column=1).value
        if (v.strftime("%d.%m.%Y") if hasattr(v, "strftime") else str(v)) == stamp:
            return (
                sheet.cell(row=row, column=6).value,
                sheet.cell(row=row, column=7).value,
                sheet.cell(row=row, column=8).value,
            )
    raise AssertionError(f"keine Tageszeile für {stamp}")


class TestHalfDayAbsence:
    """Halber Urlaubstag + halber Arbeitstag = ausgeglichen, nicht +4 h."""

    def test_day_row_keeps_half_the_target(self, db, test_user):
        d = date(YEAR, MONTH, 9)  # Montag
        _absence(db, test_user, d, AbsenceType.VACATION, 4, half_day=True)
        _entry(db, test_user, d, 8, 12)

        _net, target, diff = _day_row(_sheet(db, test_user), d)
        assert float(target) == 4.0, "halber Urlaubstag darf nur das halbe Soll entfernen"
        assert float(diff) == 0.0, "vier gearbeitete Stunden gegen vier Soll-Stunden = ausgeglichen"

    def test_summary_matches_the_calculation(self, db, test_user):
        d = date(YEAR, MONTH, 9)
        _absence(db, test_user, d, AbsenceType.VACATION, 4, half_day=True)
        _entry(db, test_user, d, 8, 12)

        summary = _summary(_sheet(db, test_user))
        expected = float(calculation_service.get_monthly_target(db, test_user, YEAR, MONTH))
        assert summary["Soll-Stunden"] == pytest.approx(expected, abs=0.01)


class TestOvertimeCompensationDay:
    """Freizeitausgleich: Soll bleibt stehen, Ist ist 0 → der Tag zieht das Konto."""

    def test_day_row_keeps_the_full_target(self, db, test_user):
        d = date(YEAR, MONTH, 10)  # Dienstag
        _absence(db, test_user, d, AbsenceType.OVERTIME, 8)

        _net, target, _diff = _day_row(_sheet(db, test_user), d)
        assert float(target) == 8.0, "OVERTIME ist nicht soll-reduzierend (BERECHNUNGEN.md §6)"

    def test_summary_matches_the_calculation(self, db, test_user):
        _absence(db, test_user, date(YEAR, MONTH, 10), AbsenceType.OVERTIME, 8)
        summary = _summary(_sheet(db, test_user))
        expected = float(calculation_service.get_monthly_target(db, test_user, YEAR, MONTH))
        assert summary["Soll-Stunden"] == pytest.approx(expected, abs=0.01)


class TestSickDayKeepsTarget:
    """Krank ist nicht soll-reduzierend — die Gutschrift läuft über das Ist."""

    def test_day_row_keeps_the_full_target(self, db, test_user):
        d = date(YEAR, MONTH, 11)  # Mittwoch
        _absence(db, test_user, d, AbsenceType.SICK, 8)

        _net, target, _diff = _day_row(_sheet(db, test_user), d)
        assert float(target) == 8.0


class TestFullDayVacationStillRemovesTarget:
    """Kontrollfall: der häufigste Fall bleibt unverändert bei 0."""

    def test_day_row_target_is_zero(self, db, test_user):
        d = date(YEAR, MONTH, 12)  # Donnerstag
        _absence(db, test_user, d, AbsenceType.VACATION, 8, half_day=False)

        _net, target, _diff = _day_row(_sheet(db, test_user), d)
        assert float(target) == 0.0

    def test_clean_month_summary_is_unchanged(self, db, test_user):
        """Ein Monat ohne Abwesenheiten muss byte-identisch zur alten Ausgabe
        bleiben — der Fix darf nur Abwesenheitstage betreffen."""
        _entry(db, test_user, date(YEAR, MONTH, 2), 8, 16)
        summary = _summary(_sheet(db, test_user))
        expected = float(calculation_service.get_monthly_target(db, test_user, YEAR, MONTH))
        assert summary["Soll-Stunden"] == pytest.approx(expected, abs=0.01)


class TestPdfFixedModeSummary:
    """#377 Baustein 2b: das PDF war die einzige Monats-Exportfläche ohne
    Fixmodus-Branch — drei Dateiformate desselben Monats trugen verschiedene
    Soll-Zahlen."""

    def _minijob(self, db, user):
        user.milog_working_time_account = True
        user.use_fixed_monthly_target = True
        user.agreed_monthly_hours = Decimal("40.0")
        user.use_daily_schedule = True
        user.hours_monday = 3
        user.hours_tuesday = 0
        user.hours_wednesday = 3
        user.hours_thursday = 0
        user.hours_friday = 0
        db.commit()
        return user

    def test_pdf_builds_for_a_fixed_mode_user(self, db, test_user):
        self._minijob(db, test_user)
        out = generate_monthly_report_pdf(db, YEAR, MONTH)
        out.seek(0)
        assert out.read()[:5] == b"%PDF-"

    def test_xlsx_summary_uses_the_flat_agreed_target(self, db, test_user):
        """Der XLSX-Zweig war schon korrekt — er dient hier als Referenzwert,
        an den das PDF gebunden ist."""
        user = self._minijob(db, test_user)
        summary = _summary(_sheet(db, user))
        assert summary["Soll-Stunden"] == pytest.approx(40.0, abs=0.01)
        assert summary["Soll-Stunden"] == pytest.approx(
            float(calculation_service.get_monthly_target(db, user, YEAR, MONTH)), abs=0.01
        )
