"""#415 — Stundenänderungen in Monats- und Jahresberichten darstellen.

Modell (``working_hours_changes``), Endpoints, UI (WorkingHoursModal) und die
Soll/Ist-Berechnung (``get_weekly_hours_for_date`` in allen Per-Tag-Schleifen)
existierten bereits. Offen war Punkt 4 des Issues: die **Darstellung**.

Konkret war das §16-Dokument selbstwidersprüchlich — die Tageszeilen rechneten
historisch korrekt, die Kopfzeile/Spalte „Wochenstunden" zeigte aber den
AKTUELLEN Vertragswert. Wer im März von 20 h auf 30 h wechselte, bekam einen
Märzbericht mit „30,0 Wochenstunden" über Tageszeilen, die für die erste
Monatshälfte mit 20 h gerechnet hatten.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import TimeEntry, WorkingHoursChange
from app.services import calculation_service
from app.services.export_service import format_weekly_hours_history
from tests.conftest import DEFAULT_TENANT_ID


def _mk_change(db, user, effective_from, weekly_hours, note=None, work_days=None):
    c = WorkingHoursChange(
        user_id=user.id,
        tenant_id=DEFAULT_TENANT_ID,
        effective_from=effective_from,
        weekly_hours=Decimal(str(weekly_hours)),
        note=note,
        work_days_per_week=work_days,
    )
    db.add(c)
    db.commit()
    return c


def _mk_day_plan_change(db, user, effective_from, days, weekly_hours=None, work_days=None):
    """#431: eine Snapshot-Zeile im TAGESPLAN-Modus. ``days`` ist ein 5-Tupel
    (Mo…Fr), je Zahl oder None."""
    values = [None if v is None else Decimal(str(v)) for v in days]
    total = sum((v for v in values if v is not None), Decimal("0"))
    c = WorkingHoursChange(
        user_id=user.id,
        tenant_id=DEFAULT_TENANT_ID,
        effective_from=effective_from,
        weekly_hours=Decimal(str(weekly_hours)) if weekly_hours is not None else total,
        use_daily_schedule=True,
        hours_monday=values[0],
        hours_tuesday=values[1],
        hours_wednesday=values[2],
        hours_thursday=values[3],
        hours_friday=values[4],
        work_days_per_week=work_days if work_days is not None else max(
            1, len([v for v in values if v])
        ),
    )
    db.add(c)
    db.commit()
    return c


def _seg(start, end, weekly_hours, use_daily_schedule=False, day_hours=(None,) * 5,
         work_days_per_week=5):
    """Kurzschreibweise fuer ein erwartetes ``ScheduleSegment``."""
    return calculation_service.ScheduleSegment(
        start=start,
        end=end,
        weekly_hours=Decimal(str(weekly_hours)),
        use_daily_schedule=use_daily_schedule,
        day_hours=tuple(None if v is None else Decimal(str(v)) for v in day_hours),
        work_days_per_week=work_days_per_week,
    )


def _load_xlsx(bio: BytesIO):
    bio.seek(0)
    return load_workbook(BytesIO(bio.read()))


# ─────────────────────────────────────────────────────────────────────
# calculation_service.weekly_hours_segments — die eine Quelle
# ─────────────────────────────────────────────────────────────────────


class TestWeeklyHoursSegments:
    def test_no_changes_is_a_single_segment(self, db, test_user):
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs == [_seg(date(2026, 3, 1), date(2026, 3, 31), 40)]

    def test_change_inside_the_period_splits_it(self, db, test_user):
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs == [
            _seg(date(2026, 3, 1), date(2026, 3, 14), 40),
            _seg(date(2026, 3, 15), date(2026, 3, 31), 30),
        ]

    def test_change_before_the_period_applies_for_the_whole_period(self, db, test_user):
        _mk_change(db, test_user, date(2026, 1, 1), 20)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs == [_seg(date(2026, 3, 1), date(2026, 3, 31), 20)]

    def test_change_after_the_period_is_ignored(self, db, test_user):
        _mk_change(db, test_user, date(2026, 4, 1), 20)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs == [_seg(date(2026, 3, 1), date(2026, 3, 31), 40)]

    def test_change_exactly_on_the_first_day_is_one_segment(self, db, test_user):
        _mk_change(db, test_user, date(2026, 3, 1), 25)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs == [_seg(date(2026, 3, 1), date(2026, 3, 31), 25)]

    def test_multiple_changes_produce_multiple_segments(self, db, test_user):
        _mk_change(db, test_user, date(2026, 3, 10), 30)
        _mk_change(db, test_user, date(2026, 3, 20), 10)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert [s.weekly_hours for s in segs] == [Decimal("40.0"), Decimal("30.0"), Decimal("10.0")]
        assert [s.start for s in segs] == [date(2026, 3, 1), date(2026, 3, 10), date(2026, 3, 20)]

    def test_change_to_the_same_value_is_not_a_visible_change(self, db, test_user):
        """Ein Eintrag, der die Stundenzahl nicht aendert, darf im Bericht
        keine Pseudo-Aenderung erzeugen."""
        _mk_change(db, test_user, date(2026, 3, 15), 40)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs == [_seg(date(2026, 3, 1), date(2026, 3, 31), 40)]

    def test_year_range_spans_all_changes(self, db, test_user):
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        _mk_change(db, test_user, date(2026, 9, 1), 20)
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )
        assert len(segs) == 3
        assert segs[-1] == _seg(date(2026, 9, 1), date(2026, 12, 31), 20)

    def test_empty_range_returns_nothing(self, db, test_user):
        assert calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 31), date(2026, 3, 1)
        ) == []

    def test_preloaded_changes_match_the_db_path(self, db, test_user):
        """Der Preload-Pfad (Hot-Loop) muss byte-identisch zum DB-Pfad sein."""
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        preload = db.query(WorkingHoursChange).filter(
            WorkingHoursChange.user_id == test_user.id
        ).all()
        assert calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31), wh_changes=preload
        ) == calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )


# ─────────────────────────────────────────────────────────────────────
# #431: das Segment traegt den VOLLSTAENDIGEN Vertrags-Snapshot
# ─────────────────────────────────────────────────────────────────────


class TestSegmentsCarryTheFullSnapshot:
    def test_uniform_segment_reports_mode_and_work_days(self, db, test_user):
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert segs[0].use_daily_schedule is False
        assert segs[0].day_hours == (None,) * 5
        assert segs[0].work_days_per_week == 5

    def test_segments_carry_day_plan(self, db, test_user):
        """Der Kernfall: eine Tagesplan-Zeile muss ihre Tageswerte ins Segment
        tragen — sonst zeigt der Bericht nur die Wochensumme, die bei diesen
        Mitarbeitenden kein einziges Tagessoll steuert."""
        _mk_day_plan_change(db, test_user, date(2026, 3, 1), (8, 5, 4, None, None))

        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )

        assert len(segs) == 2
        assert segs[1].use_daily_schedule is True
        assert segs[1].day_hours[0] == Decimal("8")
        assert segs[1].weekly_hours == Decimal("17")
        assert segs[1].work_days_per_week == 3

    def test_same_weekly_hours_but_different_day_plan_is_a_visible_change(self, db, test_user):
        """Der Mischfall: gleiche Wochensumme, verschobener Tagesplan. Die alte
        Verschmelzungsregel (nur ``weekly_hours``) machte daraus EIN Segment —
        der Bericht behauptete „keine Aenderung", obwohl sich jedes Tagessoll
        verschoben hatte."""
        _mk_day_plan_change(db, test_user, date(2026, 1, 1), (8, 5, 4, None, None))
        _mk_day_plan_change(db, test_user, date(2026, 7, 1), (4, 5, 8, None, None))

        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )

        assert [s.start for s in segs] == [date(2026, 1, 1), date(2026, 7, 1)]
        assert [s.weekly_hours for s in segs] == [Decimal("17"), Decimal("17")]
        assert segs[1].day_hours[0] == Decimal("4")

    def test_switching_the_mode_is_a_visible_change(self, db, test_user):
        """Gleiche Wochensumme, anderer Modus: 17 h auf 5 gleichmaessige Tage ist
        etwas anderes als Mo 8 / Di 5 / Mi 4."""
        _mk_change(db, test_user, date(2026, 1, 1), 17)
        _mk_day_plan_change(db, test_user, date(2026, 7, 1), (8, 5, 4, None, None))

        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )

        assert len(segs) == 2
        assert [s.use_daily_schedule for s in segs] == [False, True]

    def test_changing_only_the_work_days_is_a_visible_change(self, db, test_user):
        """40 h auf 4 statt 5 Tage aendert jedes Tagessoll (8,0 -> 10,0), obwohl
        die Wochenstunden gleich bleiben."""
        _mk_change(db, test_user, date(2026, 7, 1), 40, work_days=4)

        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )

        assert [s.work_days_per_week for s in segs] == [5, 4]

    def test_inert_day_values_in_uniform_mode_do_not_split(self, db, test_user):
        """Byte-Identitaet: ``update_user`` raeumt ``hours_*`` beim Abschalten des
        Tagesplans nicht ab. Solche Reste sind im gleichmaessigen Modus
        rechnerisch inert (``get_daily_target_for_date`` liest sie dort nicht)
        und duerfen keine Pseudo-Aenderung erzeugen."""
        test_user.use_daily_schedule = False
        test_user.hours_monday = Decimal("8.0")
        db.commit()
        c = _mk_change(db, test_user, date(2026, 7, 1), 40)
        c.hours_monday = Decimal("3.0")  # anderer Rest, weiterhin inert
        db.commit()

        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )

        assert segs == [_seg(date(2026, 1, 1), date(2026, 12, 31), 40)]

    def test_day_plan_user_without_history_reports_the_current_plan(self, db, day_plan_user):
        """Ohne erfasste Aenderung faellt der Resolver auf die User-Zeile zurueck —
        das Segment muss den Tagesplan trotzdem tragen."""
        segs = calculation_service.weekly_hours_segments(
            db, day_plan_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert len(segs) == 1
        assert segs[0].use_daily_schedule is True
        assert segs[0].day_hours[:3] == (Decimal("8"), Decimal("5"), Decimal("4"))


# ─────────────────────────────────────────────────────────────────────
# Zahlformatierung: der Wechsel auf zwei Nachkommastellen ist auf allen
# Bestandsdaten ein No-op
# ─────────────────────────────────────────────────────────────────────


class TestDeHoursExactIsAByteIdenticalUpgrade:
    """Belegt, dass ``_de_hours_exact`` die #415-Ausgabe NICHT veraendert, wo sie
    vor #431 ueberhaupt entstehen konnte — und sie genau dort korrigiert, wo
    Datei und Bildschirm auseinanderliefen."""

    def test_identical_for_every_value_the_old_column_could_hold(self):
        """``working_hours_changes.weekly_hours`` war ``Numeric(4,1)``: 0,0 bis
        60,0 in Zehntelschritten. Fuer JEDEN dieser 601 Werte liefern alte und
        neue Formatierung dieselbe Zeichenkette."""
        from app.services.export_service import _de_hours, _de_hours_exact
        divergent = [
            v for v in (Decimal(i) / 10 for i in range(0, 601))
            if _de_hours(v) != _de_hours_exact(v)
        ]
        assert divergent == []

    def test_quarter_hour_values_are_where_the_old_pair_disagreed(self):
        """Erst ``Numeric(4,2)`` (Migration 067) macht 38,25 h speicherbar. Die
        alte Formatierung rundet dort auf eine Stelle — Python half-even („38,2"),
        der Frontend-Zwilling ``toFixed(1)`` kaufmaennisch („38,3"). Die neue
        rundet gar nicht."""
        from app.services.export_service import _de_hours, _de_hours_exact
        assert _de_hours(Decimal("38.25")) == "38,2"      # alt, half-even
        assert _de_hours_exact(Decimal("38.25")) == "38,25"  # neu, verlustfrei

    def test_uniform_history_keeps_the_full_value(self):
        """Zwilling zu ``formatters.test.ts`` („keeps quarter hours in the
        uniform wording, too")."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 12, 31), Decimal("38.25")),
        ])
        assert text == "ab 01.03.2026: 38,25 Std/Woche"


# ─────────────────────────────────────────────────────────────────────
# Invariante: weekly_hours == Summe der Tageswerte
# ─────────────────────────────────────────────────────────────────────


class TestWeeklyHoursMustMatchTheDayPlanSum:
    """Die Darstellung ist der Zeile TREU — sie rechnet die Wochensumme NIE neu.

    Damit ist ``weekly_hours == Σ(hours_*)`` im Tagesplan-Modus eine Invariante,
    die die Schreibpfade halten muessen:

    * ``WorkingHoursChangeCreate.check_mode`` setzt sie beim Anlegen,
    * Migration 067 zieht sie beim Backfill nach.

    Die Tests hier nageln fest, was passiert, wenn eine Zeile sie VERLETZT —
    genau das war der Fund der Task-9-Review: der Backfill stempelte den heutigen
    Tagesplan auf Alt-Zeilen, liess ``weekly_hours`` aber auf dem historischen
    Skalar stehen.
    """

    def test_a_broken_row_prints_a_self_contradicting_sentence(self, db, test_user):
        """Der Formatter summiert die Tage NICHT selbst — er druckt, was in der
        Zeile steht. Eine Zeile mit Plan 8/5/4 und Skalar 30 erzeugt deshalb
        „= 30,0 h/Woche" neben Tagen, die 17 ergeben. Deshalb muss die
        Invariante an der QUELLE stehen, nicht im Formatter."""
        _mk_day_plan_change(
            db, test_user, date(2026, 3, 15), (8, 5, 4, None, None), weekly_hours=30
        )
        segs = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 3, 1), date(2026, 3, 31)
        )
        assert format_weekly_hours_history(segs) == (
            "ab 15.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 30,0 h/Woche"
        )

    def test_two_broken_rows_with_the_same_plan_do_not_merge(self, db, test_user):
        """Zweite Haelfte des Fundes: gleicher Tagesplan, verschiedene Skalare →
        die Segmente verschmelzen nicht (``weekly_hours`` ist Teil des
        Vergleichs), der Bericht behauptet also eine Planaenderung, die es nie
        gab. Mit korrektem Skalar verschmelzen sie."""
        _mk_day_plan_change(
            db, test_user, date(2026, 3, 1), (8, 5, 4, None, None), weekly_hours=30
        )
        _mk_day_plan_change(
            db, test_user, date(2026, 7, 1), (8, 5, 4, None, None), weekly_hours=20
        )
        broken = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )
        assert len(broken) == 3  # Basis + zwei Pseudo-Aenderungen

        # Invariante hergestellt (= was Migration 067 jetzt tut) → ein Segment
        # ab dem ersten Wirkungsdatum, keine Pseudo-Aenderung mehr.
        for row in db.query(WorkingHoursChange).filter(
            WorkingHoursChange.user_id == test_user.id
        ).all():
            row.weekly_hours = Decimal("17")
        db.commit()
        fixed = calculation_service.weekly_hours_segments(
            db, test_user, date(2026, 1, 1), date(2026, 12, 31)
        )
        assert len(fixed) == 2
        assert format_weekly_hours_history(fixed) == (
            "ab 01.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche"
        )


# ─────────────────────────────────────────────────────────────────────
# #431: Textdarstellung des Tagesplans
# ─────────────────────────────────────────────────────────────────────


class TestFormatDayPlanHistory:
    def test_format_names_the_weekdays(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 12, 31), 17,
                 use_daily_schedule=True, day_hours=(8, 5, 4, None, None),
                 work_days_per_week=3),
        ])
        assert text == "ab 01.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche"

    def test_weekdays_without_hours_are_left_out(self):
        """Ein Tag ohne Stunden ist kein Arbeitstag — weder NULL noch 0 gehoeren
        in die Kopfzeile."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 12, 31), 13,
                 use_daily_schedule=True, day_hours=(8, 0, 5, None, 0),
                 work_days_per_week=2),
        ])
        assert text == "ab 01.03.2026: Mo 8,0 / Mi 5,0 = 13,0 h/Woche"

    def test_quarter_hours_are_shown_exactly(self):
        """Die Spalten sind ``Numeric(4,2)``: 8,25 h ist ein realer Vertragswert.
        Auf eine Nachkommastelle gerundet stuende hier '8,2' — und der
        Frontend-Zwilling (``toFixed(1)``) schriebe '8,3'."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 12, 31), Decimal("17.75"),
                 use_daily_schedule=True, day_hours=("8.25", "5.5", "4", None, None),
                 work_days_per_week=3),
        ])
        assert text == "ab 01.03.2026: Mo 8,25 / Di 5,5 / Mi 4,0 = 17,75 h/Woche"

    def test_uniform_wording_is_untouched(self):
        """Byte-Identitaet: fuer Mitarbeitende ohne Tagesplan bleibt der Satz
        woertlich der von #415."""
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 14), 40),
            _seg(date(2026, 3, 15), date(2026, 3, 31), 30),
        ])
        assert text == "ab 15.03.2026: 30,0 Std/Woche"

    def test_mixed_history_joins_both_wordings(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 6, 30), 20),
            _seg(date(2026, 7, 1), date(2026, 12, 31), 17,
                 use_daily_schedule=True, day_hours=(8, 5, 4, None, None),
                 work_days_per_week=3),
        ])
        assert text == (
            "ab 01.03.2026: 20,0 Std/Woche; "
            "ab 01.07.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche"
        )

    def test_compact_drops_the_pointless_decimal(self):
        """PDF-Meta: Inline-Absatz in Schriftgroesse 8 auf Querformat — dort die
        Kurzform. 8,25 bleibt exakt, 8,0 wird zu 8."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 12, 31), Decimal("17.25"),
                 use_daily_schedule=True, day_hours=(8, 5, "4.25", None, None),
                 work_days_per_week=3),
        ], compact=True)
        assert text == "ab 01.03.2026: Mo 8 / Di 5 / Mi 4,25 = 17,25 h/Woche"

    def test_compact_leaves_the_uniform_wording_alone(self):
        """Der Kurzform-Schalter darf die #415-Formulierung nicht anfassen —
        sonst aendert sich die PDF-Meta jedes Mitarbeitenden ohne Tagesplan."""
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 14), 40),
            _seg(date(2026, 3, 15), date(2026, 3, 31), 30),
        ], compact=True)
        assert text == "ab 15.03.2026: 30,0 Std/Woche"

    def test_day_plan_without_any_hours_falls_back_to_the_weekly_sum(self):
        """Defensiv: eine (per Schema unmoegliche) Zeile ohne einen einzigen
        Tageswert darf keinen leeren Satz erzeugen."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40),
            _seg(date(2026, 3, 1), date(2026, 12, 31), 0,
                 use_daily_schedule=True, day_hours=(None,) * 5, work_days_per_week=3),
        ])
        assert text == "ab 01.03.2026: 0,0 Std/Woche"


# ─────────────────────────────────────────────────────────────────────
# Textdarstellung
# ─────────────────────────────────────────────────────────────────────


class TestFormatWeeklyHoursHistory:
    def test_single_segment_renders_empty(self):
        assert format_weekly_hours_history(
            [_seg(date(2026, 3, 1), date(2026, 3, 31), 40)]
        ) == ""

    def test_no_segments_render_empty(self):
        assert format_weekly_hours_history([]) == ""

    def test_one_change_is_named_with_date_and_value(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 14), 40),
            _seg(date(2026, 3, 15), date(2026, 3, 31), 30),
        ])
        assert text == "ab 15.03.2026: 30,0 Std/Woche"

    def test_multiple_changes_are_joined(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 3, 9), 40),
            _seg(date(2026, 3, 10), date(2026, 8, 31), 30),
            _seg(date(2026, 9, 1), date(2026, 12, 31), Decimal("20.5")),
        ])
        assert text == "ab 10.03.2026: 30,0 Std/Woche; ab 01.09.2026: 20,5 Std/Woche"


# ─────────────────────────────────────────────────────────────────────
# Abschluss-Review #431, Fund 1: Arbeitstage-Wechsel benennen
# ─────────────────────────────────────────────────────────────────────


class TestWorkDaysChangeIsNamed:
    """Eine Aenderung, die die Wochensumme unberuehrt laesst, aber die
    ARBEITSTAGE verschiebt, erzeugte eine Aenderungszeile, die zeichengleich zur
    Kopfzeile war — waehrend das Tagessoll derselben Tageszeilen sprang."""

    def test_work_days_change_at_equal_weekly_hours_is_named(self):
        """Der gemeldete Fall: 40 h auf 5 Tage → 40 h auf 4 Tage. Vorher stand
        hier „ab 16.03.2026: 40,0 Std/Woche" neben einer Kopfzeile „40,0" —
        zweimal dieselbe Zahl, waehrend das Tagessoll von 8,00 auf 10,00 ging."""
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 15), 40, work_days_per_week=5),
            _seg(date(2026, 3, 16), date(2026, 3, 31), 40, work_days_per_week=4),
        ])
        assert text == "ab 16.03.2026: 40,0 Std/Woche auf 4 Arbeitstage"

    def test_hours_and_work_days_change_together(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 15), 40, work_days_per_week=5),
            _seg(date(2026, 3, 16), date(2026, 3, 31), 20, work_days_per_week=4),
        ])
        assert text == "ab 16.03.2026: 20,0 Std/Woche auf 4 Arbeitstage"

    def test_single_work_day_is_singular(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 15), 8, work_days_per_week=5),
            _seg(date(2026, 3, 16), date(2026, 3, 31), 8, work_days_per_week=1),
        ])
        assert text == "ab 16.03.2026: 8,0 Std/Woche auf 1 Arbeitstag"

    def test_switching_back_from_a_day_plan_names_the_new_work_days(self):
        """Rueckschalten Tagesplan → gleichmaessig: die Tage kommen aus dem
        Snapshot beider Segmente, der Vergleich funktioniert ueber den
        Moduswechsel hinweg."""
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 15), 40, use_daily_schedule=True,
                 day_hours=(10, 10, 10, 10, None), work_days_per_week=4),
            _seg(date(2026, 3, 16), date(2026, 3, 31), 40, work_days_per_week=5),
        ])
        assert text == "ab 16.03.2026: 40,0 Std/Woche auf 5 Arbeitstage"

    def test_compact_uses_the_short_wording(self):
        """PDF-Meta (Schriftgroesse 8, Querformat) — „Tage" statt
        „Arbeitstage"."""
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 15), 40, work_days_per_week=5),
            _seg(date(2026, 3, 16), date(2026, 3, 31), 40, work_days_per_week=4),
        ], compact=True)
        assert text == "ab 16.03.2026: 40,0 Std/Woche auf 4 Tage"

    def test_compact_singular(self):
        text = format_weekly_hours_history([
            _seg(date(2026, 3, 1), date(2026, 3, 15), 8, work_days_per_week=5),
            _seg(date(2026, 3, 16), date(2026, 3, 31), 8, work_days_per_week=1),
        ], compact=True)
        assert text == "ab 16.03.2026: 8,0 Std/Woche auf 1 Tag"

    def test_unchanged_work_days_keep_the_frozen_415_wording(self):
        """BYTE-IDENTITAET: vor diesem Branch waren die Arbeitstage nicht
        historisiert — alle Segmente eines Mitarbeitenden trugen denselben Wert.
        Fuer JEDEN so darstellbaren Verlauf muss der Satz woertlich der von #415
        bleiben, in beiden Schreibweisen."""
        for work_days in range(1, 8):
            for compact in (False, True):
                text = format_weekly_hours_history([
                    _seg(date(2026, 1, 1), date(2026, 3, 9), 40, work_days_per_week=work_days),
                    _seg(date(2026, 3, 10), date(2026, 8, 31), 30, work_days_per_week=work_days),
                    _seg(date(2026, 9, 1), date(2026, 12, 31), Decimal("20.5"),
                         work_days_per_week=work_days),
                ], compact=compact)
                assert text == (
                    "ab 10.03.2026: 30,0 Std/Woche; ab 01.09.2026: 20,5 Std/Woche"
                ), f"work_days={work_days} compact={compact}"

    def test_only_the_changing_segment_names_the_days(self):
        """Der Zusatz haftet an der Aenderung, nicht am Zustand: die dritte
        Zeile behaelt dieselben vier Tage und nennt sie deshalb nicht noch
        einmal."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 3, 9), 40, work_days_per_week=5),
            _seg(date(2026, 3, 10), date(2026, 8, 31), 40, work_days_per_week=4),
            _seg(date(2026, 9, 1), date(2026, 12, 31), 30, work_days_per_week=4),
        ])
        assert text == (
            "ab 10.03.2026: 40,0 Std/Woche auf 4 Arbeitstage; "
            "ab 01.09.2026: 30,0 Std/Woche"
        )

    def test_day_plan_wording_is_untouched(self):
        """Im Tagesplan-Modus stehen die Arbeitstage bereits als benannte
        Wochentage da — dort kein Zusatz, auch wenn sich ihre Zahl aendert."""
        text = format_weekly_hours_history([
            _seg(date(2026, 1, 1), date(2026, 2, 28), 40, work_days_per_week=5),
            _seg(date(2026, 3, 1), date(2026, 12, 31), 17,
                 use_daily_schedule=True, day_hours=(8, 5, 4, None, None),
                 work_days_per_week=3),
        ])
        assert text == "ab 01.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche"

    def test_work_days_changed_helper_is_the_one_rule(self):
        a = _seg(date(2026, 1, 1), date(2026, 3, 9), 40, work_days_per_week=5)
        b = _seg(date(2026, 3, 10), date(2026, 12, 31), 40, work_days_per_week=4)
        assert calculation_service.work_days_changed(a, b) is True
        assert calculation_service.work_days_changed(b, b) is False
        # Ohne Vergleichspunkt wird nichts behauptet.
        assert calculation_service.work_days_changed(None, b) is False

    def test_end_to_end_in_the_xlsx_monthly_header(self, db, test_user):
        """Der Seam-Test: alle sechs #415-Flaechen holen den Text aus
        `format_weekly_hours_history` — eine Datei reicht als Nachweis, dass der
        Zusatz dort ankommt."""
        from app.services.export_service import generate_monthly_report
        # Basis-Zeile: der Zustand vor der Aenderung darf nicht vom (inzwischen
        # nachgezogenen) User-Rueckfallwert abhaengen — genau dafuer legt der
        # Schreibpfad sie an.
        _mk_change(db, test_user, date(2026, 1, 1), 40, work_days=5)
        _mk_change(db, test_user, date(2026, 3, 16), 40, work_days=4)
        test_user.work_days_per_week = 4  # aktueller Vertragswert nach dem Wechsel
        db.commit()

        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        assert sheet.cell(row=1, column=5).value == 40.0
        assert sheet.cell(row=1, column=6).value == (
            "ab 16.03.2026: 40,0 Std/Woche auf 4 Arbeitstage"
        )


# ─────────────────────────────────────────────────────────────────────
# XLSX
# ─────────────────────────────────────────────────────────────────────


class TestXlsxMonthlyReportShowsHistory:
    def test_header_uses_hours_valid_at_month_start_not_current(self, db, test_user):
        """Kernregression: die Kopfzeile darf nicht den AKTUELLEN Vertragswert
        zeigen, waehrend die Tageszeilen historisch rechnen.

        Aufbau so, dass sich beide Werte unterscheiden: eine Aenderung VOR dem
        Monat setzt den zum 01.03. gueltigen Wert auf 40 h, eine zweite Aenderung
        im Monat bringt ihn auf 20 h — ``user.weekly_hours`` (der aktuelle
        Vertragswert) steht auf 20. Die alte Implementierung schrieb 20 in die
        Kopfzeile, obwohl die Tageszeilen der ersten Monatshaelfte mit 40 h
        rechneten.
        """
        from app.services.export_service import generate_monthly_report
        _mk_change(db, test_user, date(2026, 1, 1), 40)
        _mk_change(db, test_user, date(2026, 3, 15), 20)
        test_user.weekly_hours = Decimal("20.0")  # aktueller Vertragswert
        db.commit()

        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        assert sheet.cell(row=1, column=5).value == 40.0
        assert sheet.cell(row=1, column=6).value == "ab 15.03.2026: 20,0 Std/Woche"

    def test_header_carries_the_change_note(self, db, test_user):
        from app.services.export_service import generate_monthly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        assert sheet.cell(row=1, column=6).value == "ab 15.03.2026: 30,0 Std/Woche"

    def test_no_change_leaves_the_note_cell_empty(self, db, test_user):
        from app.services.export_service import generate_monthly_report
        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        assert sheet.cell(row=1, column=5).value == 40.0
        assert not sheet.cell(row=1, column=6).value

    def test_header_note_carries_a_full_text_comment(self, db, test_user):
        """Fund D (Abschluss-Review #431): G1 ("Monat:") ist nicht leer — Excel
        ueberlaeuft F1 also NICHT und schneidet den sichtbaren Text ab, ohne
        das anzuzeigen. Ein Zellkommentar traegt den VOLLEN Satz unabhaengig
        von der Spaltenbreite/Nachbarzelle."""
        from app.services.export_service import generate_monthly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        cell = sheet.cell(row=1, column=6)
        assert cell.comment is not None
        assert cell.comment.text == "ab 15.03.2026: 30,0 Std/Woche"
        # G1 ("Monat:") bleibt unangetastet — der Fix verschiebt keine Spalte.
        assert sheet.cell(row=1, column=7).value == "Monat:"

    def test_header_note_comment_carries_the_long_multi_segment_sentence_in_full(self, db, test_user):
        """Der Extremfall aus dem Fund: eine Tagesplan-Zeile mit 53 Zeichen —
        in Spalte F (Breite 12) neben dem nicht-leeren G1 optisch praktisch
        unsichtbar, im Kommentar aber vollstaendig vorhanden."""
        from app.services.export_service import generate_monthly_report
        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))
        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        expected = "ab 15.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche"
        assert len(expected) > 40  # weit ueber Spalte-F-Breite 12 (:662) hinaus
        cell = sheet.cell(row=1, column=6)
        assert cell.value == expected
        assert cell.comment is not None
        assert cell.comment.text == expected


class TestXlsxYearlyReportShowsHistory:
    def test_overview_has_a_changes_column(self, db, test_user):
        from app.services.export_service import generate_yearly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        wb = _load_xlsx(generate_yearly_report(db, 2026))
        sheet = wb["Jahresübersicht"]
        headers = [sheet.cell(row=3, column=c).value for c in range(1, 13)]
        assert "Stundenänderungen" in headers
        col = headers.index("Stundenänderungen") + 1
        assert sheet.cell(row=4, column=col).value == "ab 15.03.2026: 30,0 Std/Woche"

    def test_overview_weekly_hours_is_the_year_start_value(self, db, test_user):
        """Wie beim Monatsbericht: die Spalte zeigt den zum 01.01. gueltigen
        Wert, nicht den heute im Vertrag stehenden."""
        from app.services.export_service import generate_yearly_report
        _mk_change(db, test_user, date(2025, 1, 1), 40)
        _mk_change(db, test_user, date(2026, 3, 15), 20)
        test_user.weekly_hours = Decimal("20.0")
        db.commit()
        wb = _load_xlsx(generate_yearly_report(db, 2026))
        sheet = wb["Jahresübersicht"]
        assert sheet.cell(row=4, column=2).value == 40.0
        assert sheet.cell(row=4, column=11).value == "ab 15.03.2026: 20,0 Std/Woche"

    def test_employee_sheet_names_the_change(self, db, test_user):
        from app.services.export_service import generate_yearly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        wb = _load_xlsx(generate_yearly_report(db, 2026))
        sheet = wb[test_user.last_name[:20]]
        row2 = [sheet.cell(row=2, column=c).value for c in range(1, 11)]
        assert "Wochenstunden:" in row2
        assert any(v == "ab 15.03.2026: 30,0 Std/Woche" for v in row2)


class TestSurfacesShowTheDayPlan:
    """#431: alle sechs §16-Flaechen muessen den Tagesplan ausweisen, nicht nur
    die Wochensumme."""

    EXPECTED = "ab 15.03.2026: Mo 8,0 / Di 5,0 / Mi 4,0 = 17,0 h/Woche"

    def test_xlsx_monthly_header(self, db, test_user):
        from app.services.export_service import generate_monthly_report
        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))
        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{test_user.last_name} {test_user.first_name}"[:31]]
        assert sheet.cell(row=1, column=6).value == self.EXPECTED

    def test_xlsx_monthly_header_number_cell_is_the_day_plan_sum(self, db, day_plan_user):
        """Die Kopf-ZAHLzelle (nicht nur der Aenderungstext) muss fuer einen
        Tagesplan-Mitarbeitenden die Wochensumme seines Plans zeigen — 8+5+4."""
        from app.services.export_service import generate_monthly_report
        wb = _load_xlsx(generate_monthly_report(db, 2026, 3))
        sheet = wb[f"{day_plan_user.last_name} {day_plan_user.first_name}"[:31]]
        assert sheet.cell(row=1, column=5).value == 17.0
        assert not sheet.cell(row=1, column=6).value  # keine Aenderung im Monat

    def test_xlsx_yearly_overview_number_cell_is_the_day_plan_sum(self, db, day_plan_user):
        from app.services.export_service import generate_yearly_report
        wb = _load_xlsx(generate_yearly_report(db, 2026))
        sheet = wb["Jahresübersicht"]
        names = [sheet.cell(row=r, column=1).value for r in range(4, 12)]
        row = 4 + names.index(f"{day_plan_user.last_name}, {day_plan_user.first_name}")
        assert sheet.cell(row=row, column=2).value == 17.0

    def test_xlsx_yearly_overview_column(self, db, test_user):
        from app.services.export_service import generate_yearly_report
        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))
        wb = _load_xlsx(generate_yearly_report(db, 2026))
        sheet = wb["Jahresübersicht"]
        assert sheet.cell(row=4, column=11).value == self.EXPECTED

    def test_xlsx_yearly_employee_sheet(self, db, test_user):
        from app.services.export_service import generate_yearly_report
        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))
        wb = _load_xlsx(generate_yearly_report(db, 2026))
        sheet = wb[test_user.last_name[:20]]
        row2 = [sheet.cell(row=2, column=c).value for c in range(1, 11)]
        assert self.EXPECTED in row2

    def test_ods_monthly_and_yearly(self, db, test_user):
        import zipfile
        from app.services.ods_export_service import generate_monthly_report, generate_yearly_report
        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))
        monthly = generate_monthly_report(db, 2026, 3)
        monthly.seek(0)
        assert self.EXPECTED in zipfile.ZipFile(monthly).read("content.xml").decode("utf-8")
        yearly = generate_yearly_report(db, 2026)
        yearly.seek(0)
        assert self.EXPECTED in zipfile.ZipFile(yearly).read("content.xml").decode("utf-8")

    def test_pdf_meta_uses_the_compact_form(self, db, test_user, monkeypatch):
        """Die PDF-Meta ist ein Inline-Absatz in Schriftgroesse 8 — dort steht die
        Kurzform. Geprueft am Text, den der Exporter baut (der fertige PDF-Strom
        ist komprimiert und nicht durchsuchbar)."""
        from app.services import export_service
        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))

        captured = []
        original = export_service.format_weekly_hours_history

        def _spy(segments, compact=False):
            text = original(segments, compact=compact)
            captured.append(text)
            return text

        monkeypatch.setattr(export_service, "format_weekly_hours_history", _spy)
        out = export_service.generate_monthly_report_pdf(db, 2026, 3)
        out.seek(0)
        assert out.read()[:5] == b"%PDF-"
        assert "ab 15.03.2026: Mo 8 / Di 5 / Mi 4 = 17 h/Woche" in captured

    def test_monthly_report_api_carries_the_day_plan(self, db, test_admin, test_user):
        from fastapi.testclient import TestClient
        from app.database import get_db
        from app.middleware.auth import get_current_user, require_admin
        from tests.test_endpoints import test_app

        _mk_day_plan_change(db, test_user, date(2026, 3, 15), (8, 5, 4, None, None))

        def _override_db():
            yield db
        test_app.dependency_overrides[get_db] = _override_db
        test_app.dependency_overrides[get_current_user] = lambda: test_admin
        test_app.dependency_overrides[require_admin] = lambda: test_admin
        try:
            r = TestClient(test_app).get("/api/admin/reports/monthly?month=2026-03")
            assert r.status_code == 200, r.text
            row = [x for x in r.json() if x["user_id"] == str(test_user.id)][0]
            assert row["weekly_hours_changes"] == [{
                "effective_from": "2026-03-15",
                "weekly_hours": 17.0,
                "use_daily_schedule": True,
                "day_hours": [8.0, 5.0, 4.0, None, None],
                "work_days_per_week": 3,
                # 5 → 3 Arbeitstage; im Tagesplan-Modus rendert der Zwilling den
                # Befund nicht (die Wochentage stehen dort namentlich da), das
                # Feld sagt trotzdem die Wahrheit.
                "work_days_changed": True,
            }]
        finally:
            test_app.dependency_overrides.clear()

    def test_weekly_report_api_carries_the_day_plan(self, db, test_admin, test_user):
        from fastapi.testclient import TestClient
        from app.database import get_db
        from app.middleware.auth import get_current_user, require_admin
        from tests.test_endpoints import test_app

        # KW 12/2026 = Mo 16.03. – So 22.03.; die Aenderung liegt im Fenster.
        _mk_day_plan_change(db, test_user, date(2026, 3, 18), (8, 5, 4, None, None))

        def _override_db():
            yield db
        test_app.dependency_overrides[get_db] = _override_db
        test_app.dependency_overrides[get_current_user] = lambda: test_admin
        test_app.dependency_overrides[require_admin] = lambda: test_admin
        try:
            r = TestClient(test_app).get("/api/admin/reports/weekly?week_start=2026-03-16")
            assert r.status_code == 200, r.text
            row = [x for x in r.json() if x["user_id"] == str(test_user.id)][0]
            assert row["weekly_hours_changes"] == [{
                "effective_from": "2026-03-18",
                "weekly_hours": 17.0,
                "use_daily_schedule": True,
                "day_hours": [8.0, 5.0, 4.0, None, None],
                "work_days_per_week": 3,
                "work_days_changed": True,  # 5 → 3, wie im Monatsbericht
            }]
        finally:
            test_app.dependency_overrides.clear()


class TestPdfMonthlyReportShowsHistory:
    def test_pdf_builds_with_a_change_in_the_period(self, db, test_user):
        """Der PDF-Meta-Text wird um die Aenderung ergaenzt — reportlab muss den
        laengeren Paragraph fehlerfrei bauen (unbalancierte Tags => 500)."""
        from app.services.export_service import generate_monthly_report_pdf
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        out = generate_monthly_report_pdf(db, 2026, 3)
        out.seek(0)
        assert out.read()[:5] == b"%PDF-"


# ─────────────────────────────────────────────────────────────────────
# ODS — Parität zu XLSX (CLAUDE.md: drei Exporter parieren)
# ─────────────────────────────────────────────────────────────────────


class TestOdsParity:
    def test_monthly_ods_carries_the_change_note(self, db, test_user):
        import zipfile
        from app.services.ods_export_service import generate_monthly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        out = generate_monthly_report(db, 2026, 3)
        out.seek(0)
        content = zipfile.ZipFile(out).read("content.xml").decode("utf-8")
        assert "ab 15.03.2026: 30,0 Std/Woche" in content

    def test_monthly_ods_note_carries_an_annotation(self, db, test_user):
        """Fund D — ODS-Pendant: dieselbe Ueberlauf-Zelle (die Nachbarzelle
        "Monat:" ist auch hier nicht leer, s. `_monthly_sheet`), derselbe
        ``office:annotation``-Kommentar statt einer breiteren Spalte."""
        import zipfile
        from app.services.ods_export_service import generate_monthly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        out = generate_monthly_report(db, 2026, 3)
        out.seek(0)
        content = zipfile.ZipFile(out).read("content.xml").decode("utf-8")
        assert "<office:annotation" in content
        # Der Kommentartext steht zusaetzlich zum sichtbaren Zellinhalt drin —
        # beide tragen denselben Satz.
        assert content.count("ab 15.03.2026: 30,0 Std/Woche") >= 2
        # Fix-Welle 4 #4: das ODF-Inhaltsmodell verlangt office:annotation VOR
        # den text:p-Kindern von table:table-cell — ein reines "kommt im
        # content.xml vor"-Assert (wie oben) belegt das NICHT, ein Kommentar
        # nach dem text:p wuerde denselben String liefern. Deshalb hier
        # explizit die Reihenfolge INNERHALB der annotierten Zelle pruefen.
        ann_idx = content.index("<office:annotation")
        cell_open = content.rindex("<table:table-cell", 0, ann_idx)
        cell_close = content.index("</table:table-cell>", ann_idx)
        cell_xml = content[cell_open:cell_close]
        assert cell_xml.index("<office:annotation") < cell_xml.index("<text:p"), (
            "office:annotation muss laut ODF-Inhaltsmodell VOR text:p in der "
            "Zelle stehen (LibreOffice ist beim Import tolerant, andere "
            "ODF-Consumer nicht)"
        )

    def test_yearly_ods_overview_has_the_changes_column(self, db, test_user):
        import zipfile
        from app.services.ods_export_service import generate_yearly_report
        _mk_change(db, test_user, date(2026, 3, 15), 30)
        out = generate_yearly_report(db, 2026)
        out.seek(0)
        content = zipfile.ZipFile(out).read("content.xml").decode("utf-8")
        assert "Stundenänderungen" in content
        assert "ab 15.03.2026: 30,0 Std/Woche" in content


# ─────────────────────────────────────────────────────────────────────
# API — der Monatsbericht transportiert die Änderungen in die UI
# ─────────────────────────────────────────────────────────────────────


class TestMonthlyReportApi:
    def test_report_carries_weekly_hours_changes(self, db, test_admin, test_user):
        from fastapi.testclient import TestClient
        from app.database import get_db
        from app.middleware.auth import get_current_user, require_admin
        from tests.test_endpoints import test_app

        _mk_change(db, test_user, date(2026, 3, 15), 30)

        def _override_db():
            yield db
        test_app.dependency_overrides[get_db] = _override_db
        test_app.dependency_overrides[get_current_user] = lambda: test_admin
        test_app.dependency_overrides[require_admin] = lambda: test_admin
        try:
            r = TestClient(test_app).get("/api/admin/reports/monthly?month=2026-03")
            assert r.status_code == 200, r.text
            row = [x for x in r.json() if x["user_id"] == str(test_user.id)][0]
            assert row["weekly_hours"] == 40.0
            assert row["weekly_hours_changes"] == [{
                "effective_from": "2026-03-15",
                "weekly_hours": 30.0,
                # #431: der Snapshot ist vollstaendig, im gleichmaessigen Modus
                # sind die Tageswerte durchgehend leer.
                "use_daily_schedule": False,
                "day_hours": [None] * 5,
                "work_days_per_week": 5,
                # Abschluss-Review #431, Fund 1: reine Stundenaenderung, die
                # Arbeitstage bleiben — der Bildschirm schreibt denselben Satz
                # wie vor diesem Branch.
                "work_days_changed": False,
            }]
        finally:
            test_app.dependency_overrides.clear()

    def test_report_flags_a_work_days_change_for_the_screen(self, db, test_admin, test_user):
        """Der Bildschirm sieht nur die Aenderungen (`segments[1:]`), nie das
        Basissegment davor — der Befund „Arbeitstage geaendert" muss deshalb
        mitgeliefert werden, sonst kann der Zwilling den Zusatz „auf 4
        Arbeitstage" fuer die erste Aenderung eines Zeitraums nicht schreiben und
        Bildschirm und Datei sagen Verschiedenes."""
        from fastapi.testclient import TestClient
        from app.database import get_db
        from app.middleware.auth import get_current_user, require_admin
        from tests.test_endpoints import test_app

        _mk_change(db, test_user, date(2026, 1, 1), 40, work_days=5)
        _mk_change(db, test_user, date(2026, 3, 16), 40, work_days=4)
        test_user.work_days_per_week = 4
        db.commit()

        def _override_db():
            yield db
        test_app.dependency_overrides[get_db] = _override_db
        test_app.dependency_overrides[get_current_user] = lambda: test_admin
        test_app.dependency_overrides[require_admin] = lambda: test_admin
        try:
            r = TestClient(test_app).get("/api/admin/reports/monthly?month=2026-03")
            assert r.status_code == 200, r.text
            row = [x for x in r.json() if x["user_id"] == str(test_user.id)][0]
            assert row["weekly_hours_changes"] == [{
                "effective_from": "2026-03-16",
                "weekly_hours": 40.0,
                "use_daily_schedule": False,
                "day_hours": [None] * 5,
                "work_days_per_week": 4,
                "work_days_changed": True,
            }]
        finally:
            test_app.dependency_overrides.clear()

    def test_report_without_changes_has_an_empty_list(self, db, test_admin, test_user):
        from fastapi.testclient import TestClient
        from app.database import get_db
        from app.middleware.auth import get_current_user, require_admin
        from tests.test_endpoints import test_app

        def _override_db():
            yield db
        test_app.dependency_overrides[get_db] = _override_db
        test_app.dependency_overrides[get_current_user] = lambda: test_admin
        test_app.dependency_overrides[require_admin] = lambda: test_admin
        try:
            r = TestClient(test_app).get("/api/admin/reports/monthly?month=2026-03")
            row = [x for x in r.json() if x["user_id"] == str(test_user.id)][0]
            assert row["weekly_hours_changes"] == []
        finally:
            test_app.dependency_overrides.clear()
